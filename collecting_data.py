import datetime as dt
import logging
from typing import List, Dict, Any, Tuple
import pandas as pd
import yfinance as yf
import ta
import numpy as np
from pypfopt import expected_returns
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas_datareader import data as web
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed

BASE_FORECAST_URL = "https://tradingeconomics.com/forecast"
FORECAST_SPECS: Dict[str, Dict[str, Any]] = {
    # endpoint: {"index_pos": int, "mapping": {col_idx: col_name, ...}}
    "interest-rate": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "stock-market": {"index_pos": 1, "mapping": {1: "Index", 2: "Last", 4: "Q2/25", 5: "Q3/25", 6: "Q4/25", 7: "Q1/26"}},
    "currency": {"index_pos": 1, "mapping": {1: "Currency Pair", 2: "Last", 4: "Q2/25", 5: "Q3/25", 6: "Q4/25", 7: "Q1/26"}},
    "wages": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "unemployment-rate": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "gdp": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "2025", 4: "2026", 5: "2027"}},
    "consumer-price-index-cpi": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "inflation-rate": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "consumer-confidence": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "business-confidence": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "balance-of-trade": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
    "corporate-profits": {"index_pos": 0, "mapping": {0: "Country", 1: "Last", 3: "Q2/25", 4: "Q3/25", 5: "Q4/25", 6: "Q1/26"}},
}
COMMODITY_CATEGORIES: Dict[str, Tuple[str, int, Dict[int, str]]] = {
    "Crude Oil": (
        "Energy", 
        0,
        {
            0: "Category",  
            1: "Price",      
            2: "Signal",
            3: "Q2/25",
            4: "Q3/25",
            5: "Q4/25",
            6: "Q1/26",
        }
    ),
    "Gold": (
        "Metals",
        0,
        {
            0: "Category",   
            1: "Signal",
            2: "Q2/25",
            3: "Q3/25",
            4: "Q4/25",
            5: "Q1/26",
        }
    ),
}
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

def configure_logging(level: int = logging.INFO) -> None:
    """
    Configure logging format and level.
    """
    fmt = '%(asctime)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=level, format=fmt)

def download_data(tickers: List[str], start: dt.date, end: dt.date) -> pd.DataFrame:
    """
    Download historical OHLCV data from Yahoo Finance for given tickers.
    Raises ValueError if the result is empty.
    """
    logging.info("Downloading data for %d tickers: %s to %s", len(tickers), start, end)
    data = yf.download(tickers, start=start, end=end)
    if data.empty:
        raise ValueError("No data downloaded. Verify tickers or date range.")
    data.index = pd.to_datetime(data.index)
    return data


def scrape_forecast(endpoint: str, index_pos: int, col_map: Dict[int, str]) -> pd.DataFrame:
    """
    Scrape a Trading Economics forecast table by endpoint.
    Uses col_map to map td indices to column names, and index_pos for the index column.
    """
    url = f"{BASE_FORECAST_URL}/{endpoint}"
    logging.info("Scraping forecast for %s", endpoint)
    resp = requests.get(url, headers=HEADERS, timeout=10)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, 'html.parser')
    table = soup.find('table')
    if not table:
        raise RuntimeError(f"No table on page {url}")
    records: List[Dict[str, Any]] = []
    for row in table.find_all('tr')[1:]:
        cells = row.find_all('td')
        if len(cells) <= max(col_map):
            continue
        rec: Dict[str, Any] = {}
        for idx, name in col_map.items():
            rec[name] = cells[idx].get_text(strip=True)
        records.append(rec)
    df = pd.DataFrame(records)
    df.set_index(col_map[index_pos], inplace=True)
    return df


from bs4 import BeautifulSoup
import requests
import pandas as pd

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0.0.0 Safari/537.36"
    )
}

def scrape_commodity_forecast(
    category: str, 
    index_pos: int, 
    col_map: Dict[int, str]
) -> pd.DataFrame:
    url = f"{BASE_FORECAST_URL}/commodity"
    resp = requests.get(url, headers=HEADERS, timeout=10)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    for tbl in soup.find_all("table"):
        thead = tbl.find("thead")
        if not thead:
            continue
        first_th = thead.find("th")
        if not first_th or first_th.get_text(strip=True) != category:
            continue

        # parse the body
        records = []
        for row in tbl.find("tbody").find_all("tr"):
            cells = row.find_all("td")
            rec = {
                col_map[i]: cells[i].get_text(strip=True)
                for i in col_map
                if i < len(cells)
            }
            records.append(rec)

        df = pd.DataFrame(records)
        # set index from the correct column
        df.set_index(col_map[index_pos], inplace=True)
        return df

    raise RuntimeError(f"No '{category}' table found on commodity page.")

def get_close_series(data: pd.DataFrame, tickers: List[str]) -> pd.DataFrame:
    """
    Extract and scale 'Close' prices; interpolate missing values.
    Returns DataFrame of closes.
    """
    try:
        close = data['Close'].copy()
    except KeyError:
        logging.exception("'Close' missing in data.")
        raise
    uk = [t for t in close.columns if t.endswith('.L')]
    if uk:
        close.loc[:, uk] = close.loc[:, uk].div(100)
    avail = [t for t in tickers if t in close.columns]
    miss = set(tickers) - set(avail)
    if miss:
        logging.warning("Missing tickers: %s", miss)
    return close[avail].interpolate()


def compute_technical_indicators(
    data: pd.DataFrame,
    tickers: List[str],
    rsi_window: int = 14
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Compute RSI and MACD (+ signal) for each ticker in 'Close'.
    Returns: (RSI_df, MACD_df, MACD_signal_df) with MultiIndex columns.
    """
    close = get_close_series(data, tickers)
    rsi_d, macd_d, macd_sig_d = {}, {}, {}
    for t in tickers:
        if t not in close:
            logging.warning("Skipping indicators for %s", t)
            continue
        s = close[t]
        rsi_d[t] = ta.momentum.RSIIndicator(s, window=rsi_window).rsi()
        mac = ta.trend.MACD(s)
        macd_d[t] = mac.macd()
        macd_sig_d[t] = mac.macd_signal()
    rsi_df = pd.DataFrame(rsi_d)
    macd_df = pd.DataFrame(macd_d)
    macd_sig_df = pd.DataFrame(macd_sig_d)
    rsi_df.columns = pd.MultiIndex.from_product([['RSI'], rsi_df.columns])
    macd_df.columns = pd.MultiIndex.from_product([['MACD'], macd_df.columns])
    macd_sig_df.columns = pd.MultiIndex.from_product([['MACD Signal'], macd_sig_df.columns])
    return rsi_df, macd_df, macd_sig_df

def fetch_fred_series(symbol: str, start: str, end: str, freq: str = 'M') -> pd.Series:
    """
    Fetch a FRED series and resample forward-fill to freq.
    """
    s = web.DataReader(symbol, 'fred', start, end).squeeze()
    if freq:
        s = s.resample(freq).ffill()
    return s


def macro_data() -> pd.DataFrame:
    """
    Compile monthly macro indicators and SP500 history into one clean DataFrame.
    """
    today = dt.date.today()
    start, end = '2010-01-01', today.isoformat()
    # SP500
    sp = yf.download('^GSPC', start=start, end=end)
    
    sp500_monthly = pd.DataFrame(sp["Close"].resample("ME").ffill())

    sp500_monthly.columns = ["SP500_Close"]

    sp500_monthly["SP500_Return"] = sp500_monthly["SP500_Close"].pct_change() * 100

    sp500_monthly.dropna(inplace=True)

    # FRED series
    series_map = [
        ('Inflation', 'CPIAUCSL'),
        ('US_GDP', 'GDP'),
        ('GBP_USD', 'DEXUSUK'),
        ('US_Consumer_Sentiment', 'UMCSENT'),
        ('US_Unemployment_Rate', 'UNRATE'),
        ('US_Interest_Rate', 'DGS10'),
        ('US_Wages', 'AWHNONAG')
    ]
    results: Dict[str, pd.Series] = {}
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fetch_fred_series, sym, start, end): name
                   for name, sym in series_map}
        for fut in as_completed(futures):
            name = futures[fut]
            try:
                results[name] = fut.result()
            except Exception:
                logging.exception("Failed FRED %s", name)
    df = pd.DataFrame(results)
    macro_df = df.join(sp500_monthly, how='inner')
    return macro_df

def format_and_save_excel(
    filename: str,
    sheets: Dict[str, pd.DataFrame],
    apply_table: bool = False
) -> None:
    logging.info("Creating Excel %s", filename)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.sort_index().to_excel(writer, sheet_name=name)

        if apply_table and 'Exponential Returns' in writer.sheets:
            wb = writer.book
            ws = wb['Exponential Returns']

            max_row, max_col = ws.max_row, ws.max_column
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tbl = Table(displayName="ExpRetTbl", ref=table_ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            ws.add_table(tbl)

            # proper fills
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # apply conditional formatting to columns B & C
            for col_letter in ['B', 'C']:
                rng = f"{col_letter}2:{col_letter}{max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator='lessThan',
                        formula=['0'],
                        fill=red_fill,
                        stopIfTrue=True
                    )
                )
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(
                        operator='greaterThan',
                        formula=['0'],
                        fill=green_fill,
                        stopIfTrue=True
                    )
                )
    logging.info("Saved Excel %s", filename)

# -----------------------------------------------------------------------------
# Main Workflow
# -----------------------------------------------------------------------------
def main() -> None:
    """
    Orchestrate data download, indicator computation, forecast scraping,
    macro data retrieval, and single-pass Excel export.
    """
    configure_logging()
    today = dt.date.today()

    # 1. Download market data
    tickers = [
        ""
    ]    
    start_date = '2000-01-01'  
    data = download_data(tickers, start_date, today)
    close = get_close_series(data, tickers)
    high = data['High']
    low = data['Low']
    open_p = data['Open']
    volume = data['Volume']
    weekly_close = close.resample('W').last()
    daily_ret = close.pct_change().dropna()
    weekly_ret = weekly_close.pct_change().dropna()
    rsi_df, macd_df, macd_sig_df = compute_technical_indicators(data, tickers)
    ema_vol = daily_ret.ewm(span=252, adjust=False).std()
    weekly_ema_vol = weekly_ret.ewm(span=52, adjust=False).std()
    forecasts: Dict[str, pd.DataFrame] = {}
    for ep, spec in FORECAST_SPECS.items():
        try:
            df = scrape_forecast(ep, spec['index_pos'], spec['mapping'])
            forecasts[ep.replace('-', '_').title()] = df
        except Exception:
            logging.exception("Forecast failed: %s", ep)

    for key, (cat, idx, cmap) in COMMODITY_CATEGORIES.items():
        try:
            df = scrape_commodity_forecast(cat, idx, cmap)
            forecasts[key.replace(' ', '_')] = df
        except Exception:
            logging.exception("Commodity forecast failed: %s", key)
    macro_df = macro_data()
    sheets_data = {
        "Close": close,
        "High": high,
        "Low": low,
        "Open": open_p,
        "Volume": volume,
        "Weekly Close": weekly_close,
        "Historic Returns": daily_ret,
        "Historic Weekly Returns": weekly_ret,
        "RSI": rsi_df,
        "MACD": macd_df,
        "MACD Signal": macd_sig_df,
        **forecasts,
        "Macro Data": macro_df
    }

    one_year = today - dt.timedelta(days=365)
    close_1y = close.loc[close.index >= pd.to_datetime(one_year)]
    weekly_close_1y = weekly_close.loc[weekly_close.index >= pd.to_datetime(one_year)]
    l = len(close_1y)
    ret_1y = close_1y.pct_change().dropna()
    exp_ret_ema = expected_returns.ema_historical_return(
        close, span=l, frequency=l, compounding=True, returns_data=False
    )
    exp_ret_week = expected_returns.ema_historical_return(
        weekly_close, span=(0.5*len(weekly_close_1y)), frequency=(0.5*len(weekly_close_1y)), compounding=True
    )
    latest_ema_vol = ema_vol.iloc[-1]
    latest_wk_vol = weekly_ema_vol.iloc[-1]
    one_year_raw = close_1y.iloc[-1] / close_1y.iloc[0] - 1

    er_ema_df = pd.DataFrame({
        "EMA Returns": exp_ret_ema,
        "EMA Weekly Returns": exp_ret_week,
        "EMA Daily Volatility": latest_ema_vol,
        "EMA Weekly Volatility": latest_wk_vol
    })
    dr_df = pd.DataFrame({
        "Returns": one_year_raw,
        "Vol": ret_1y.std(),
        "Annual Vol": ret_1y.std() * np.sqrt(len(ret_1y))
    })

    file_raw = f"Portfolio_Optimisation_Data_{today}.xlsx"
    forecast_file = f"Portfolio_Optimisation_Forecast_{today}.xlsx"

    format_and_save_excel(file_raw, sheets_data)
    format_and_save_excel(
        forecast_file,
        {"Exponential Returns": er_ema_df, "Daily Returns": dr_df},
        apply_table=True
    )


if __name__ == '__main__':
    main()
  
