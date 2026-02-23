"""
End-to-end portfolio optimisation and reporting pipeline.
 
This script loads forecasts, returns, and classifications; builds risk models;  
runs a suite of long-only optimisers via `PortfolioOptimiser`; and exports both 
weights and diagnostics to an Excel workbook with conditional formatting.

Pipeline outline
----------------  
1) **Data ingestion and alignment** 
  
   - Loads recent daily/weekly/monthly returns, score/return forecasts, per-ticker 
     sector/industry and market caps, and factor/benchmark series.
   
   - Builds annual and weekly covariance estimates using a shrinkage routine that 
     can blend idiosyncratic and factor components.
  
   - Resamples, aligns, and prunes histories to a common ticker universe. 

2) **Risk model construction**

   - `ann_cov` (annualised covariance) and `weekly_cov = ann_cov / 52`.
 
   - Optional Black–Litterman (BL) prior covariance (`sigma_prior`) and view 
     uncertainty from forecast SE to derive a BL posterior inside the optimiser.

3) **Bounds and gating**
 
   - Lower/upper per-ticker bounds computed by `compute_mcap_bounds` from market cap,
     forecasted return and score, ex-ante volatility and a factor-tilt predictor.
 
   - Enforced long-only, budget, industry and sector caps via the optimiser.

4) **Optimisers executed**
  
   - **MSR**: Maximum Sharpe ratio.
  
   - **Sortino**: Maximum Sortino ratio (downside risk, 1-year weekly).
  
   - **MIR**: Maximum Information Ratio (active mean over tracking error).
  
   - **MSP**: Score-over-CVaR (Rockafellar–Uryasev 1-year CVaR).
  
   - **BL**: MSR under Black–Litterman posterior (μ_bl, Σ_bl).
  
   - **Deflated MSR**: Maximises Deflated Sharpe (controls selection bias).
  
   - **Adjusted MSR**: Maximises Adjusted Sharpe (skew/kurtosis correction).
  
   - **Composite CCP portfolios** (all long-only, same constraint set):
  
       • `comb_port`   = Sharpe + Sortino + BL-Sharpe − proximity to MIR/MSP.  
  
            Advantage: stability from anchoring to two practical baselines.
  
       • `comb_port1`  = Sharpe + Sortino + BL-Sharpe + IR(1y) + Score/CVaR.  
  
            Advantage: balanced near-term alpha, risk and tails; RU CVaR exact at end.
  
       • `comb_port2`  = Sharpe + Sortino + BL-Sharpe − L1 proximity to MIR/MSP.  
  
            Advantage: sparser tilts, turnover-friendly.
  
       • `comb_port3`  = Sharpe + Sortino + IR(5y) + Score/CVaR − L2 to BL.  
            
            Advantage: long-horizon discipline around BL equilibrium.
       
       • `comb_port4`  = As in 3 with L1 to BL + **sector risk caps** (linearised).  
           
            Advantage: limits sector risk concentration with sparse BL tilts.
       
       • `comb_port5`  = Sharpe + Sortino + BL-Sharpe + IR(1y) + Score/CVaR + ASR.  
            
            Advantage: higher-moment quality via adjusted Sharpe.
       
       • `comb_port6`  = Sharpe + Sortino + IR(5y) + Score/CVaR + ASR − L2 to BL.  
            
            Advantage: long-horizon + tails + higher moments, regularised to BL.
       
       • `comb_port7`  = As in 6 with L1 to BL + **sector risk caps** (linearised).  
            
            Advantage: sector budgets + sparse BL tilts + long-horizon risk.

       • `comb_port8`  = Sharpe + Sortino + BL-Sharpe + Score/CVaR + Adjusted Sharpe 
            
            Advantage: pure reward composite that balances mean–variance, downside/tails,
            and higher moments; BL posterior alignment **without** proximity penalties; 
            auto-scaling equalises gradient pushes across terms.

       • `comb_port12` = Sortino + MDP + BL-Sharpe + Deflated Sharpe + Adjusted Sharpe
            + Up/Down + UPM-LPM.
            Advantage: broad multi-objective reward composite without direct Sharpe term.


5) **Reporting**
  
   - Exports per-portfolio weights, industry/sector breakdowns (% of capital), 
     per-portfolio and per-ticker performance diagnostics, covariance summary, 
     and gating bounds to named Excel sheets.
  
   - Applies conditional formatting (tables, colour scales, “traffic-light” rules).

Inputs & configuration
----------------------
- Paths, ticker list, benchmark, dates, sector caps, risk-free rates, gamma weights,
  and Excel sheet names are drawn from `config`.

- Market/factor/benchmark series are fetched through `RatioData` and `yfinance`.

Outputs
-------
- Excel workbook updated in place with the following sheets (replaced if present):
  `Ticker Performance`, `Today_Buy_Sell`, `Covariance`, `Covariance Description`,
  `Bounds`, `Weights`, `Industry Breakdown`, `Sector Breakdown`,
  `Portfolio Performance`.

Notes
-----
- `yfinance` download calls require network connectivity and may incur API delays.
- All optimisers enforce the same long-only, budget, and box/sector/industry caps;
  some composite variants also include CCP linearised sector risk-contribution caps.
"""


import datetime as dt
import logging
import re
from typing import Tuple, List, Dict, Callable
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule 

import portfolio_functions as pf
from functions.cov_functions import shrinkage_covariance
from portfolio_optimisers import PortfolioOptimiser
from data_processing.ratio_data import RatioData
from financial_forecast_data5 import FinancialForecastData
import config


def build_fx_factor_returns(
    ratio_data,
    tickers,
    weekly_index,
    base_ccy: str,
):
    """
    Construct weekly foreign-exchange factor returns relative to a base currency.

    Purpose
    -------
    Portfolio and factor risk are evaluated in a single base currency. Where
    holdings trade in mixed local currencies, this function creates explicit FX
    return factors so covariance estimation can model translation risk directly.

    Construction
    ------------
    1) Determine each ticker's local trading currency.
   
    2) Obtain converter price series for each local currency into ``base_ccy``.
   
    3) For each non-base currency ``c``, compute:
   
       ``r_fx_c,t = P_fx_c,t / P_fx_c,t-1 - 1``.
   
    4) Align all series to ``weekly_index`` and return columns named ``FX_<ccy>``.

    Advantages of this modelling approach
    -------------------------------------
    - Improves covariance decomposition by separating equity and currency shocks.
   
    - Reduces the risk of conflating FX translation effects with idiosyncratic noise.
   
    - Preserves a coherent base-currency framework across global universes.

    Parameters
    ----------
    ratio_data : object
        Data adaptor exposing ticker-currency and FX converter retrieval helpers.
    tickers : sequence[str]
        Universe tickers.
    weekly_index : pd.Index
        Weekly timestamps used for alignment.
    base_ccy : str
        Base reporting currency.

    Returns
    -------
    pd.DataFrame | None
        FX factor return matrix aligned to ``weekly_index``, or ``None`` if data
        are unavailable or no non-base currencies are present.
    """
    
    try:
    
        col_ccy = ratio_data._ticker_ccy(pd.Index(tickers)).fillna(base_ccy)
    
    except Exception:
    
        return None

    try:
    
        fx_by_ccy = ratio_data._fx_converters_by_ccy(
            ccys = col_ccy,
            base_ccy = base_ccy,
            interval = "1wk",
            index = weekly_index,
        )
    
    except Exception:
    
        return None

    fx_rets = {}

    for ccy, fx_price in fx_by_ccy.items():
    
        if ccy == base_ccy:
    
            continue
    
        s = fx_price.pct_change().reindex(weekly_index).fillna(0.0)
    
        fx_rets[f"FX_{ccy}"] = s

    if not fx_rets:
    
        return None

    return pd.DataFrame(fx_rets, index = weekly_index)


r = RatioData()

tickers = config.tickers

money_in_portfolio = config.MONEY_IN_PORTFOLIO

MIN_STD = 1e-2

MAX_STD = 2


logging.basicConfig(
    level = logging.INFO,
    format = "%(asctime)s [%(levelname)s] %(message)s"
)


def ensure_headers_are_strings(
    df: pd.DataFrame
) -> pd.DataFrame:
    """
    Ensure all DataFrame headers are strings.

    Coerces each column name to `str` (empty string if None) and ensures the index
    has a non-None string name.

    Args:
        df: Input DataFrame.

    Returns:
        The same DataFrame with stringified column names and a string index name.
    """

    df.columns = [str(col) if col is not None else "" for col in df.columns]
    
    if df.index.name is None:
   
        df.index.name = "Index"
   
    else:
   
        df.index.name = str(df.index.name)
    
    return df


def _load_cached_cov_matrix(
    *,
    base_dir,
    tickers: list[str],
) -> tuple[pd.DataFrame | None, str | None]:
    """
    Try to load a cached covariance matrix (prefer today's file, else latest).
    Returns (cov_df, path_str).
    """
   
    tickers = [str(t).upper() for t in tickers]
   
    base_dir = config.BASE_DIR if base_dir is None else base_dir
   
    today_path = base_dir / f"cov_matrix_{config.TODAY}.pkl"

    candidates = []
   
    if today_path.exists():
   
        candidates.append(today_path)

    pattern = re.compile(r"cov_matrix_(\d{4}-\d{2}-\d{2})\.pkl$")
   
    dated = []
   
    for p in base_dir.glob("cov_matrix_*.pkl"):
   
        if p == today_path:
   
            continue
   
        m = pattern.search(p.name)
   
        if m:
   
            try:
   
                d = dt.date.fromisoformat(m.group(1))
   
                dated.append((d, p))
   
            except Exception:
   
                continue
   
    dated.sort(key=lambda x: x[0], reverse=True)
   
    for _, p in dated:
   
        candidates.append(p)

    for path in candidates:
   
        try:
   
            cov_loaded = pd.read_pickle(path)
   
        except Exception:
   
            continue
   
        if not isinstance(cov_loaded, pd.DataFrame):
   
            continue
   
        cov_loaded.index = cov_loaded.index.astype(str).str.upper()
   
        cov_loaded.columns = cov_loaded.columns.astype(str).str.upper()
   
        cov_loaded = cov_loaded.reindex(index = tickers, columns = tickers)
   
        if cov_loaded.shape[0] != len(tickers) or cov_loaded.shape[1] != len(tickers):
   
            continue
   
        diag = np.diag(cov_loaded.values)
   
        if not np.isfinite(diag).all():
   
            continue
   
        return cov_loaded, str(path)

    return None, None


def _load_cached_sigma_prior(
    *,
    base_dir,
    tickers: list[str],
) -> tuple[pd.DataFrame | None, str | None]:
    """
    Try to load a cached BL prior covariance (prefer today's file, else latest).
    Returns (sigma_prior_df, path_str).
    """
  
    tickers = [str(t).upper() for t in tickers]
  
    base_dir = config.BASE_DIR if base_dir is None else base_dir
  
    today_path = base_dir / f"sigma_prior_{config.TODAY}.pkl"

    candidates = []
  
    if today_path.exists():
  
        candidates.append(today_path)

    pattern = re.compile(r"sigma_prior_(\d{4}-\d{2}-\d{2})\.pkl$")
  
    dated = []
  
    for p in base_dir.glob("sigma_prior_*.pkl"):
  
        if p == today_path:
  
            continue
  
        m = pattern.search(p.name)
  
        if m:
  
            try:
  
                d = dt.date.fromisoformat(m.group(1))
  
                dated.append((d, p))
  
            except Exception:
  
                continue
  
    dated.sort(key = lambda x: x[0], reverse = True)
  
    for _, p in dated:
  
        candidates.append(p)

    for path in candidates:
  
        try:
  
            sigma_loaded = pd.read_pickle(path)
  
        except Exception:
  
            continue
  
        if not isinstance(sigma_loaded, pd.DataFrame):
  
            continue
  
        sigma_loaded.index = sigma_loaded.index.astype(str).str.upper()
  
        sigma_loaded.columns = sigma_loaded.columns.astype(str).str.upper()
  
        sigma_loaded = sigma_loaded.reindex(index = tickers, columns = tickers)
  
        if sigma_loaded.shape[0] != len(tickers) or sigma_loaded.shape[1] != len(tickers):
  
            continue
  
        diag = np.diag(sigma_loaded.values)
  
        if not np.isfinite(diag).all():
  
            continue
  
        return sigma_loaded, str(path)

    return None, None


def load_excel_data() -> Tuple[
    pd.DataFrame,
    List[str],
    np.ndarray, 
    pd.DataFrame,  
    pd.DataFrame, 
    pd.Series,    
    pd.Series,
    pd.Series,   
    pd.Series,     
    pd.Series,      
    pd.Series,
    pd.Series,
    pd.Series, 
    pd.Series
]:
    """
    Load and align all data required by the optimiser.

    This function pulls returns and forecast sheets, aligns indices to the configured
    ticker universe, computes annualised covariance with shrinkage, and prepares
    auxiliary inputs (scores, sectors, industries, market caps, factor predictor).

    Data sources:
        - Weekly/daily returns and factor series from `RatioData`.
        - Excel sheets:
            • PORTFOLIO_FILE / "Combination Forecast": forecasted Returns, SE, Score,
            Low/High Returns.
            • FORECAST_FILE / "Analyst Data": Industry, Sector, marketCap, beta.
            • DATA_FILE / "Signal Scores": daily signal snapshot (last row).

    Risk model:
        - `annCov`: annual covariance (possibly factor-shrunk); `weeklyCov = annCov/52`.
        - `sigma_prior`: prior covariance for BL (idiosyncratic-only variant).

    Returns:
        A tuple:
            weekly_ret_5y:   pd.DataFrame of weekly returns (5y window).
            tickers:         List[str] current universe from config.
            weeklyCov:       np.ndarray, annual covariance / 52.
            annCov:          np.ndarray, annual covariance (aligned).
            annCov_desc_df:  pd.DataFrame, shrinkage description/attribution.
            cov_df:          pd.DataFrame, annual covariance with labels.
            Signal:          pd.Series, latest signal score per ticker.
            beta:            pd.Series, analyst beta per ticker.
            score:           pd.Series, model score per ticker.
            comb_rets:       pd.Series, forecast expected returns per ticker.
            ticker_ind:      pd.Series, industry per ticker.
            ticker_sec:      pd.Series, sector per ticker.
            comb_std:        pd.Series, forecast standard error per ticker.
            bear_rets:       pd.Series, stressed (low) forecast returns.
            bull_rets:       pd.Series, optimistic (high) forecast returns.
            ticker_mcap:     pd.Series, market capitalisation.
            sigma_prior:     np.ndarray, prior covariance for BL (aligned).
            factor_pred:     pd.Series, factor-tilt predictor used in bounds.

    Raises:
        AssertionError: if key indices fail to align to `tickers`.
        FileNotFoundError / ValueError: if required sheets are missing.
    """

    weekly_ret = r.weekly_rets
   
    daily_ret = r.daily_rets
    
    daily_open = r.open
    
    daily_close = r.close

    macro_weekly = r.macro_data
  
    monthly_ret = daily_close.resample('ME').last().pct_change().dropna()
    
    weekly_ret.index = pd.to_datetime(weekly_ret.index)
  
    weekly_ret.sort_index(ascending = True, inplace = True)
      
    weekly_ret.columns = tickers
    
    comb_data = pd.read_excel(config.PORTFOLIO_FILE, sheet_name = "Combination Forecast", index_col = 0)
    
    comb_data.index = comb_data.index.str.upper()
    
    comb_data = comb_data.reindex(tickers)
    
    factor_data = pd.read_excel(config.FORECAST_FILE, sheet_name = "Factor Exponential Regression", index_col = 0)
    
    factor_pred = factor_data['Returns']
    
    factor_r2 = factor_data['r2']
    
    comb_std = comb_data['SE']
    
    score = comb_data["Score"]
    
    comb_rets = comb_data["Returns"]
    
    safe_rets = comb_data["Safe Returns"]
  
    bear_rets = comb_data["Low Returns"]
  
    bull_rets = comb_data["High Returns"]

    cov_path = config.BASE_DIR / f"cov_matrix_{config.TODAY}.pkl"
   
    sigma_prior_path = config.BASE_DIR / f"sigma_prior_{config.TODAY}.pkl"
   
    annCov = None
   
    annCov_desc = None
   
    extras = None
   
    annCov, cov_used_path = _load_cached_cov_matrix(
        base_dir = config.BASE_DIR,
        tickers = list(tickers),
    )
   
    print('annCov:', annCov)
   
    if annCov is not None:
   
        annCov_desc = annCov.describe()
   
        eigvals = np.linalg.eigvalsh(annCov.values)
   
        diag_vol = np.sqrt(np.clip(np.diag(annCov.values), 0, None))
   
        corr_vol = np.corrcoef(diag_vol, comb_std.reindex(tickers).to_numpy())[0, 1]
   
        extras = {
            "weights": {},
            "eig_min": float(eigvals[0]),
            "eig_max": float(eigvals[-1]),
            "eig_trace": float(eigvals.sum()),
            "vol_forecast_corr": float(corr_vol),
        }
   
        logging.info("Loaded cached covariance from %s", cov_used_path)
   
    sigma_prior, sigma_prior_used_path = _load_cached_sigma_prior(
        base_dir = config.BASE_DIR,
        tickers = list(tickers),
    )
    
    ticker_data = pd.read_excel(config.FORECAST_FILE, sheet_name = "Analyst Data", index_col = 0)
    
    ticker_ind = ticker_data["Industry"]
  
    ticker_mcap = ticker_data['marketCap']
  
    ticker_sec = ticker_data['Sector']

    sector_map = ticker_sec

    industry_map = ticker_ind
    
    d_to_e = ticker_data['debtToEquity'] / 100
    
    tax = ticker_data['Tax Rate']
    
    weekly_ret_5y = weekly_ret.loc[weekly_ret.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
   
    base_ccy = getattr(config, "BASE_CCY", "USD")
   
    weekly_ret_5y_base = r.convert_returns_to_base(
        ret_df = weekly_ret_5y,
        base_ccy = base_ccy,
        interval = "1wk"
    )
    
    if annCov is None or sigma_prior is None:
        
        daily_ret_5y = daily_ret.loc[daily_ret.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
        
        daily_close_5y = daily_close.loc[daily_close.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
        
        daily_open_5y = daily_open.loc[daily_open.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
        
        macro_weekly_5y = macro_weekly.loc[macro_weekly.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
    
        monthly_ret_5y = monthly_ret.loc[monthly_ret.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]
        
        factor_weekly, index_weekly, ind_weekly, sec_weekly = r.factor_index_ind_sec_weekly_rets(
            merge = False
        )

        daily_ret_5y_base = r.convert_returns_to_base(
            ret_df = daily_ret_5y,
            base_ccy = base_ccy,
            interval = "1d"
        )

        daily_close_5y_base = r.convert_returns_to_base(
            ret_df = daily_close_5y,
            base_ccy = base_ccy,
            interval = "1d"
        )
        
        daily_open_5y_base = r.convert_returns_to_base(
            ret_df = daily_open_5y,
            base_ccy = base_ccy,
            interval = "1d"
        )

        fund_exposures_weekly = None
       
        if getattr(config, "COV_USE_FUND_FACTORS", True):
       
            try:
       
                fdata = FinancialForecastData(tickers = list(tickers), quiet = True)
       
                weekly_prices = r.weekly_close.reindex(
                    index = weekly_ret_5y_base.index
                ).reindex(columns = tickers)
                
                shares_out = r.shares_outstanding
                
                fund_exposures_weekly = fdata.build_pit_fundamental_exposures_weekly(
                    tickers = list(tickers),
                    weekly_index = weekly_ret_5y_base.index,
                    weekly_prices = weekly_prices,
                    shares_outstanding = shares_out,
                    report_lag_days = getattr(config, "FUND_REPORT_LAG_DAYS", 90),
                    price_lag_weeks = getattr(config, "FUND_PRICE_LAG_WEEKS", 1),
                    winsor_p = getattr(config, "FUND_WINSOR_P", (0.05, 0.95)),
                    robust_scale = getattr(config, "FUND_ROBUST_SCALE", "mad"),
                    min_coverage = getattr(config, "FUND_MIN_COVERAGE", 0.6),
                )
                
            except Exception as exc:
              
                logging.warning("Failed to build PIT fundamental exposures: %s", exc)
            
            
        fx_factors_weekly = build_fx_factor_returns(
            ratio_data = r,
            tickers = tickers,
            weekly_index = weekly_ret_5y_base.index,
            base_ccy = base_ccy,
        )
            
        monthly_ret_5y_base = r.convert_returns_to_base(
            ret_df = monthly_ret_5y,
            base_ccy = base_ccy,
            interval = "1mo"
        )
            
        factor_weekly_base = r.convert_returns_to_base(
            ret_df = factor_weekly, 
            base_ccy = base_ccy,
            interval = "1wk"
        )
            
        index_weekly_base = r.convert_returns_to_base(
            ret_df = index_weekly,
            base_ccy = base_ccy,
            interval = "1wk"
        )
            
        ind_weekly_base = r.convert_returns_to_base(
            ret_df = ind_weekly,
            base_ccy = base_ccy, 
            interval = "1wk"
        )
            
        sec_weekly_base = r.convert_returns_to_base(
            ret_df = sec_weekly,
            base_ccy = base_ccy, 
            interval = "1wk"
        )
    
    if annCov is None:
    
        annCov, annCov_desc, extras = shrinkage_covariance(
            daily_5y = daily_ret_5y_base, 
            weekly_5y = weekly_ret_5y_base, 
            monthly_5y = monthly_ret_5y_base, 
            comb_std = comb_std, 
            common_idx = comb_rets.index,
            ff_factors_weekly = factor_weekly_base,
            index_returns_weekly = index_weekly_base,
            industry_returns_weekly = ind_weekly_base,
            sector_returns_weekly = sec_weekly_base,
            macro_factors_weekly = macro_weekly_5y,
            fx_factors_weekly = fx_factors_weekly,
            fund_exposures_weekly = fund_exposures_weekly,
            sector_map = sector_map,
            industry_map = industry_map,
            daily_open_5y = daily_open_5y_base,
            daily_close_5y = daily_close_5y_base,
            use_excess_ff = False,
            use_log_returns = getattr(config, "COV_USE_LOG_RETURNS", True),
            use_oas = getattr(config, "COV_USE_OAS", True),
            use_block_prior = getattr(config, "COV_USE_BLOCK_PRIOR", True),
            use_regime_ewma = getattr(config, "COV_USE_REGIME_EWMA", True),
            use_glasso = getattr(config, "COV_USE_GLOSSO", True),
            use_fund_factors = getattr(config, "COV_USE_FUND_FACTORS", True),
            use_fx_factors = getattr(config, "COV_USE_FX_FACTORS", True),
            use_factor_term_structure = getattr(config, "COV_USE_TERM_STRUCTURE", False),
            description = True,
        )
     
        try:
     
            annCov.to_pickle(cov_path)
     
        except Exception as exc:
     
            logging.warning("Failed to cache covariance to %s: %s", cov_path, exc)
    
    print('------------------------------------------------------------')
    
    print('\nextras\n', extras) 
        
    if sigma_prior is None:
        
        sigma_prior = shrinkage_covariance(
            daily_5y = daily_ret_5y_base, 
            weekly_5y = weekly_ret_5y_base, 
            monthly_5y = monthly_ret_5y_base, 
            comb_std = comb_std, 
            common_idx = tickers, 
            w_F = 0,       
            ff_factors_weekly = factor_weekly_base,
            index_returns_weekly = index_weekly_base,
            industry_returns_weekly = ind_weekly_base,
            sector_returns_weekly = sec_weekly_base,
            macro_factors_weekly = macro_weekly_5y,
            fx_factors_weekly = fx_factors_weekly,
            fund_exposures_weekly = fund_exposures_weekly,
            sector_map = sector_map,
            industry_map = industry_map,
            use_log_returns = getattr(config, "COV_USE_LOG_RETURNS", True),
            use_oas = getattr(config, "COV_USE_OAS", True),
            use_block_prior = getattr(config, "COV_USE_BLOCK_PRIOR", True),
            use_regime_ewma = getattr(config, "COV_USE_REGIME_EWMA", True),
            use_glasso = getattr(config, "COV_USE_GLOSSO", True),
            use_fund_factors = getattr(config, "COV_USE_FUND_FACTORS", True),
            use_fx_factors = getattr(config, "COV_USE_FX_FACTORS", True),
            use_factor_term_structure = getattr(config, "COV_USE_TERM_STRUCTURE", False),
        )
     
        try:
     
            sigma_prior.to_pickle(sigma_prior_path)
     
        except Exception as exc:
     
            logging.warning("Failed to cache sigma_prior to %s: %s", sigma_prior_path, exc)
   
    else:
   
        logging.info("Loaded cached sigma_prior from %s", sigma_prior_used_path)
    
    print('sigma_prior', np.isfinite(sigma_prior).all())
        
    for t in sigma_prior.columns:
        
        if not np.isfinite(sigma_prior[t]).all():
        
            print(t)
    
    annCov = annCov.reindex(index = tickers, columns = tickers)
        
    weeklyCov = annCov / 52
    
    cov_df = pd.DataFrame(annCov, index = tickers, columns = tickers)
    
    annCov_desc_df = pd.DataFrame(annCov_desc.T)
        
    Signal = pd.read_excel(config.DATA_FILE, sheet_name = "Signal Scores", index_col = 0).iloc[-1]
    
    beta_data = pd.read_excel(config.FORECAST_FILE, sheet_name = "COE", index_col = 0)["Beta_Levered_Used"]

    beta_data.index = beta_data.index.str.upper()

    beta = beta_data.reindex(tickers)
    
    print('end of load_excel_data')
    
    return (weekly_ret_5y, weekly_ret_5y_base, tickers, weeklyCov, annCov, annCov_desc_df, cov_df,
            Signal, beta, 
            score, comb_rets, safe_rets, ticker_ind, ticker_sec, comb_std,
            bear_rets, bull_rets, ticker_mcap, sigma_prior, factor_pred, factor_r2,
            d_to_e, tax)


def get_buy_sell_signals_from_scores(
    tickers: List[str],
    signal_score: pd.Series
) -> Tuple[pd.DataFrame, List[bool], List[bool]]:
    """
    Derive simple buy/sell flags from a score vector.

    Rule: for ticker t, Buy = (score_t > 0), Sell = (score_t < 0).

    Args:
        tickers: Ordered list of tickers to evaluate.
        signal_score: pd.Series of scores indexed by ticker (case-insensitive).

    Returns:
        df:   pd.DataFrame with columns [Ticker, Buy, Sell, Score].
        buy_flags:  List[bool] aligned to `tickers`.
        sell_flags: List[bool] aligned to `tickers`.

    Notes:
        Missing tickers in `signal_score` default to Score=0 (neither buy nor sell).
    """

    signal_score.index = signal_score.index.str.upper()

    buy_flags = [bool(signal_score.get(t, 0) > 0) for t in tickers]
   
    sell_flags = [bool(signal_score.get(t, 0) < 0) for t in tickers]

    df = pd.DataFrame({
        "Ticker": tickers,
        "Buy": buy_flags,
        "Sell": sell_flags,
        "Score": [signal_score.get(t, 0) for t in tickers],
    })
    
    return df, buy_flags, sell_flags


def _add_table(
    ws, 
    table_name: str
) -> None:
    """
    Convert the used range of an openpyxl worksheet into a styled Excel table.

    Creates a table covering A1:(lastcol,lastrow) with `TableStyleMedium9`.

    Args:
        ws: openpyxl worksheet object.
        table_name: Name to assign to the created table.

    Side Effects:
        Modifies `ws` in place (adds a table).
    """

    last_col = get_column_letter(ws.max_column)
    
    table = Table(displayName = table_name, ref = f"A1:{last_col}{ws.max_row}")
    
    table.tableStyleInfo = TableStyleInfo(
        name = "TableStyleMedium9",
        showFirstColumn = False,
        showLastColumn = False,
        showRowStripes = True,
        showColumnStripes = False,
    )
    
    ws.add_table(table)


def add_weight_cf(
    ws
) -> None:
    """
    Apply “traffic-light” conditional formatting to numeric weight columns.

    Rules applied to each numeric column (rows 2…max_row):
      
        - Red fill if value < 0.01 (tiny / effectively zero).
      
        - Yellow fill if 1.95 ≤ value ≤ 2.05 (aggregated check; e.g., sum validation).
      
        - Green fill if cell is neither 0 nor 2 (treated as a “valid” entry).
   
    Also sets number format "0.00" for all numeric cells in data rows.

    Args:
        ws: openpyxl worksheet containing a weight table with headers in row 1.

    Side Effects:
        Adds conditional formatting rules and number formats to `ws`.
    """

    red = PatternFill(start_color = "FFC7CE", end_color = "FFC7CE", fill_type = "solid")
   
    green = PatternFill(start_color = "C6EFCE", end_color = "C6EFCE", fill_type = "solid")
   
    yellow = PatternFill(start_color = "FFEB9B", end_color = "FFEB9B", fill_type = "solid")

    first_data_col = 2                         
   
    last_data_col = ws.max_column

    for col_idx in range(first_data_col, last_data_col + 1):
   
        col_letter = get_column_letter(col_idx)
        
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"

        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator = "lessThan", formula=['0.01'], fill = red, stopIfTrue = True)
        )
        
       
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator = "between", formula=['1.95', '2.05'], fill = yellow, stopIfTrue = True)
        )
       
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula = [f"AND({col_letter}2<>0,{col_letter}2<>2)"], fill = green)
        )

    for row in ws.iter_rows(min_row = 2, min_col = first_data_col, max_col = last_data_col):
        
        for cell in row:
            
            cell.number_format = "0.00"


def write_excel_results(
    excel_file: str, 
    sheets: dict[str, pd.DataFrame]
) -> None:
    """
    Write result DataFrames to Excel and apply styling.

    Creates/replaces the following sheets (keys must be present in `sheets`):
        - "Ticker Performance"
        - "Today_Buy_Sell"
        - "Covariance"
        - "Covariance Description"
        - "Bounds"
        - "Weights"
        - "Industry Breakdown"
        - "Sector Breakdown"
        - "Portfolio Performance"

    Formatting:
   
        - Converts each written range to an Excel table with a consistent style.
   
        - Applies green/red fills to Buy/Sell booleans.
   
        - Applies 3-colour scale to the covariance numeric area.
   
        - Applies traffic-light rules to weight sheets.

    Args:
        excel_file: Path to the workbook (will be opened with mode='a' and
            `if_sheet_exists='replace'`).
        sheets: Mapping from sheet name (key) to DataFrame (value).

    Raises:
        KeyError: if a required sheet key is missing.
        ValueError: if DataFrame shapes are incompatible with expected formatting.
    """

    green = PatternFill(start_color = "C6EFCE", end_color = "C6EFCE", fill_type = "solid")
   
    red = PatternFill(start_color = "FFC7CE", end_color = "FFC7CE", fill_type = "solid")

    with pd.ExcelWriter(excel_file, engine = "openpyxl", mode = "a", if_sheet_exists = "replace") as writer:
        
        tperf = ensure_headers_are_strings(
            df = sheets["Ticker Performance"].copy()
        )
        
        tperf.to_excel(writer, sheet_name="Ticker Performance", index = True)
        
        ws_tp = writer.sheets["Ticker Performance"]

        tbs = ensure_headers_are_strings(
            df = sheets["Today_Buy_Sell"].copy()
        )
        
        tbs.to_excel(writer, sheet_name = "Today_Buy_Sell", index = False)

        ws = writer.sheets["Today_Buy_Sell"]
        
        n = tbs.shape[0] + 1
        
        ws.conditional_formatting.add(f"B2:B{n}", CellIsRule(operator = "equal", formula = ["TRUE"],  fill = green))
       
        ws.conditional_formatting.add(f"B2:B{n}", CellIsRule(operator = "equal", formula = ["FALSE"], fill = red))
       
        ws.conditional_formatting.add(f"C2:C{n}", CellIsRule(operator = "equal", formula = ["TRUE"],  fill = green))
       
        ws.conditional_formatting.add(f"C2:C{n}", CellIsRule(operator = "equal", formula = ["FALSE"], fill = red))
        
        _add_table(
            ws = writer.sheets['Ticker Performance'], 
            table_name = "TickerPerformanceTable"
        )
        
        _add_table(
            ws = ws, 
            table_name = "TodayBuySellTable"
        )

        cov = ensure_headers_are_strings(
            df = sheets["Covariance"].copy()
        )

        bnds = ensure_headers_are_strings(
            df = sheets["Bounds"].copy()
        )
        
        bnds.to_excel(writer, sheet_name = "Bounds")
        
        bnds_ws = writer.sheets["Bounds"]
        
        _add_table(
            ws = bnds_ws, 
            table_name = "BoundsTable"
        )

        cov.to_excel(writer, sheet_name = "Covariance")
        
        cov_ws = writer.sheets["Covariance"]
        
        _add_table(
            ws = cov_ws, 
            table_name = "CovarianceTable"
        )

        data = cov.values.astype(float)

        min_val = data.min()
        
        mean_val = data.mean()
        
        max_val = data.max()
        
        color_rule = ColorScaleRule(
            start_type = 'num', start_value = str(min_val), start_color = 'FFFFFF',
            mid_type = 'num', mid_value = str(mean_val), mid_color = 'FFDD99',
            end_type = 'num', end_value = str(max_val), end_color = 'FF0000'
        )

        rng = f"B2:{get_column_letter(cov_ws.max_column)}{cov_ws.max_row}"
       
        cov_ws.conditional_formatting.add(rng, color_rule)
        
        cov_desc = ensure_headers_are_strings(
            df = sheets['Covariance Description'].copy()
        )
        
        cov_desc.to_excel(writer, sheet_name = "Covariance Description", index = True)
        
        cov_desc_ws = writer.sheets["Covariance Description"]
                
        _add_table(
            ws = cov_desc_ws, 
            table_name = "CovarianceDescriptionTable"
        )


        def dump_weight_sheet(
            df_key: str,
            sheet_name: str, 
            tablename: str
        ) -> None:
            """
            Write one weights-style worksheet and apply standard formatting controls.

            Workflow:
           
            1) sanitise headers to strings;
           
            2) write the selected DataFrame to ``sheet_name``;
           
            3) apply weight-focused conditional formatting rules;
           
            4) convert the used range into a styled Excel table.

            Parameters
            ----------
            df_key : str
                Key identifying the source DataFrame in ``sheets``.
            sheet_name : str
                Destination worksheet name.
            tablename : str
                Excel table display name to assign.
            """
       
            df = ensure_headers_are_strings(
                df = sheets[df_key].copy()
            )
       
            df.to_excel(writer, sheet_name = sheet_name, index = False)
       
            ws = writer.sheets[sheet_name]
       
            add_weight_cf(
                ws = ws
            )
       
            _add_table(
                ws = ws, 
                table_name = tablename
            )


        dump_weight_sheet(
            df_key = "Weights", 
            sheet_name = "Portfolio Weights", 
            tablename = "WeightsTable"
        )

        for key, sheet_name, tablename in [
            ("Portfolio Performance", "Portfolio Performance", "PortfolioPerformanceTable"),
            ("Industry Breakdown", "Industry Breakdown", "IndustryBreakdownTable"),
           ("Sector Breakdown", "Sector Breakdown", "SectorBreakdownTable"),
        ]:
       
            df = ensure_headers_are_strings(
                df = sheets[key].copy()
            )
           
            df.to_excel(writer, sheet_name = sheet_name, index = "Industry" not in key)
           
            _add_table(
                ws = writer.sheets[sheet_name],
                table_name = tablename
            )
            

def benchmark_rets(
    benchmark: str, 
    start: dt.date, 
    end: dt.date, 
    steps: int
) -> float:
    """
    Download benchmark prices and compute weekly returns and trailing annual return.

    Supported `benchmark` strings (case-insensitive):
       
        'SP500'  → '^GSPC'
       
        'NASDAQ' → '^IXIC'
       
        'FTSE'   → '^FTSE'
       
        'FTSE ALL-WORLD' → 'VWRL.L'
       
        'ALL'    → equal-weighted average of SP500, NASDAQ, FTSE annualised returns

    Computation:
    
        - Price series resampled to weekly last; returns = pct_change().dropna().
        
        - `last_year_rets` filters to the last calendar year window (`config.YEAR_AGO`).
        
        - Trailing annual return = ∏(1 + weekly_ret) − 1 over `last_year_rets`.

    Args:
        benchmark: Name from the supported list above.
        start: Start date for download.
        end: End date for download.
        steps: Periods per year used by external annualisers when `benchmark='ALL'`.

    Returns:
        benchmark_ann_ret: float, trailing annual return over the last year window.
        rets: pd.Series, full weekly returns over [start, end].
        last_year_rets: pd.Series, weekly returns within the last year.

    Raises:
        ValueError: for unsupported `benchmark`.
        Exception: network/IO errors from `yfinance` and resampling.
    """
    
    if benchmark.upper() == 'SP500':
        
        close = yf.download('^GSPC', start = start, end = end, auto_adjust = True)['Close'].squeeze()
    
    elif benchmark.upper() == 'NASDAQ':
        
        close = yf.download('^IXIC', start = start, end = end, auto_adjust = True)['Close'].squeeze()
    
    elif benchmark.upper() == 'FTSE':
        
        close = yf.download('^FTSE', start = start, end = end, auto_adjust = True)['Close'].squeeze()
        
    elif benchmark.upper() == 'FTSE ALL-WORLD':
        
        close = yf.download('VWRL.L', start = start, end = end, auto_adjust = True)['Close'].squeeze()
    
    elif benchmark.upper() == 'ALL':
    
        close_sp = yf.download('^GSPC', start = start, end = end, auto_adjust = True)['Close'].squeeze()
     
        close_nd = yf.download('^IXIC', start = start, end = end, auto_adjust = True)['Close'].squeeze()
     
        close_ft = yf.download('^FTSE', start = start, end = end, auto_adjust = True)['Close'].squeeze()
    
        rets_sp = close_sp.resample('W').last().pct_change().dropna()
     
        rets_nd = close_nd.resample('W').last().pct_change().dropna()
     
        rets_ft = close_ft.resample('W').last().pct_change().dropna()
    
        return float((pf.annualise_returns(rets_sp, steps) +
                      pf.annualise_returns(rets_nd, steps) +
                      pf.annualise_returns(rets_ft, steps)) / 3)
    
    else:
     
        raise ValueError("Unsupported benchmark")
    
    rets = close.resample('W').last().pct_change().dropna()
    
    last_year_rets = rets[rets.index >= pd.to_datetime(config.YEAR_AGO)]
    
    benchmark_ann_ret = (last_year_rets + 1).prod() - 1
    
    return benchmark_ann_ret, rets, last_year_rets 


def compute_adv_dollar(
    close: pd.DataFrame,
    volume: pd.DataFrame,
    window: int = 20,
    fx_to_usd: pd.DataFrame | None = None,
    min_valid_ratio: float = 0.6,
) -> pd.Series:
    """
    Trailing average daily *dollar* volume over `window` trading days.

    close, volume: DataFrames aligned by index and columns (tickers).
    fx_to_usd: optional DataFrame (same shape) to convert local prices to USD.
               If None, assumes `close` already in USD/base currency.

    Returns a pd.Series (index=tickers) with ADV$.
    """
    
    close = close.copy()
    
    volume = volume.copy()

    if fx_to_usd is not None:
    
        close = close * fx_to_usd

    dollar_vol = (close * volume).astype(float)

    roll = dollar_vol.rolling(window = window, min_periods = int(window * min_valid_ratio)).mean()
    
    adv = roll.iloc[-1].fillna(0.0)

    return adv


def compute_amihud(
    close: pd.DataFrame,
    volume: pd.DataFrame,
    window: int = 60,
    fx_to_usd: pd.DataFrame | None = None,
    min_valid_ratio: float = 0.6,
    eps: float = 1e-12,
) -> pd.Series:
    """
    Amihud (2002) illiquidity: mean_t ( |r_t| / $Vol_t ) over a trailing `window`.

    `close`, `volume` as DataFrames (index dates, columns tickers).
    `r_t` is close-to-close pct return. `$Vol_t` is price*volume in USD/base.

    Returns a pd.Series (index=tickers) with the Amihud measure (higher=worse liquidity).
    """
    
    close_ = close.copy()
    
    if fx_to_usd is not None:
    
        close_ = close_ * fx_to_usd

    dollar_vol = (close_ * volume).astype(float).clip(lower=eps)
    
    ret = close.pct_change().abs()

    ratio = (ret / dollar_vol).replace([np.inf, -np.inf], np.nan)

    ami = ratio.rolling(window = window, min_periods = int(window * min_valid_ratio)).mean().iloc[-1]
    
    return ami.fillna(ami.median() if np.isfinite(ami.median()) else 0.0)


def compute_mcap_bounds(
    market_cap: pd.Series,
    er: pd.Series,
    vol: pd.Series,
    score: pd.Series,
    tickers: pd.Index,
    ticker_sec: pd.Series,
    min_all: float,
    max_all: float,
    max_all_l: float,
    min_all_u: float,
    factor_pred: pd.Series,
    factor_r2: pd.Series,
    *,
    adv_dollar: pd.Series | None = None,        
    amihud: pd.Series | None = None,              
    sector_limits: dict | None = None,            
    restricted: pd.Series | None = None,         
    a_mcap: float = 0.5,      
    d_vol: float = 1.0,   
    b_adv: float = 0.5,       
    h_amihud: float = 0.25,  
    fp_exp: float = 1.0,
    r2_exp: float = 1.0,
    winsor_q: float = 0.99,  
    ub_pow_cap: float = 0.5,  
    lb_pow_cap: float = 1.0, 
    lb_max_as_frac_of_ub: float = 0.5,
    sector_multipliers: dict | None = None,  
    return_diag: bool = False,
    small_cap_threshold: float = 0.1,
):
    """
    Derive per-asset lower and upper bounds from capacity, liquidity, and sector structure.

    Economic rationale
    ------------------
    The routine builds practical long-only box bounds ``lb_i <= w_i <= ub_i`` that
    combine market-cap scale, forecast quality, score quality, and optional trading
    liquidity. The objective is to preserve deployability while preventing fragile
    allocations in low-capacity names.

    Core capacity score
    -------------------
    For each asset ``i``:

    ``eligible_i = 1{mu_i > 0 and score_i > 0 and not restricted_i}``

    ``fp_mult_i = 1 + max(factor_pred_i, 0)^fp_exp * r2_i^r2_exp``

    ``cap_i = market_cap_i^a_mcap * fp_mult_i / vol_i^d_vol``.

    Optional liquidity corrections:

    ``cap_i <- cap_i * adv_i^b_adv``  (ADV-based scaling),

    ``cap_i <- cap_i / (1 + h_amihud * amihud_i)``  (illiquidity penalty).

    After winsorisation, the normalised capacity share is:

    ``frac_i = cap_i / sum_j cap_j``.

    Upper envelope construction
    ---------------------------
    - Sector-aware allocation (if caps are provided) uses score-weighted intra-sector
      shares.
 
    - Base upper envelope:
      ``ub_base_i = (sector_alloc_i ^ ub_pow_cap) * score01_i``.
 
    - Apply sector multipliers, global clips, and hard overrides for selected
      sector/size regimes.

    Lower envelope construction
    ---------------------------
    - Base lower envelope:
      ``lb_base_i = (frac_i ^ lb_pow_cap) * score01_i``.
  
    - Clip globally and enforce a relative constraint:
      ``lb_i <= lb_max_as_frac_of_ub * ub_i``.
 
    - Blend with legacy references based on square-root and full market-cap shares
      to avoid unstable minima.

    Feasibility repair layer
    ------------------------
    - If ``sum(lb) > 1``, lower bounds are proportionally scaled.
 
    - If ``sum(ub) < 1``, residual mass is distributed through available headroom
      while respecting sector caps and hard per-name ceilings.
 
    - Enforce ``ub_i >= lb_i`` after each repair stage.

    Advantages
    ----------
    - Integrates deployability constraints before optimisation, reducing solver stress.
 
    - Incorporates both conviction (returns/scores) and trading realism (ADV/Amihud).
 
    - Preserves sector-budget consistency while maintaining cross-sectional flexibility.
 
    - Adds deterministic feasibility fixes, reducing downstream infeasible runs.

    Parameters
    ----------
    market_cap, er, vol, score : pd.Series
        Per-ticker capacity and signal inputs.
    tickers : pd.Index
        Ordered optimisation universe.
    ticker_sec : pd.Series
        Sector label per ticker.
    min_all, max_all, max_all_l, min_all_u : float
        Global floor/ceiling controls for bounds.
    factor_pred, factor_r2 : pd.Series
        Factor tilt and explanatory strength inputs for capacity scaling.
    adv_dollar, amihud : pd.Series | None
        Optional liquidity controls.
    sector_limits : dict | None
        Optional sector cap dictionary.
    restricted : pd.Series | None
        Optional boolean hard-exclusion flags.
    a_mcap, d_vol, b_adv, h_amihud, fp_exp, r2_exp : float
        Exponents/sensitivity parameters.
    winsor_q, ub_pow_cap, lb_pow_cap, lb_max_as_frac_of_ub : float
        Envelope shape and clipping parameters.
    sector_multipliers : dict | None
        Optional sector tilt multipliers.
    return_diag : bool
        If ``True``, include intermediate diagnostics in the return.
    small_cap_threshold : float
        Threshold for additional small-cap upper-bound tightening.

    Returns
    -------
    tuple
        ``(lb, ub, frac)`` or ``(lb, ub, frac, diag)`` when diagnostics are requested.
    """
    
    idx = pd.Index(tickers)
    
    S = lambda x, fill = 0.0: pd.Series(x, index = getattr(x, "index", idx)).reindex(idx).astype(float).fillna(fill)

    mc = S(market_cap, 0.0).clip(lower = 0.0)
    
    mu = S(er, 0.0)
    
    sc = S(score, 0.0)
    
    sec = pd.Series(ticker_sec, index = getattr(ticker_sec, "index", idx)).reindex(idx)
    
    fp = S(factor_pred, 0.0)
    
    r2 = S(factor_r2, 0.0).clip(lower = 0.0, upper = 1.0)
    
    sig = S(vol, np.nan)
    
    eps = 1e-12

    sig = sig.fillna(sig.median()).clip(lower = eps)

    if (sc > 0).any():

        sc_max = sc.max()
        
        sc01 = sc / sc_max

    else:

        sc01 = pd.Series(0.0, index = idx)

    elig = (mu > 0) & (sc > 0)

    if restricted is not None:

        restr = pd.Series(restricted, index = idx).fillna(False).astype(bool)

        elig &= (~restr)

    fp_mult = 1.0 + ((fp.clip(lower = 0.0) ** fp_exp) * (r2 ** r2_exp))

    cap = (mc.clip(lower = eps) ** a_mcap) * fp_mult / (sig ** d_vol)

    if adv_dollar is not None:

        adv = S(adv_dollar, 0.0).clip(lower = eps)

        cap *= adv ** b_adv

    if amihud is not None:

        ami = S(amihud, 0.0).clip(lower = 0.0)

        cap *= 1.0 / (1.0 + h_amihud * ami)

    cap = cap.clip(upper = float(cap.quantile(winsor_q)))

    cap = cap.where(elig, 0.0)

    tot = float(cap.sum())
    
    if tot <= eps:
    
        lb = pd.Series(0.0, index = idx)
        
        ub = lb.copy()
        
        frac = lb.copy()
        
        if return_diag:
        
            return (lb, ub, frac, {"note": "no eligible capacity"}) 
        
        else:
            
            return (lb, ub, frac)

    frac = cap / tot 

    mcap_sqrt = np.sqrt(mc.clip(lower=0.0))

    mcap_sqrt_sum = float(mcap_sqrt[elig].sum())

    if mcap_sqrt_sum <= eps:

        mcap_sqrt_sum = 1.0

    frac_sqrt = mcap_sqrt / mcap_sqrt_sum

    mcap_full_sum = float(mc[elig].sum())

    if mcap_full_sum <= eps:

        mcap_full_sum = 1.0

    frac_full = mc / mcap_full_sum 
    
    if sector_limits:

        sec_caps = pd.Series({k: float(v) for k, v in sector_limits.items()})

        sec_key = sec.fillna("UNKNOWN").astype(str)

        by_sec = (frac * sc01).groupby(sec_key)

        share_in_sec = by_sec.transform(lambda s: s / max(s.sum(), eps))

        ub_sec_alloc = share_in_sec * sec_key.map(sec_caps).fillna(0.0)

    else:

        ub_sec_alloc = frac  

    ub_base = (ub_sec_alloc.pow(ub_pow_cap) * sc01).fillna(0.0)

    if sector_multipliers:

        mult = sec.map(lambda x: sector_multipliers.get(x, 1.0)).astype(float).fillna(1.0)

        ub_base *= mult

    ub = ub_base.clip(lower = min_all_u, upper = max_all).where(elig, 0.0)

    hard_cap = 0.025
    
    TRILLION_THRESHOLD = 1e12
    
    hcc_mask = sec.isin(["Consumer Staples", "Consumer Discretionary"]) & (mc < TRILLION_THRESHOLD)

    ub.loc[hcc_mask] = np.minimum(ub.loc[hcc_mask], hard_cap)

    lb_base = (frac.pow(lb_pow_cap) * sc01).fillna(0.0)

    lb = lb_base.where(elig, 0.0).clip(lower = min_all, upper = max_all_l)

    lb = np.minimum(lb, lb_max_as_frac_of_ub * np.maximum(ub, 0.0))
    
    if elig.any():
        
        lb_candidates = pd.concat(
            [
                lb.rename("low"),     
                frac_sqrt.rename("mcap_sqrt"),
                frac_full.rename("full_weight"),
            ],
            axis = 1,
        )

        lb_legacy = lb_candidates.min(axis=1).clip(lower=min_all)

        lb = lb.where(~elig, lb_legacy)
        
    small_mask = elig & (frac_sqrt < small_cap_threshold)
    
    if small_mask.any():
        
        ub_small = np.minimum(4.0 * frac_sqrt, ub) 
        
        ub = ub.where(~small_mask, ub_small)

    if hasattr(config, "TICKER_EXEMPTIONS") and config.TICKER_EXEMPTIONS:

        for t in config.TICKER_EXEMPTIONS:

            if t in idx:

                lb.loc[t] = 0.01
                
                ub.loc[t] = min(0.075, max_all) 

    ub = np.maximum(ub, lb)

    sum_lb = float(lb.sum())

    if sum_lb > 1.0 - 1e-12:

        lb *= (1.0 / max(sum_lb, eps))

    sum_ub = float(ub.sum())

    if sum_ub < 1.0 - 1e-12:

        headroom = (max_all - ub).clip(lower = 0.0)

        headroom.loc[hcc_mask] = np.minimum(headroom.loc[hcc_mask], hard_cap - ub.loc[hcc_mask]).clip(lower = 0.0)

        if sector_limits:
           
            sec_caps_abs = sec.map(lambda x: sector_limits.get(x, 0.0)).fillna(0.0)
           
            sec_used = ub.groupby(sec).transform('sum')
           
            sec_room = (sec_caps_abs - sec_used).clip(lower = 0.0)
           
            headroom = np.minimum(headroom, sec_room)

        room = float(headroom.sum())
        
        if room > eps:
        
            incr = headroom * ((1.0 - sum_ub) / room)
        
            ub = (ub + incr).clip(upper = max_all)

            ub.loc[hcc_mask] = np.minimum(ub.loc[hcc_mask], hard_cap)
    
    ub = np.maximum(ub, lb)

    is_trillion_plus = (mc >= TRILLION_THRESHOLD) | (mc.index.isin(config.TICKER_EXEMPTIONS))

    ub = ub.where(is_trillion_plus, ub.clip(upper=0.025))

    ub = np.maximum(ub, lb)

    lb = lb / 2
    
    lb = np.minimum(lb, 0.025)
    
    lb = np.maximum(lb, min_all)

    if return_diag:
     
        diag = dict(
            eligible = elig, 
            frac = frac,
            ub_base = ub_base,
            lb_base = lb_base,
            sum_lb = float(lb.sum()), 
            sum_ub = float(ub.sum())
        )
        
        return lb, ub, frac, diag

    return lb, ub, frac



def main() -> None:
    """
    Run the full optimisation workflow and export results.

    Steps:
       
        1) Download benchmark data and compute weekly/annual returns.
      
        2) Load Excel-based forecasts and analytics; build annual/weekly covariance
        (including a BL prior covariance).
      
        3) Build per-ticker bounds using market cap, scores, volatility and factor tilt.
      
        4) Instantiate `PortfolioOptimiser` with returns, covariances, histories,
        benchmark, classifications, bounds, BL inputs, sector caps, and rates.
      
        5) Solve base optimisers (MSR, Sortino, MIR, MSP, BL), DSR and ASR variants,
        and composite CCP portfolios (`comb_port`, …, `comb_port7`).
      
        6) Compute annual and weekly volatilities for each solution.
      
        7) Build per-ticker weight tables and industry/sector breakdowns (% of capital).
      
        8) Produce portfolio and ticker performance reports via `PortfolioAnalytics`.
      
        9) Write all outputs to Excel with conditional formatting and tables.

    Side Effects:
     
        - Logs progress to the root logger.
     
        - Appends/replaces sheets in `config.PORTFOLIO_FILE`.
     
        - Prints intermediate optimised weights and volatility figures.

    Raises:
        AssertionError: if key indices do not align to the configured tickers.
        RuntimeError / ValueError: bubbled up from optimiser/calibration steps or IO.
    """
   
    logging.info("Starting portfolio optimisation script...")
        
    benchmark_ret, benchmark_weekly_rets, last_year_benchmark_weekly_rets = benchmark_rets(
        benchmark = config.benchmark, 
        start = config.FIVE_YEAR_AGO, 
        end = config.TODAY, 
        steps = 52
    )

    logging.info("Loading Excel data...")
  
    (weekly_ret, weekly_ret_5y_base, tickers, weekly_cov, ann_cov, ann_cov_desc, cov_df, signal_score, beta, comb_score, comb_rets, safe_rets, ticker_ind, ticker_sec, comb_ann_std, bear_rets, bull_rets, mcap, sigma_prior, factor_pred, factor_r2, d_to_e, tax) = load_excel_data()
    
    assert comb_rets.index.equals(pd.Index(tickers)), "comb_rets index ≠ tickers"
    
    assert cov_df.index.equals(pd.Index(tickers)), "cov_df index ≠ tickers"
    
    assert cov_df.columns.equals(pd.Index(tickers)), "cov_df columns ≠ tickers"
    
    assert comb_rets.isna().sum() == 0, "Some tickers have no forecasted return!"

    buy_sell_today_df, buy_flags, sell_flags = get_buy_sell_signals_from_scores(
        tickers = tickers, 
        signal_score = signal_score
    )
    
    logging.info("Data loaded and signals created.")
     
    rf_rate = config.RF
    
    w_max = config.MAX_WEIGHT
    
    w_min = 2 / money_in_portfolio
    
    max_all_l = w_max / 2
    
    min_all_u = w_min * 2
    
    tickers_index = pd.Index(tickers)
    
    mcap_bnd_l, mcap_bnd_h, mcap_vol_beta, diag = compute_mcap_bounds(
        market_cap = mcap, 
        er = safe_rets, 
        vol = comb_ann_std, 
        score = comb_score,
        tickers = tickers_index, 
        ticker_sec = ticker_sec,
        min_all = w_min, 
        max_all = w_max, 
        max_all_l = max_all_l, 
        min_all_u = min_all_u,
        factor_pred = factor_pred,
        factor_r2 = factor_r2,
        return_diag = True
    )
    
    print("Capacity Diagnostics:\n", diag)
    
    print("Market Cap Bounds Low:\n", mcap_bnd_l)
    
    print("Market Cap Bounds High:\n", mcap_bnd_h)
    
    mcap["SGLP.L"] = mcap.max()
    
    bounds_df = pd.DataFrame({
        'mid': mcap_vol_beta,
        'Low': mcap_bnd_l,
        'High': mcap_bnd_h
    })
    
    last_year_weekly_rets = weekly_ret_5y_base.loc[weekly_ret_5y_base.index >= pd.to_datetime(config.YEAR_AGO)]   
        
    n_last_year_weeks = len(weekly_ret_5y_base)
    
    last_5_year_weekly_rets = weekly_ret_5y_base.loc[weekly_ret_5y_base.index >= pd.to_datetime(config.FIVE_YEAR_AGO)]    
    
    pa = pf.PortfolioAnalytics(cache = False)     
    
    logging.info("Optimising Portfolios...")
   
    opt = PortfolioOptimiser(
        er = safe_rets, 
        cov = ann_cov, 
        scores = comb_score,
        weekly_ret_1y = last_year_weekly_rets,
        last_5_year_weekly_rets = last_5_year_weekly_rets,
        benchmark_weekly_ret = last_year_benchmark_weekly_rets,
        ticker_ind = ticker_ind, 
        ticker_sec = ticker_sec,
        bnd_h = mcap_bnd_h, 
        bnd_l = mcap_bnd_l,
        comb_std = comb_ann_std, 
        sigma_prior = sigma_prior, 
        mcap = mcap,
        sector_limits = config.sector_limits,
        rf_annual = config.RF, 
        rf_week = config.RF_PER_WEEK,
        gamma = config.GAMMA,
    )

    base_bundle = opt.get_base_portfolios_cached()
    
    w_bl, mu_bl, sigma_bl = base_bundle["BL"]
    
    weights_by_label: Dict[str, pd.Series] = {
        "MSR": base_bundle["MSR"],
        "Sortino": base_bundle["Sortino"],
        "MIR": base_bundle["MIR"],
        "MSP": base_bundle["MSP"],
        "MDP": base_bundle["MDP"],
        "BL": w_bl,
        "Deflated_MSR": base_bundle["DSR"],
        "Adjusted_MSR": base_bundle["ASR"],
        "UPM-LPM": base_bundle["ULPM"],
        "Upside/Downside": base_bundle["UDRP"],
    }
    
    optimiser_registry: Dict[str, Callable[[], pd.Series]] = {
        "GMV": opt.min_variance,
        "Combination": opt.comb_port,
        "Combination1": opt.comb_port1,
        "Combination2": opt.comb_port2,
        "Combination3": opt.comb_port3,
        "Combination4": opt.comb_port4,
        "Combination5": opt.comb_port5,
        "Combination6": opt.comb_port6,
        "Combination7": opt.comb_port7,
        "Combination8": opt.comb_port8,
        "Combination9": opt.comb_port9,
        "Combination10": opt.comb_port10,
        "Combination11": opt.comb_port11,
        "Combination12": opt.comb_port12,
        "Combination13": opt.comb_port13,
    }
    
    default_optimisers = [
        "MSR",
        "Sortino",
        "MIR",
        "GMV",
        "MDP",
        "MSP",
        "BL",
        "Deflated_MSR",
        "Adjusted_MSR",
        "UPM-LPM",
        "Upside/Downside",
        "Combination",
        "Combination1",
        "Combination2",
        "Combination3",
        "Combination4",
        "Combination5",
        "Combination6",
        "Combination7",
        "Combination8",
        "Combination9",
        "Combination10",
        "Combination11",
        "Combination12",
        "Combination13",
    ]
    
    selected_raw = getattr(config, "OPTIMISERS_TO_RUN", default_optimisers)
    
    if selected_raw is None:
        
        selected_labels = list(default_optimisers)
    
    elif isinstance(selected_raw, str):
        
        selected_labels = [s.strip() for s in selected_raw.split(",") if s.strip()]
    
    else:
        
        selected_labels = [str(s).strip() for s in selected_raw if str(s).strip()]
    
    if not selected_labels:
        
        selected_labels = list(default_optimisers)
    
    unknown_labels = [label for label in selected_labels if label not in default_optimisers]
    
    if unknown_labels:
        
        logging.warning("Ignoring unknown OPTIMISERS_TO_RUN entries: %s", unknown_labels)
    
    selected_labels = [label for label in selected_labels if label in default_optimisers]
    
    selected_labels = list(dict.fromkeys(selected_labels))
    
    if not selected_labels:
        
        selected_labels = list(default_optimisers)
    
    logging.info("Optimiser profile: %s", getattr(config, "OPT_PROFILE", "exact"))
    
    logging.info("Optimisers selected: %s", ", ".join(selected_labels))
    
    for label in selected_labels:
        
        if label not in weights_by_label:
            
            w_val = optimiser_registry[label]()
            
            if not isinstance(w_val, pd.Series):
                
                w_val = pd.Series(np.asarray(w_val, float).ravel(), index = ann_cov.index, name = label)
            
            weights_by_label[label] = w_val
        
        weights_by_label[label] = (
            weights_by_label[label]
            .reindex(ann_cov.index)
            .astype(float)
            .fillna(0.0)
        )
        
        logging.info("\n%s Weights:\n%s", label, weights_by_label[label])
    
    selected_weights = {label: weights_by_label[label] for label in selected_labels}
    
    W = np.vstack([selected_weights[label].to_numpy(dtype = float) for label in selected_labels])
    
    cov_ann_np = ann_cov.to_numpy(dtype = float)
  
    cov_week_np = weekly_cov.to_numpy(dtype = float)
    
    var_ann = np.einsum("ki,ij,kj->k", W, cov_ann_np, W, optimize = True)
  
    var_week = np.einsum("ki,ij,kj->k", W, cov_week_np, W, optimize = True)
    
    vols_ann = {
        label: float(np.sqrt(max(var_ann[i], 0.0)))
        for i, label in enumerate(selected_labels)
    }
    
    vols_weekly = {
        label: float(np.sqrt(max(var_week[i], 0.0)))
        for i, label in enumerate(selected_labels)
    }
    
    
    var = pd.Series(np.diag(ann_cov), index = ann_cov.index)
    
    std = np.sqrt(var).clip(lower = MIN_STD, upper = MAX_STD)
    
    weights_cols = {label: selected_weights[label] for label in selected_labels}
    
    weights_df = pd.DataFrame({
        **weights_cols,
        "Expected Return": comb_rets,
        "Vol": std,
        "Score": comb_score,
    })
    
    weights = weights_df.reindex(tickers)
    
    cols_to_mult = [
        col for col in weights.columns
        if col not in ["Expected Return", "Vol", "Score", "Sector", "Industry"]
    ]
    
    weights[cols_to_mult] = weights[cols_to_mult] * money_in_portfolio
    
    portfolio_weights = weights.reset_index().rename(columns = {"index": "Ticker"})
    
    weight_cols = [
        c for c in portfolio_weights.columns
        if c not in ["Ticker", "Expected Return", "Vol", "Score", "Sector", "Industry"]
    ]
    
    metric_cols = ["Expected Return", "Vol", "Score"]

    portfolio_weights_base = portfolio_weights.drop(
        columns = ["Industry", "Sector"],
        errors = "ignore"
    )

    ticker_ind_s = pd.Series(ticker_ind, copy = False).rename("Industry")
   
    ticker_ind_s.index = ticker_ind_s.index.astype(str)
    
    portfolio_weights_base["Ticker"] = portfolio_weights_base["Ticker"].astype(str)
    
    portfolio_weights_with_ind = portfolio_weights_base.merge(
        ticker_ind_s,
        left_on = "Ticker",
        right_index = True,
        how = "left"
    )
    
    if "Industry" not in portfolio_weights_with_ind.columns:
        
        portfolio_weights_with_ind["Industry"] = "Unknown"
    
    portfolio_weights_with_ind["Industry"] = portfolio_weights_with_ind["Industry"].fillna("Unknown")
    
    g_ind = portfolio_weights_with_ind.groupby("Industry", dropna = False)

    industry_weights = g_ind[weight_cols].sum()

    industry_weights_pct = industry_weights.div(money_in_portfolio).mul(100)

    industry_metrics_mean = g_ind[metric_cols].mean()

    industry_breakdown_percent = (pd.concat([industry_weights_pct, industry_metrics_mean], axis = 1).reset_index())

    ticker_sec_s = pd.Series(ticker_sec, copy = False).rename("Sector")
  
    ticker_sec_s.index = ticker_sec_s.index.astype(str)
    
    portfolio_weights_with_sec = portfolio_weights_base.merge(
        ticker_sec_s,
        left_on = "Ticker",
        right_index = True,
        how = "left"
    )
    
    if "Sector" not in portfolio_weights_with_sec.columns:
        
        portfolio_weights_with_sec["Sector"] = "Unknown"
    
    portfolio_weights_with_sec["Sector"] = portfolio_weights_with_sec["Sector"].fillna("Unknown")
    
    g_sec = portfolio_weights_with_sec.groupby("Sector", dropna = False)

    sector_weights = g_sec[weight_cols].sum()
    
    sector_weights_pct = sector_weights.div(money_in_portfolio).mul(100)

    sector_metrics_mean = g_sec[metric_cols].mean()

    sector_breakdown_percent = (
        pd.concat([sector_weights_pct, sector_metrics_mean], axis = 1)
        .reset_index()
    )


    logging.info("Generating portfolio performance reports...")
    
    report_name_map = {
        "MSR": "MSR",
        "Sortino": "Sortino",
        "BL": "Black-Litterman",
        "MIR": "MIR",
        "GMV": "GMV",
        "MDP": "MDP",
        "MSP": "MSP",
        "Deflated_MSR": "Deflated MSR",
        "Adjusted_MSR": "Adjusted MSR",
        "UPM-LPM": "UPM-LPM",
        "Upside/Downside": "Up/Down Cap",
        "Combination": "Combination",
        "Combination1": "Combination1",
        "Combination2": "Combination2",
        "Combination3": "Combination3",
        "Combination4": "Combination4",
        "Combination5": "Combination5",
        "Combination6": "Combination6",
        "Combination7": "Combination7",
        "Combination8": "Combination8",
        "Combination9": "Combination9",
        "Combination10": "Combination10",
        "Combination11": "Combination11",
        "Combination12": "Combination12",
        "Combination13": "Combination13",
    }
    
    report_weights = {
        report_name_map[label]: selected_weights[label].to_numpy(dtype = float)
        for label in selected_labels
    }
    
    report_vols_weekly = {
        report_name_map[label]: vols_weekly[label]
        for label in selected_labels
    }
    
    report_vols_annual = {
        report_name_map[label]: vols_ann[label]
        for label in selected_labels
    }
    
    performance_df = pa.report_portfolio_metrics_batch(
        weights = report_weights,
        vols_weekly = report_vols_weekly,
        vols_annual = report_vols_annual,
        comb_rets = comb_rets,
        bear_rets = bear_rets,
        bull_rets = bull_rets,
        comb_score = comb_score,
        last_year_weekly_rets = last_year_weekly_rets,
        last_5y_weekly_rets = last_5_year_weekly_rets,
        n_last_year_weeks = n_last_year_weeks,
        rf_rate = rf_rate,
        beta = beta,
        benchmark_weekly_rets = last_year_benchmark_weekly_rets,
        benchmark_ret = benchmark_ret,
        mu_bl = mu_bl,
        sigma_bl = sigma_bl,
        d_to_e = d_to_e,
        tax = tax,
    )
        
    ticker_performance = pa.report_ticker_metrics(
        tickers = tickers, 
        last_year_weekly_rets = last_year_weekly_rets,
        last_5y_weekly_rets = last_5_year_weekly_rets,
        n_last_year_weeks = n_last_year_weeks,
        weekly_cov = weekly_cov, 
        ann_cov = ann_cov, 
        comb_rets = comb_rets,
        bear_rets = bear_rets,
        bull_rets = bull_rets,
        comb_score = comb_score,
        rf = rf_rate,
        beta = beta,
        benchmark_weekly_rets = benchmark_weekly_rets,
        benchmark_ann_ret = benchmark_ret,
        bl_ret = mu_bl,
        bl_cov = sigma_bl,
        tax = tax,
        d_to_e = d_to_e,
        forecast_file = config.FORECAST_FILE
    )
    
    logging.info("Writing results to Excel...")
    
    ticker_performance_df = pd.DataFrame(ticker_performance)
    
    sheets_to_write = {
        "Ticker Performance": ticker_performance_df,
        "Today_Buy_Sell": buy_sell_today_df,
        "Covariance": cov_df,
        "Covariance Description": ann_cov_desc,
        "Bounds": bounds_df,
        "Weights": portfolio_weights,
        "Industry Breakdown": industry_breakdown_percent,
        "Sector Breakdown": sector_breakdown_percent,
        "Portfolio Performance": performance_df,
    }

    write_excel_results(
        excel_file = config.PORTFOLIO_FILE, 
        sheets = sheets_to_write
    )
  
    logging.info("Upload Complete")


if __name__ == "__main__":
    
    main()
