"""
Writes various DataFrames to an Excel workbook and applies conditional formatting and tables.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import datetime as dt

import config


def export_results(sheets, output_excel_file=None):
    
    if not output_excel_file:
        output_excel_file = config.FORECAST_FILE
        
    with pd.ExcelWriter(
        output_excel_file,
        mode='a' if os.path.exists(output_excel_file) else 'w',
        engine='openpyxl',
        **({'if_sheet_exists':'replace'} if os.path.exists(output_excel_file) else {})
    ) as writer:
        
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=True)

    wb = load_workbook(output_excel_file)
    
    red_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9B', end_color='FFEB9B', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    for sheet_name, df in sheets.items():
      
        ws = wb[sheet_name]  

        for cell in ws[1]:
            if isinstance(cell, MergedCell):
                continue

            if not isinstance(cell.value, str):
                cell.value = "" if cell.value is None else str(cell.value)
     
        max_row = ws.max_row
        max_col = ws.max_column

        header_map = {
            cell.value: get_column_letter(cell.column)
            for cell in ws[1]
            if cell.value is not None
        }

        cp_col  = header_map.get('Current Price')
        low_col = header_map.get('Low Price')
        avg_col = header_map.get('Avg Price')
        med_col = header_map.get('Median Price')
        high_col = header_map.get('High Price')
        ret_col = header_map.get('Returns')
        low_ret_col = header_map.get('Low Returns')
        high_ret_col = header_map.get('High Returns')
        score_col = header_map.get('Score')
        buy_col = header_map.get('Buy')
        sell_col = header_map.get('Sell')
        msr_col = header_map.get('MSR')
        sortino_col = header_map.get('Sortino')
        mir_col = header_map.get('MIR')
        comb_col = header_map.get('Combination')
                
        if ret_col:
            
            ws.conditional_formatting.add(
                f"{ret_col}2:{ret_col}{max_row}",
                CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{ret_col}2:{ret_col}{max_row}",
                CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
            )
        
        if low_ret_col:
            
            ws.conditional_formatting.add(
                f"{low_ret_col}2:{low_ret_col}{max_row}",
                CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{low_ret_col}2:{low_ret_col}{max_row}",
                CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
            )
        
        if high_ret_col:
            
            ws.conditional_formatting.add(
                f"{high_ret_col}2:{high_ret_col}{max_row}",
                CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{high_ret_col}2:{high_ret_col}{max_row}",
                CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
            )
            
        if avg_col:

            fr = f"{avg_col}2<{cp_col}2"
            fg = f"{avg_col}2>{cp_col}2"
            
            ws.conditional_formatting.add(
                f"{avg_col}2:{avg_col}{max_row}",
                FormulaRule(formula=[fr], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{avg_col}2:{avg_col}{max_row}",
                FormulaRule(formula=[fg], fill=green_fill)
            )
            
        if med_col:
            
            fr = f"{med_col}2<{cp_col}2"
            fg = f"{med_col}2>{cp_col}2"
            
            ws.conditional_formatting.add(
                f"{med_col}2:{med_col}{max_row}",
                FormulaRule(formula=[fr], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{med_col}2:{med_col}{max_row}",
                FormulaRule(formula=[fg], fill=green_fill)
            )
            
        if low_col:
            
            fr = f"{low_col}2<{cp_col}2"
            fg = f"{low_col}2>{cp_col}2"
            
            ws.conditional_formatting.add(
                f"{low_col}2:{low_col}{max_row}",
                FormulaRule(formula=[fr], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{low_col}2:{low_col}{max_row}",
                FormulaRule(formula=[fg], fill=green_fill)
            )
        
        if high_col:
            
            fr = f"{high_col}2<{cp_col}2"
            fg = f"{high_col}2>{cp_col}2"
            
            ws.conditional_formatting.add(
                f"{high_col}2:{high_col}{max_row}",
                FormulaRule(formula=[fr], fill=red_fill)
            )
            
            ws.conditional_formatting.add(
                f"{high_col}2:{high_col}{max_row}",
                FormulaRule(formula=[fg], fill=green_fill)
            )
        
        if score_col:
            
            ws.conditional_formatting.add(
                f"{score_col}2:{score_col}{max_row}",
                CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            )           
            
            ws.conditional_formatting.add(
                f"{score_col}2:{score_col}{max_row}",
                CellIsRule(operator='between', formula=['-10', '0'], fill=red_fill)
            )  
            
            ws.conditional_formatting.add(
                f"{score_col}2:{score_col}{max_row}",
                CellIsRule(operator='greaterThan', formula=['9'], fill=green_fill)
            )  
            
            ws.conditional_formatting.add(
                f"{score_col}2:{score_col}{max_row}",
                CellIsRule(operator='between', formula=['1', '5'], fill=orange_fill)
            )           
             
            ws.conditional_formatting.add(
                f"{score_col}2:{score_col}{max_row}",
                CellIsRule(operator='between', formula=['6', '9'], fill=yellow_fill)
            )      
        
        for col in (buy_col, sell_col):
           
            if not col:
                continue  
            
            rng = f"{col}2:{col}{max_row}"
            
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='equal', formula=['TRUE'],  fill=green_fill)
            )
            
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill)
            )

        for col in (msr_col, sortino_col, mir_col, comb_col):
            
            if not col:
                continue
            
            rng = f"{col}2:{col}{max_row}"
            
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan",   formula=['0.01'], fill=red_fill,  stopIfTrue=True)
            )
            
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='greaterThan', formula=['0'],    fill=green_fill)
            )
            
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="between",     formula=['1.95','2.05'], fill=yellow_fill, stopIfTrue=True)
            )
            
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f"AND({col}2<>0,{col}2<>2)"], fill=green_fill)
            )
        
        cov_name = 'Covariance' 
        
        if cov_name in sheets:
            
            cov_df = sheets[cov_name]
            cov_ws = wb[cov_name]

            data = cov_df.values.astype(float)
            min_val, mean_val, max_val = data.min(), data.mean(), data.max()

            color_rule = ColorScaleRule(
                start_type='num', start_value=str(min_val), start_color='FFFFFF',
                mid_type='num',   mid_value=str(mean_val), mid_color='FFDD99',
                end_type='num',   end_value=str(max_val), end_color='FF0000'
            )

            rng = f"B2:{get_column_letter(cov_ws.max_column)}{cov_ws.max_row}"
            
            cov_ws.conditional_formatting.add(rng, color_rule)            
          
        last_col_letter = get_column_letter(max_col)
        
        table = Table(
            displayName=ws.title.replace(" ", "") + "Table",
            ref=f"A1:{last_col_letter}{max_row}"
        )
        
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        
        ws.add_table(table)

    wb.save(output_excel_file)
    wb.close()
    
    wb.close()
    
