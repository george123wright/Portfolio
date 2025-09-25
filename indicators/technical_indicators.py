from __future__ import annotations

"""
Indicator-driven signal calibration, backtesting, and composite scoring.

Overview
--------
This module implements a complete pipeline for:

1) loading OHLCV data from an Excel workbook,

2) precomputing reusable indicator inputs,

3) grid-search calibration of indicator parameters per ticker subject to
   statistical significance gates,

4) converting indicator "pulses" into persistent positions with explicit
   rules for signal lag, cooldown, and pulse conflicts,

5) evaluating performance (Sharpe ratio, t-statistic, and one-sided p-value),

6) assigning integer weights to statistically defensible indicator
   configurations relative to buy-and-hold performance, and

7) producing a time-indexed composite integer "signal score" per ticker that
   aggregates weighted positions across indicators.

Data Model
----------
Input data are read from a workbook with sheets "High", "Low", "Close",
and "Volume". Each sheet is a wide, sheet-style table (columns are tickers,
rows are dates) and is parsed into a `pd.DataFrame` indexed by timestamp.
The code enforces a strictly ascending DatetimeIndex and converts timezone-
aware timestamps to tz-naïve UTC for Excel interoperability.

Precomputation and Gating
-------------------------
For each ticker a `Precomp` structure holds NumPy arrays of prices, returns,
and indicator precomputations to avoid repeated O(T) passes. Two boolean
masks gate signal generation:

- Trend gate: `trend_mask_t = 1{ ADX_t >= ADX_ENTRY }` with ADX computed via
  the standard DI+/DI− construction. This is applied to pro-trend indicators
  (MACD, EMA, ATR, VWAP, OBV) and negated for contra-trend indicators (RSI,
  Bollinger Bands, Stochastic). Neutral indicators (MFI) are ungated by ADX.

- Volume gate: `vol_ok_t = 1{ Volume_t >= median(Volume_{t-19:t}) }`
  (20-bar rolling median), applied to indicators that require robust volume
  context (ATR, VWAP, OBV, MACD, EMA).

Only timestamps where the relevant gate is True can emit pulses.

Indicators and Signal Pulses
----------------------------
Each indicator maps to a buy/sell pulse construction that is then converted
to a persistent position with `pulses_to_positions` (see below).

1) Relative Strength Index (RSI; anti-trend):
   
   - RSI is the classic Wilder oscillator computed over a window `w`:
   
     Let 
     
        `u_t = max(C_t − C_{t−1}, 0)`, 
        
        `d_t = max(C_{t−1} − C_t, 0)`.
        
     Smoothed averages `U_t` and `D_t` are typically EMA/RMA of `u_t` and `d_t`.
     
        `RS_t = U_t / D_t` and `RSI_t = 100 − (100 / (1 + RS_t))`.
   
   - Pulses:
     buy at an upward cross of the buy threshold `b`: 
     
        `RSI_t > b` 
    
        `RSI_{t−1} <= b`. 
    
    sell at a downward cross of `s`: 
    
        `RSI_t < s` 
        
        `RSI_{t−1} >= s`. 
        
    Typical `b < 50 < s` with `b` low (oversold) and `s` high (overbought).

2) Money Flow Index (MFI; neutral):
   
   - Typical price 
   
        `TP_t = (H_t + L_t + C_t) / 3`.
    
    Money flow 
    
        `MF_t = TP_t * Volume_t`.
    
    Over a window `w`, positive/negative sums are
    
        `MF_plus = Σ MF_t * 1{TP_t > TP_{t−1}}`, 
        
        `MF_minus = Σ MF_t * 1{TP_t < TP_{t−1}}`.
    
    The money flow ratio `R_t = MF_plus / MF_minus`, `MFI_t = 100 − 100 / (1 + R_t)`.
   
   - Pulses: upward cross of `b` for buy; downward cross of `s` for sell.

3) Bollinger Bands (BB; anti-trend):
   
   - Midline and dispersion over window `w`:
        
        `mid_t = SMA_w(C)_t`, `sd_t = std_w(C)_t` (population `ddof=0`).
    
    Upper and lower bands: 
    
        `upper_t = mid_t + nσ * sd_t`,
    
        `lower_t = mid_t − nσ * sd_t`, 
    
    where `nσ` is the standard-deviation multiplier.
   
   - Pulses: buy on cross above `lower_t`; sell on cross below `upper_t`.

4) Average True Range breakout (ATR; pro-trend):
  
   - True Range 
   
        `TR_t = max( H_t − L_t, |H_t − C_{t−1}|, |L_t − C_{t−1}| )`.
     
        ATR is a rolling mean of TR over `w_atr`.
   
   - Price breakout thresholds from prior extrema over window `w_break`:
     
     `hi_prev_t = max(H_{t−w_break : t−1})`, `lo_prev_t = min(L_{t−w_break : t−1})`.
     
     Buy threshold: 
     
        `up_th_t = hi_prev_t + m * ATR_{t−1}`.
     
     Sell threshold: 
        
        `dn_th_t = lo_prev_t − m * ATR_{t−1}`.
  
   - Pulses:
   
        buy if `C_t > up_th_t`
        
        sell if `C_t < dn_th_t`.

5) Stochastic oscillator (anti-trend):
   
   - `%K_t = 100 * (C_t − min(L_{t−k+1:t})) / (max(H_{t−k+1:t}) − min(L_{t−k+1:t}))`.
     `%D_t = SMA_d(%K)_t`.
     
     Buy is conditioned on oversold, sell on overbought:
        
        buy if `%K_t < L` *and* `%K` crosses above `%D`;
     
        sell if `%K_t > U` *and* `%K` crosses below `%D`.

6) EMA crossover (pro-trend):
   - Exponentially weighted moving averages `EMA_f`, `EMA_s` with `f < s`.
    
     Buy on cross up:
     
        `EMA_f` crosses above `EMA_s`.
     
     Sell on cross down:
     
        `EMA_f` crosses below `EMA_s`.

7) MACD (pro-trend):
   - MACD line 
   
        `M_t = EMA_f(C)_t − EMA_s(C)_t`,
        
    signal 
    
        `S_t = EMA_g(M)_t`.
     
     Buy on `M` crossing above `S`
     
     sell on `M` crossing below `S`.

8) Rolling VWAP (pro-trend):
    
    - Windowed VWAP over `w` bars:
    
         `VWAP_t = (Σ_{i=t−w+1..t} TP_i * V_i) / (Σ_{i=t−w+1..t} V_i)`,
     
     with efficient computation via cumulative sums.
  
   - Pulses: buy on cross above VWAP; sell on cross below VWAP.

9) On-Balance Volume divergence (pro-trend entry on bullish divergence, exit
   on bearish divergence):
   
   - OBV cumulative series: 
   
        `OBV_t = OBV_{t−1} + sign(C_t − C_{t−1}) * V_t`
     
     (0 contribution if `C_t == C_{t−1}`).
   
   - Over lookback `w`, detect divergence vs prior price/OBV extrema:
     
        bullish: 
        
            `C_t < min(C_{t−w:t−1})` but `OBV_t >= min(OBV_{t−w:t−1})`;
     
        bearish: 
        
            `C_t > max(C_{t−w:t−1})` but `OBV_t <= max(OBV_{t−w:t−1})`.
  
   - Pulses: buy on bullish divergence; sell on bearish divergence.

Pulse → Position Mapping
------------------------
Pulses are converted to persistent positions with a Numba kernel that enforces:

- Signal lag `L`: a pulse at `t` affects the position only from `t + L`.

- Cooldown `K`: after any flip, additional pulses are ignored for the next `K`
  bars.

- Conflict mode when buy and sell are simultaneous at the effective time:
  "sell_wins" (default), "buy_wins", or "mutual_exclude" (drop both).

The resulting position series `pos_t ∈ {−1, 0, +1}` persists until flipped.

Backtest Alignment and Statistics
---------------------------------
Returns use simple arithmetic daily returns

`r_t = (C_t / C_{t−1}) − 1`.

To avoid look-ahead:

    `X_t = pos_{t−1} * r_t` is the realised PnL proxy.
    
Only valid (non-NaN) returns are included.

For `n` valid observations,

- mean return: 

    `mu = (1/n) * Σ X_t`,

- standard deviation: 

    `sd = sqrt( (1/(n−1)) * Σ (X_t − mu)^2 )`,

- annualised Sharpe: 
    
    `Sharpe = sqrt(A) * (mu / sd)` with
  
  `A = 252` trading days,
  
- t-statistic: 

    `t = mu / (sd / sqrt(n))`,
    
- one-sided p-value for `H1: mu > 0`:

    `p = t_sf(max(t, 0), df = n−1)`.

Calibration and Weighting
-------------------------
Each indicator is tuned by a grid search over its parameter space. A coarse
screen can be applied, followed by full or "local" refinement around the
best coarse parameters. A configuration is considered significant if:
`p_value < SIGNIFICANCE_ALPHA` and `Sharpe >= MIN_SHARPE_FOR_SIG`.

Relative weight assignment compares the calibrated Sharpe to buy-and-hold
Sharpe; the improvement `Δ = Sharpe_strategy − Sharpe_BH` maps to an integer
weight in {0, 1, 2, 3, 4} by thresholding
(≥ 1.00 → 4, ≥ 0.50 → 3, ≥ 0.25 → 2, > 0 → 1; else 0).

Composite Score
---------------
Per-bar composite score is the integer sum of `weight_indicator * position_indicator`
over all indicators and is finally clamped to `±SCORE_CLAMP`. This score
represents weighted agreement among active indicators after gating and
conflict resolution; it is *not* an expected return.

Outputs
-------
- Calibration results are upserted to an Excel sheet ("Calibration") with
  parameters, weights, Sharpe, t-statistic, p-value, number of flips, and
  period metadata.

- Composite scores are written to a sheet ("Signal Scores") with conditional
  colour formatting (green for positive, red for negative).

Assumptions and Limitations
---------------------------
- The framework uses daily bars and assumes 252 trading days per year.

- All statistics are in-sample and ignore transaction costs, slippage,
  borrowing costs, and market impact.

- Significance testing is simple one-sided Student t without multiple-
  hypothesis correction across grid points.

- Indicators are combined linearly; dependencies between indicators are not
  explicitly modelled.
"""

import logging
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List

import numpy as np
import pandas as pd
from scipy import stats
import ta
import datetime as dt

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from numba import njit, prange

from math import erfc

import json
import config


UPDATE_PARAM: bool = False

CONFLICT_MODE: str = "sell_wins"  

SIGNIFICANCE_ALPHA = 0.05
MIN_SHARPE_FOR_SIG = 0.30

USE_COARSE_SCREEN: bool = False           

REFINE_STRATEGY: str = "full"           
LOCAL_WIDTH: int = 1                   

SCREEN_ALPHA: float = SIGNIFICANCE_ALPHA    
SCREEN_MIN_SHARPE: float = MIN_SHARPE_FOR_SIG

COARSE_GRIDS = {
    "rsi": {
        "buy": [25, 30, 35], 
        "sell": [65, 75, 85]
    },
    "mfi": {
        "buy": [20, 25, 30], 
        "sell": [75, 80, 85]
    },
    "bb": {
        "num_std": [1.5, 2.0, 2.5]
    },
    "atr": {
        "mult": [1.25, 1.5, 2.0, 2.5]
    },
    "stoch": {
        "lower": [15, 20, 25], 
        "upper": [75, 80, 85]
    },
    "ema": {
        "fast": [8, 12, 16],
        "slow": [20, 26, 34]
    },
    "macd": {
        "fast": [8, 12, 16],
        "slow": [20, 26, 34],
        "signal": [6, 9, 12]
    },
    "vwap": {
        "window": [10, 20, 30]
    },
    "obv": {
        "lookback": [15, 20, 30]
    },
}

try:

    CALIBRATION_FILE = Path(config.CALIBRATION_FILE)

except Exception:

    CALIBRATION_FILE = Path("indicator_calibration.xlsx")

CALIB_SHEET = "Calibration"

ANNUALISATION = 252

RSI_BUY_GRID = list(np.arange(1, 49, 0.5))
RSI_SELL_GRID = list(np.arange(51, 100, 0.5))
RSI_WINDOW_GRID = list(range(5, 31))

MFI_BUY_GRID = list(np.arange(1, 49, 0.5))
MFI_SELL_GRID = list(np.arange(51, 100, 0.5))
MFI_WINDOW_GRID = list(range(5, 31))

BB_STD_GRID = list(np.arange(0.75, 3, 0.05))
BB_WINDOW_GRID = list(range(5, 31))

ATR_MULT_GRID = list(np.arange(0.75, 3, 0.05))
ATR_WINDOW_GRID = list(range(5, 31))
ATR_BREAK_GRID = list(range(10, 51, 2))

STOCH_LOW_GRID = list(np.arange(1, 49, 0.5))
STOCH_HIGH_GRID = list(np.arange(51, 100, 0.5))

EMA_FAST_GRID = list(range(5, 21))          
EMA_SLOW_GRID = list(range(20, 61, 2))

MACD_FAST_GRID = list(range(7, 16))
MACD_SLOW_GRID = list(range(20, 34))
MACD_SIGNAL_GRID = list(range(6, 12))

VWAP_WIN_GRID = list(range(5, 40))

OBV_LOOKBACK_SCAN = list(range(5, 40))

MODE_SELL_WINS = 0
MODE_BUY_WINS = 1
MODE_MUTUAL = 2

_MODE_MAP = {
    "sell_wins": MODE_SELL_WINS, 
    "buy_wins": MODE_BUY_WINS, 
    "mutual_exclude": MODE_MUTUAL
}

DEFAULT_WEIGHTS = {
    "macd": 1,
    "ema": 1, 
    "atr": 1,
    "rsi": 1, 
    "bb": 1,
    "stoch": 1, 
    "obv": 1,
    "mfi": 1, 
    "vwap": 1
}

PRO_TREND = {"macd", "ema", "atr", "vwap", "obv"}

ANTI_TREND = {"rsi", "bb", "stoch"}

NEUTRAL_TREND = {"mfi"}

VOL_REQUIRED = {"atr", "vwap", "obv", "macd", "ema"}

INDICATOR_ORDER = ["macd", "ema", "atr", "rsi", "bb", "stoch", "obv", "mfi", "vwap"]

EMA_FAST: int = 12
EMA_SLOW: int = 26

BB_STD = 2
BB_WINDOW: int = 20

STOCH_K: int = 14
STOCH_D: int = 3

ATR_WINDOW: int = 14
ATR_BREAK_WINDOW: int = 20
ATR_MULTIPLIER: float = 1.5  

OBV_LOOKBACK: int = 20

ADX_WINDOW: int = 14
ADX_ENTRY: float = 25.0
ADX_EXIT: float = 20.0 

RSI_WINDOW: int = 14
RSI_BUY_THRESH: float = 30.0
RSI_SELL_THRESH: float = 70.0

VWAP_WINDOW: int = 20

MFI_WINDOW: int = 14
MFI_BUY_THRESH: float = 20.0
MFI_SELL_THRESH: float = 80.0

SCORE_CLAMP: int = 10

SIGNAL_LAG: int = 0
COOLDOWN_BARS: int = 0

CALIB_COLS = [
    "Ticker", "Indicator", "ParamJSON", "Weight",
    "Sharpe", "SharpeBH", "TStat", "PValue", "NTrades",
    "Start", "End", "LastUpdated"
]


logger = logging.getLogger(__name__)

logger.setLevel(logging.INFO)

if not logger.handlers:

    _h = logging.StreamHandler()

    _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

    logger.addHandler(_h)


def _col(
    df: pd.DataFrame, 
    tk: str, 
    name: str
) -> pd.Series:
    """
    Return one ticker column from a wide, sheet-style DataFrame with strict validation.

    Parameters
    ----------
    df : pd.DataFrame
        A 2-D table whose columns are tickers and whose rows are dates.
    tk : str
        Ticker symbol to extract, e.g. "AAPL".
    name : str
        Human-readable sheet name for error messages (e.g. "Close").

    Returns
    -------
    pd.Series
        The column `df[tk]`. The index, dtype, and metadata are preserved.

    Raises
    ------
    KeyError
        If `tk` is not present in `df.columns`.

    Rationale
    ---------
    Downstream indicator and scoring logic assumes that the input series exist and
    are properly aligned across High/Low/Close/Volume sheets. This guard converts a
    latent data quality issue (missing ticker) into an explicit failure with a clear
    message naming the offending sheet.
    """

    if tk not in df.columns:
    
        raise KeyError(f"{name} sheet does not contain ticker '{tk}'")
    
    return df[tk]


def read_sheet(
    excel_file: str | Path, 
    sheet: str
) -> pd.DataFrame:
    """
    Read a single Excel sheet as a DateTime-indexed DataFrame, sorted by index.

    Parameters
    ----------
    excel_file : str | Path
        Path to the workbook.
    sheet : str
        Sheet name to read.

    Returns
    -------
    pd.DataFrame
        DataFrame indexed by datetime (parsed with `parse_dates=True`).

    Raises
    ------
    ValueError
        If the sheet’s index is not datetime-typed.
    Exception
        Any read errors are logged and re-raised.

    Notes
    -----
    The function enforces a datetime index because subsequent indicator and
    backtest logic assumes time-ordered data. The frame is sorted ascending
    by index to guarantee time order.
    """

    try:

        df = pd.read_excel(excel_file, sheet_name=sheet, index_col = 0, header = 0, parse_dates = True)

    except Exception as exc:

        logger.exception("Parsing %s!%s failed (%s)", excel_file, sheet, type(exc).__name__)

        raise

    if not pd.api.types.is_datetime64_any_dtype(df.index):

        raise ValueError(f"Sheet '{sheet}' has non-datetime index")

    return df.sort_index()


def load_data(
    excel_file: str | Path
) -> Dict[str, pd.DataFrame]:
    """
    Load OHLCV data from an Excel workbook into a dict of DataFrames.

    Parameters
    ----------
    excel_file : str | Path
        Path to workbook containing sheets: "High", "Low", "Close", "Volume".

    Returns
    -------
    Dict[str, pd.DataFrame]
        Keys: "high", "low", "close", "volume". Values are DateTime-indexed frames.

    Notes
    -----
    This function calls `read_sheet` for each sheet and provides a consistent
    mapping of logical names → sheet names expected downstream.
    """

    mapping = {
        "high": "High",
        "low": "Low", 
        "close": "Close", 
        "volume": "Volume"
    }
    
    out: Dict[str, pd.DataFrame] = {}
    
    for key, sheet in mapping.items():
    
        out[key] = read_sheet(
            excel_file = excel_file,
            sheet = sheet
        )
    
    return out


def _excel_safe_datetimes(
    df: pd.DataFrame
) -> pd.DataFrame:
    """
    Return a copy with timezone-aware datetimes converted to tz-naive UTC.

    Parameters
    ----------
    df : pd.DataFrame
        Arbitrary DataFrame that may contain datetime-like columns or objects.

    Returns
    -------
    pd.DataFrame
        Copy of `df` with:
        - tz-aware pandas datetime columns converted to UTC then made tz-naive.
        - object columns that contain datetime/timestamp objects coerced similarly.

    Why
    ---
    Excel (via openpyxl) cannot store tz-aware datetimes. This function ensures
    all datetimes are timezone-naive (interpreted as UTC) to avoid
    `ValueError: Excel does not support datetimes with timezones`.

    Details
    -------
    For pandas datetime columns with dtype `DatetimeTZDtype`, we apply:
        ts_utc_naive = ts.dt.tz_convert("UTC").dt.tz_localize(None)

    For object columns, each element `x` is mapped:
    - if `x` is pd.Timestamp with tz: `x.tz_convert("UTC").tz_localize(None)`
    - if `x` is datetime with tzinfo: `x.astimezone(UTC).replace(tzinfo=None)`
    - otherwise left unchanged.
    """
   
    out = df.copy()
   
    DatetimeTZDtype = getattr(pd, "DatetimeTZDtype", None)
   
    for col in out.columns:
   
        s = out[col]
   
        dtype = s.dtype
   
        is_tzaware = (
            (DatetimeTZDtype is not None and isinstance(dtype, DatetimeTZDtype))
            or getattr(dtype, "tz", None) is not None
        )
       
        if is_tzaware:
       
            out[col] = s.dt.tz_convert("UTC").dt.tz_localize(None)
       
        elif pd.api.types.is_datetime64_dtype(dtype):
       
            pass
       
        elif pd.api.types.is_object_dtype(dtype):
       
       
            def _strip_tz(
                x
            ):
       
                if isinstance(x, pd.Timestamp):
            
                    return x.tz_convert("UTC").tz_localize(None) if x.tz is not None else x
            
                if isinstance(x, dt.datetime):
                    
                    if x.tzinfo is not None:
            
                        return x.astimezone(dt.timezone.utc).replace(tzinfo = None)  
                    
                    else:
                        
                        return x
            
                return x
            
            
            out[col] = s.map(_strip_tz)
   
    return out


def load_calibration_df() -> pd.DataFrame:
    """
    Load the calibration table from the calibration workbook (if present).

    Returns
    -------
    pd.DataFrame
        A DataFrame with columns:
        ["Ticker","Indicator","ParamJSON","Weight","Sharpe","TStat","PValue",
        "NTrades","Start","End","LastUpdated"].
        Empty frame if file/sheet is missing or upon read errors.

    Notes
    -----
    The sheet name is given by the module constant `CALIB_SHEET`.
    """
   
    if not CALIBRATION_FILE.exists():
   
        return pd.DataFrame(columns = CALIB_COLS)

    try:

        df = pd.read_excel(CALIBRATION_FILE, sheet_name = CALIB_SHEET, engine = "openpyxl")

        return df.reindex(columns = CALIB_COLS, fill_value = np.nan)
  
    except Exception:
  
        return pd.DataFrame(columns = CALIB_COLS)


def save_calibration_df(
    df: pd.DataFrame
) -> None:
    """
    Write the calibration DataFrame into the calibration workbook.

    Parameters
    ----------
    df : pd.DataFrame
        Calibration data. Columns will be reindexed to the canonical set.

    Notes
    -----
    All datetime columns are passed through `_excel_safe_datetimes` before write.
    The file is overwritten (mode="w") each time.
    """

    df = df.reindex(columns = CALIB_COLS)

    df = _excel_safe_datetimes(
        df = df
    )

    with pd.ExcelWriter(CALIBRATION_FILE, engine = "openpyxl", mode = "w", datetime_format = "yyyy-mm-dd hh:mm:ss") as w:
      
        df.to_excel(w, sheet_name = CALIB_SHEET, index = False)


def _is_sig(
    sharpe: float,
    p_value: float, 
    min_sharpe: float,
    alpha: float
) -> bool:
    """
    Determine whether a backtested configuration is statistically and economically significant.

    This gate is deliberately simple and conservative. It requires BOTH:
   
    (1) statistical significance of the mean daily return (one-sided) at level `alpha`,
   
    (2) an economic filter that the annualised Sharpe ratio meets or exceeds `min_sharpe`.

    Formally, let X_t denote the daily strategy return series produced by a given
    parameterisation (constructed as position_{t-1} × simple_return_t; see `_eval_stats_from_pos`).
    Let μ̂ and s denote the sample mean and standard deviation of {X_t} over n valid days.
    Let A be the annualisation factor (e.g., 252).

    - Annualised Sharpe ratio:
    
        Sharpe = sqrt(A) × ( μ̂ / s )

    - One-sided t-statistic for H₁: μ > 0:
    
        t = μ̂ / ( s / sqrt(n) )

    - One-sided p-value under a Student-t with df = n − 1:
    
        p = sf(max(t, 0), df = n−1)

    This function returns True iff:
    
        (p_value < alpha) AND (sharpe ≥ min_sharpe)

    Parameters
    ----------
    sharpe : float
        Annualised Sharpe ratio estimate based on daily returns.
    p_value : float
        One-sided p-value for the hypothesis H1: mean(daily strategy return) > 0.
    min_sharpe : float
        Minimum annualised Sharpe threshold required for practical relevance.
    alpha : float
        Statistical significance level for the one-sided t-test.

    Returns
    -------
    bool
        True if both the statistical and the Sharpe filters pass; False otherwise.

    Notes
    -----
    The function does not apply a multiple-testing correction nor a buy-and-hold
    comparison, although those concepts are used elsewhere in the module
    (e.g., weight setting). The intention is to gate obviously weak parameterisations
    before expensive refinement.
    """

    base_gate = (p_value < alpha) and (sharpe >= min_sharpe)

    return base_gate


def _neighbor_window(
    full_grid: List[float] | List[int], 
    center: float, 
    width: int = 1
) -> List[float]:
    """
    Select an index-based neighbourhood around `center` within a 1-D, ordered grid.

    The grid is treated as an ordered discrete set 
    
        G = {g_0, …, g_{m−1}}. 
        
    The function first identifies 
    
        i* = argmin_i |g_i − center|,
        
    then returns the slice
   
        G[max(0, i*−width) : min(m, i*+width+1)].

    This is used to define a “local refinement” region around a coarse optimum when
    REFINE_STRATEGY == "local", limiting the Cartesian search to nearby candidates and
    reducing the multiple-testing burden.
    
    Parameters
    ----------
    full_grid : list[float | int]
        Full, ordered grid of candidate values.
    center : float
        Centre value to locate in `full_grid`. If not exactly present, the closest
        element by absolute difference is used.
    width : int, default 1
        Number of neighbours to include on each side of the centre.

    Returns
    -------
    list[float]
        A contiguous slice of `full_grid` containing at most `2*width+1` values,
        preserving order.

    Edge Cases
    ----------
    If `center` lies beyond the edges, the slice is clipped at the ends; if
    `width == 0`, the result is the nearest grid point only.
    """

    arr = list(full_grid)
    
    try:
    
        i = arr.index(center)  
   
    except ValueError:

        i = int(np.argmin([abs(x - center) for x in arr]))
  
    lo = max(0, i - width)
  
    hi = min(len(arr), i + width + 1)
  
    return arr[lo: hi]


def load_calibration_map() -> dict[str, dict[str, dict]]:
    """
    Load the saved calibration table and convert it into a nested dictionary.

    Output Schema
    -------------
    Returns a mapping:
        {
        "<Ticker>": {
            "<indicator>": {
                "params": dict,       # parsed from ParamJSON
                "weight": int,        # 0..4
                "sharpe": float,      # annualised
                "t_stat": float,      # daily mean t-statistic
                "p_value": float,     # one-sided, mean > 0
                "n_trades": int       # number of position flips
            },
            ...
        },
        ...
        }

    Notes
    -----
    - Missing or malformed JSON is safely treated as an empty parameter dict.
    
    - Indicator keys are normalised to lower case to ensure consistent lookups.
    
    - All numeric fields are coerced to their primitive types with sensible defaults.
    """
    
    df = load_calibration_df()
   
    out: dict[str, dict[str, dict]] = {}
   
    for _, row in df.iterrows():
   
        tk = str(row["Ticker"])
   
        raw = row.get("ParamJSON")
   
        s = "" if pd.isna(raw) else str(raw)
   
        try:
   
            params = json.loads(s) if s else {}
   
        except Exception:
   
            params = {}
   
        out.setdefault(tk, {})[str(row["Indicator"]).lower()] = {
            "params": params,
            "weight": int(row.get("Weight", 0) or 0),
            "sharpe": float(row.get("Sharpe", 0.0) or 0.0),
            "t_stat": float(row.get("TStat", 0.0) or 0.0),
            "p_value": float(row.get("PValue", 1.0) or 1.0),
            "n_trades": int(row.get("NTrades", 0) or 0),
        }
        
    return out


def write_calibration_rows(
    rows: list[dict]
) -> None:
    """
    Upsert calibration rows into the workbook by (Ticker, Indicator) key.

    Process
    -------
    1) Load the existing calibration DataFrame, if any.
    
    2) Replace existing rows that share the same composite key (Ticker, Indicator).
    
    3) Append new rows for previously unseen keys.
    
    4) Persist the result via `save_calibration_df`, after normalising columns and
    datetime types.

    Parameters
    ----------
    rows : list[dict]
        Each dict is expected to contain the canonical calibration columns. Missing
        columns are introduced with NaNs; extra keys are ignored.

    Idempotency
    -----------
    Running the function multiple times with the same `rows` is idempotent with
    respect to (Ticker, Indicator), ensuring reproducible persistence.

    Timezone Handling
    -----------------
    Any datetime columns are made timezone-naïve for Excel compatibility by
    `_excel_safe_datetimes`.
    """

    old = load_calibration_df()

    if not old.empty:

        old = _excel_safe_datetimes(
            df = old
        )

    new = pd.DataFrame(rows)
      
    new = new.reindex(columns = CALIB_COLS)

    if old.empty:
  
        df = new
  
    else:
        old = old.reindex(columns = CALIB_COLS)
  
        key = ["Ticker", "Indicator"]
       
        old_idx = old.set_index(key).index
       
        new_idx = new.set_index(key).index
       
        keep_old_mask = ~old_idx.isin(new_idx)

        pieces = []
       
        old_keep = old.loc[keep_old_mask]
       
        if not old_keep.empty:
       
            pieces.append(old_keep.reset_index(drop = True))
       
        pieces.append(new.reset_index(drop = True))
       
        df = pd.concat(pieces, ignore_index = True)

    save_calibration_df(
        df = df
    )


def save_scores(
    excel_file: str | Path, 
    scores: pd.DataFrame
) -> None:
    """
    Write the composite per-bar signal score to Excel with conditional formatting.

    The score sheet contains a 'Date' column followed by one column per ticker; cell
    values are integer scores in the range [−SCORE_CLAMP, +SCORE_CLAMP]. Conditional
    formatting shades > 0 green and < 0 red to aid scanning.

    Index handling
    --------------
    If the score index is timezone-aware, it is converted to UTC and made tz-naive,
    because Excel does not support timezone-aware datetimes.
    
    Parameters
    ----------
    excel_file : str | Path
        Target workbook.
    scores : pd.DataFrame
        DateTime-indexed DataFrame; columns are tickers; values are integer scores.

    Behavior
    --------
    - Creates/replaces a sheet named "Signal Scores".
   
    - Prepends a "Date" column derived from the index. If the index is timezone-aware,
    it is converted to UTC and made tz-naïve (Excel constraint).
   
    - Applies openpyxl conditional formatting:
        
        > 0 shaded green, < 0 shaded red.
        
    Why the Score is Integer-Valued
    -------------------------------
    Each per-indicator signal contributes a discrete position in {−1, 0, +1}
    multiplied by an integer weight. The aggregate is therefore integer-valued,
    and is finally clamped to a finite range to avoid outliers dominating any
    downstream use in dashboards.
    """

    excel_file = Path(excel_file)

    out = scores.copy()
    
    out.insert(0, "Date", out.index.tz_convert("UTC").tz_localize(None) if out.index.tz is not None else out.index)
    
    out = out.reset_index(drop=True)

    with pd.ExcelWriter(
        excel_file, 
        engine = "openpyxl", 
        mode = "a" if excel_file.exists() else "w",
        datetime_format = "yyyy-mm-dd hh:mm:ss"
    ) as w:
      
        try:
      
            book = w.book
        
            if "Signal Scores" in book.sheetnames:
        
                del book["Signal Scores"]
        
        except Exception:
        
            pass

        out.to_excel(w, sheet_name = "Signal Scores", index = False)

        ws = w.book["Signal Scores"]
       
        green = PatternFill("solid", start_color = "90EE90", end_color = "90EE90")
       
        red  =  PatternFill("solid", start_color = "FFC7CE", end_color = "FFC7CE")

        last_col = get_column_letter(
            col_idx = ws.max_column
        )
        
        rng = f"B2:{last_col}{ws.max_row}"
        
        ws.conditional_formatting.add(rng, CellIsRule(operator = "greaterThan", formula = ["0"], fill = green))
        
        ws.conditional_formatting.add(rng, CellIsRule(operator = "lessThan",  formula = ["0"], fill = red))


def _build_gate(
    pc: Precomp, 
    is_pro_trend: bool, 
    is_anti_trend: bool,
    vol_required: bool
) -> np.ndarray:
    """
    Construct a boolean “trading permission” mask per bar, enforcing trend and volume filters.

    Inputs
    ------
    - is_pro_trend: 
    
        if True, only permit bars for which the ADX-based trend mask is True.
   
    - is_anti_trend: 
    
        if True, only permit bars for which the trend mask is False.
   
    - vol_required: 
    
        if True, only permit bars that satisfy the volume sufficiency mask.

    The returned array `gate` (shape (T,)) acts as a multiplicative mask on all
    buy/sell pulse matrices, ensuring that signal generation is conditional on
    structural regimes (trend or non-trend) and liquidity/volume sufficiency.

    Notes
    -----
    - The trend mask is constructed from ADX ≥ ADX_ENTRY (with hysteresis if desired),
    hence it proxies “trending” regimes, while its complement proxies “mean-reverting”.
   
    - The volume mask requires the bar’s volume to be at least the rolling median of
    the last VOL_CONFIRM_WIN bars, mitigating microstructure noise in illiquid regimes.
    """
    
    gate = np.ones(pc.c.shape[0], dtype = np.bool_)

    if is_pro_trend:

        gate &= pc.trend_mask

    elif is_anti_trend:

        gate &= ~pc.trend_mask

    if vol_required:

        gate &= pc.vol_ok

    return gate


@dataclass
class IndicatorResult:
    """
    Container for the best indicator configuration and performance stats.

    Fields
    ------
    name : str
        Indicator identifier (e.g., "rsi").
    params : dict
        Chosen parameterisation (e.g., {"buy_thresh": 25.0, "sell_thresh": 75.0}).
    sharpe : float
        Annualised Sharpe (see `_eval_stats_from_pos` for the formula).
    t_stat : float
        One-sample t-statistic for mean daily return > 0.
    p_value : float
        One-sided p-value under Student-t with df = N-1.
    n_trades : int
        Estimated number of flips (position changes) during the backtest.
    """

    name: str
  
    params: dict
  
    sharpe: float
  
    t_stat: float
  
    p_value: float
    
    n_trades: int
    

def _mode_id(
    mode: str
) -> int:
    """
    Map the human-readable conflict mode to the integer code used by the Numba kernel.

    Conflict semantics when buy and sell pulses occur simultaneously at the effective time:
    - "sell_wins"      → 0 : prioritise exits/shorts over entries/longs
   
    - "buy_wins"       → 1 : prioritise entries/longs over exits/shorts
   
    - "mutual_exclude" → 2 : drop both pulses (no flip)

    This permits deterministic resolution within a single fused pass in `pulses_to_positions`.

    Parameters
    ----------
    mode : {"sell_wins","buy_wins","mutual_exclude"}
        Conflict resolution policy when buy and sell pulses occur simultaneously.

    Returns
    -------
    int
        0 for "sell_wins", 1 for "buy_wins", 2 for "mutual_exclude".

    Raises
    ------
    ValueError
        For unknown mode strings.
    """

    try:

        return _MODE_MAP[mode]

    except KeyError:

        raise ValueError(f"Unknown CONFLICT_MODE: {mode}")


def crosses_over(
    x: np.ndarray, 
    y: np.ndarray
) -> np.ndarray:
    """
    Detect cross-up events between two time-aligned arrays.

    Definition
    ----------
    At time t >= 1, a cross-up occurs if:
        
        (x_t > y_t) and (x_{t−1} <= y_{t−1}).
    
    Time t = 0 is defined as False. NaN comparisons follow NumPy rules and
    suppress signals where inputs are not available.

    Parameters
    ----------
    x, y : np.ndarray
        Arrays with a shared leading axis for time. Additional trailing dimensions
        are broadcast as per NumPy broadcasting to evaluate multiple series at once.

    Returns
    -------
    np.ndarray of bool
        Boolean array with the same broadcasted shape as x and y.
    """

    x_prev = np.roll(x, 1, axis = 0)
    
    x_prev[0, ...] = np.nan
    
    y_prev = np.roll(y, 1, axis = 0)
    
    y_prev[0, ...] = np.nan
    
    return (x > y) & (x_prev <= y_prev)


def crosses_under(
    x: np.ndarray,
    y: np.ndarray
) -> np.ndarray:
    """
    Detect cross-down events between two time-aligned arrays.

    Definition
    ----------
    At time t >= 1, a cross-down occurs if:
   
        (x_t < y_t) and (x_{t−1} >= y_{t−1}).
    
    Time t = 0 is False. NaN handling follows NumPy semantics.
    """

    x_prev = np.roll(x, 1, axis = 0)
    
    x_prev[0, ...] = np.nan

    y_prev = np.roll(y, 1, axis = 0)
    
    y_prev[0, ...] = np.nan

    return (x < y) & (x_prev >= y_prev)


def crosses_over_prerolled(
    x, 
    x_prev,
    y, 
    y_prev
):
    """
    Cross-up detector assuming `x_prev` and `y_prev` are already the lagged values.

    This micro-utility exists to avoid repeated `np.roll` when performance matters.
    Inputs are expected to have identical shapes, with the same broadcasting rules
    as in `crosses_over`.
    """
    
    return (x > y) & (x_prev <= y_prev)


def crosses_under_prerolled(
    x, 
    x_prev, 
    y, 
    y_prev
):
    """
    Cross-down detector assuming `x_prev` and `y_prev` are already the lagged values.

    See `crosses_under` for the definition of a cross-down event.
    """
    
    return (x < y) & (x_prev >= y_prev)


@njit(parallel = True, cache = True, fastmath = True)
def pulses_to_positions(
    buys, 
    sells, 
    signal_lag, 
    cooldown_bars, 
    mode_id
):
    """
    Convert instantaneous buy/sell pulses into a persistent position time series.

    State Machine
    -------------
    For each scenario p (i.e., column) and each time t:

    1) Effective pulses are read at index (t − signal_lag). If tt < 0, both are False.
    
    2) If `cooldown_bars > 0` and the scenario is cooling down, pulses are ignored,
    and the cooldown counter is decremented.
    
    3) If both buy and sell are True, resolve using `mode_id`:
    
        0 = sell_wins,
        
        1 = buy_wins, 
        
        2 = mutual_exclude (both dropped).
   
    4) Position update:
   
    - buy → position := +1; start cooldown.
   
    - sell → position := −1; start cooldown.
   
    - else → position := previous position (persistence).

    Parameters
    ----------
    buys, sells : np.ndarray[bool], shape (T, P)
        Boolean pulse matrices over time and parameter scenarios.
    signal_lag : int
        Number of bars to delay the reaction to a pulse (models execution latency).
    cooldown_bars : int
        Bars for which new pulses are ignored after a position change (debouncing).
    mode_id : int
        Conflict resolution code (0, 1, or 2 as described).

    Returns
    -------
    np.ndarray[int8], shape (T, P)
        Position process taking values in {−1, 0, +1}.

    Rationale
    ---------
    Lag avoids look-ahead bias from same-bar signal/return alignment.
    Cooldown reduces churning by enforcing a minimum holding period after flips.

    Notes
    -----
    The kernel is Numba-compiled for speed. It expects C-contiguous boolean
    arrays for best performance.
    """

    T, P = buys.shape
   
    pos = np.zeros((T, P), np.int8)
   
    for p in prange(P):
   
        cool = 0
   
        for t in range(T):
   
            tt = t - signal_lag
   
            b = bool(buys[tt, p]) if tt >= 0 and cool == 0 else False
   
            s = bool(sells[tt, p]) if tt >= 0 and cool == 0 else False
   
            if cool > 0:
   
                cool -= 1
   
            if b and s:
   
                if mode_id == 0:
                    
                    b = False       
   
                elif mode_id == 1:
                    
                    s = False    
   
                else: 
                    
                    b = s = False            
   
            prev = 0 if t == 0 else pos[t - 1, p]
           
            if b:
           
                pos[t, p] = 1
                
                cool = cooldown_bars
           
            elif s:
           
                pos[t, p] = -1
                
                cool = cooldown_bars
           
            else:
           
                pos[t, p] = prev
    
    return pos


def _eval_stats_from_pos(
    pos: np.ndarray,
    r: np.ndarray,
    annualisation: float
):
    """
    Compute Sharpe ratio, t-statistic, p-value, and sample size from positions and returns.

    Alignment
    ---------
    To eliminate look-ahead, position at time t−1 is applied to the return realised
    between t−1 and t:
   
        X_t := pos_{t−1} * r_t  for t = 1..T−1.
   
    Only non-NaN r_t are considered; the corresponding X_t are used in statistics.

    Statistics
    ----------
    Let n be the number of valid X_t values, μ be their sample mean, and s be the
    unbiased sample standard deviation:
   
        μ = (1/n) * Σ_{t} X_t
   
        s = sqrt( (1/(n−1)) * Σ_{t} (X_t − μ)^2 )

    Annualised Sharpe ratio:
   
        Sharpe = sqrt(A) * (μ / s)
   
    where A is the annualisation factor (e.g., 252 for trading days).

    One-sample t-statistic for H1: μ > 0:

        t = μ / (s / sqrt(n))

    One-sided p-value under Student t with df = n−1:

        p = sf(max(t, 0), df = n−1)

    where sf is the survival function.

    Return Values
    -------------
    sharpe : np.ndarray[float], shape (P,)
        Annualised Sharpe for each scenario.
    t_stat : np.ndarray[float], shape (P,)
        One-sided t-statistic.
    p_val : np.ndarray[float], shape (P,)
        One-sided p-values for H1: μ > 0.
    n : int
        Number of valid observations (shared across scenarios after alignment).

    Edge Cases
    ----------
    If n < 2 or s == 0, Sharpe is set to −inf to push the scenario to the bottom
    of any ranking, t_stat is 0, and p_val is 1.
    """

    valid = ~np.isnan(r)

    pos_eff = pos[:-1, :]

    r_eff = r[1:]

    valid_eff = valid[1:]

    if not valid_eff.any():

        P = pos.shape[1]

        return (np.full(P, -np.inf), np.zeros(P), np.ones(P), 0)

    pos_eff = pos_eff[valid_eff, :].astype(np.float64)

    r_eff = r_eff[valid_eff]

    ret_mat = pos_eff * r_eff[:, None]

    mean = np.nanmean(ret_mat, axis=0)

    std = np.nanstd(ret_mat, axis =0, ddof=1)

    n = ret_mat.shape[0]
    
    with np.errstate(divide = "ignore", invalid = "ignore"):
       
        sharpe = np.where(std > 0, np.sqrt(annualisation) * mean / std, -np.inf)
       
        t_stat = np.where(std > 0, mean / (std / np.sqrt(max(n, 1))), 0.0)
   
    df = max(n - 1, 1)
   
    p_val = stats.t.sf(np.maximum(t_stat, 0.0), df = df)
   
    return sharpe, t_stat, p_val, n


def _bh_sharpe_from_returns(
    r: np.ndarray,
    annualisation: float = ANNUALISATION
) -> float:
    """
    Compute buy-and-hold Sharpe using log returns to compound correctly.

    Method
    ------
    Given simple returns r_t, define geometric (log) returns:
   
        g_t = ln(1 + r_t)
   
    after aligning on t = 1..T−1 (to match strategy alignment).
   
    The annualised Sharpe is:
   
        Sharpe_BH = sqrt(A) * mean(g_t) / std(g_t, ddof=1)

    NaNs are dropped. If the standard deviation is zero or not finite, −inf is returned.
    """
   
    r_eff = r[1:]
   
    valid = ~np.isnan(r_eff)
   
    if not valid.any():
   
        return float("-inf")
   
    g = np.log1p(r_eff[valid].astype(np.float64))
   
    mu = np.nanmean(g)
   
    sd = np.nanstd(g, ddof=1)
   
    if not np.isfinite(sd) or sd <= 0:
   
        return float("-inf")
   
    return float(np.sqrt(annualisation) * (mu / sd))


def _best_index(
    sharpe, 
    t_stat, 
    p_val, 
    min_sharpe, 
    alpha
) -> int:
    """
    Select the index of the best scenario, preferring statistically significant ones.

    Rule
    ----
   
    1) Compute a boolean mask of scenarios with p_val < alpha and sharpe >= min_sharpe.
   
    2) If any are significant, pick the argmax Sharpe among them.
   
    3) Otherwise, pick the global argmax Sharpe (even if not significant).

    Returns
    -------
    int
        Column index suitable for indexing the statistics arrays.
    """

    sig = (p_val < alpha) & (sharpe >= min_sharpe)

    if sig.any():
        
        cand = np.where(sig, sharpe, -np.inf)  
    
    else:
        
        cand = sharpe

    return int(np.nanargmax(cand))


def true_range(
    high: pd.Series,
    low: pd.Series,
    close: pd.Series
) -> pd.Series:
    """
    Calculate Wilder's True Range (TR).

    Definition
    ----------
    For each time t >= 1:
       
        TR_t = max(
            high_t − low_t,
            |high_t − close_{t − 1}|,
            |low_t  − close_{t − 1}|
        )
    TR_0 is defined by pandas alignment (typically NaN due to lagged close).

    Use
    ---
    This is the base for ATR and ADX constructions. It captures gaps between bars
    and intraday range in a single volatility proxy.
    """

    tr = pd.concat(
        [high - low, (high - close.shift(1)).abs(), (low - close.shift(1)).abs()],
        axis = 1
    ).max(axis = 1)
    
    return tr


def atr_series(
    high: pd.Series,
    low: pd.Series, 
    close: pd.Series,
    window: int
) -> pd.Series:
    """
    Compute Average True Range (ATR) as a simple moving average of True Range.

    Definition
    ----------
    Given TR_t, ATR over a window w is:
   
        ATR_t = SMA_w(TR)_t
   
    i.e., the arithmetic mean of the last w TR values.

    Parameters
    ----------
    window : int
        Rolling window length. A Wilder RMA/EMA variant is also common; this
    implementation uses SMA for simplicity and determinism.

    Returns
    -------
    pd.Series of float
        ATR with NaN for the first `window−1` bars.

    Notes
    -----
    This implementation uses a simple moving average (SMA) of TR. Some variants
    use Wilder’s RMA/EMA. Consistency matters more than the exact smoother.
    """

    tr = true_range(
        high = high,
        low = low, 
        close = close
    )

    return tr.rolling(window, min_periods = window).mean()


def adx_series(
    high: pd.Series, 
    low: pd.Series, 
    close: pd.Series, 
    *, 
    window: int = ADX_WINDOW
) -> pd.Series:
    """
    Compute the Average Directional Index (ADX) following the DI+/DI− construction.

    Steps
    -----
    1) True Range as in `true_range`.
  
    2) Directional movement:
  
        up_move_t   = max(high_t − high_{t−1}, 0)
  
        down_move_t = max(low_{t−1} − low_t, 0)
  
    Then:
  
        +DM_t = up_move_t   if up_move_t > down_move_t else 0
  
        −DM_t = down_move_t if down_move_t > up_move_t else 0

    3) Wilder-style exponential smoothing with α = 1 / window:
    
        ATR_t     = EMA(TR_t; α)
    
        +DM_ema_t = EMA(+DM_t; α)
    
        −DM_ema_t = EMA(−DM_t; α)

    4) Directional indicators (percent):
    
        +DI_t = 100 * (+DM_ema_t / ATR_t)
    
        −DI_t = 100 * (−DM_ema_t / ATR_t)

    5) Directional index and its smoothing:

        DX_t  = 100 * |(+DI_t − −DI_t)| / (+DI_t + −DI_t)

        ADX_t = EMA(DX_t; α)

    Returns
    -------
    pd.Series
        ADX in [0, 100], with NaN where insufficient history exists.

    Interpretation
    --------------
    Common practice treats ADX >= 25 as a trending regime. This module uses ADX
    both to gate pro-trend indicators and to exclude anti-trend ones where trend is strong.
    """

    tr = true_range(
        high = high, 
        low = low, 
        close = close
    )
    
    alpha = 1.0 / window

    atr = tr.ewm(alpha = alpha, adjust = False).mean().replace(0, np.nan)

    up_move = (high.diff()).clip(lower = 0.0)

    down_move = (-low.diff()).clip(lower = 0.0)

    plus_dm = up_move.where(up_move > down_move, 0.0)

    minus_dm = down_move.where(down_move > up_move, 0.0)
    
    h_atr = 100 / atr

    plus_di = h_atr * plus_dm.ewm(alpha = alpha, adjust = False).mean() 

    minus_di = h_atr * minus_dm.ewm(alpha = alpha, adjust = False).mean() 

    denom = (plus_di + minus_di).replace(0, np.nan)

    dx = 100 * (plus_di - minus_di).abs() / denom

    adx = dx.ewm(alpha = alpha, adjust = False).mean()

    return adx


@dataclass
class Precomp:
    """
    Packed precomputations for one ticker to avoid repeated O(T) passes.
    """
   
    idx: pd.DatetimeIndex

    c: np.ndarray
   
    h: np.ndarray
   
    l: np.ndarray
   
    v: np.ndarray
   
    r: np.ndarray

    trend_mask: np.ndarray
   
    vol_ok: np.ndarray

    rsi: np.ndarray
   
    bb_mid: np.ndarray
   
    bb_std: np.ndarray
   
    atr: np.ndarray
   
    hi_prev: np.ndarray
   
    lo_prev: np.ndarray
   
    mfi: np.ndarray
   
    stoch_k: np.ndarray
   
    stoch_d: np.ndarray
   
    ema_fast: np.ndarray
   
    ema_slow: np.ndarray
   
    obv: np.ndarray
   
    price_low_prev: np.ndarray
   
    price_high_prev: np.ndarray
   
    obv_low_prev: np.ndarray
   
    obv_high_prev: np.ndarray

    _ema_cache: dict | None = None                 
   
    _macd_sig_cache: dict | None = None      
   
    _macd_line_cache: dict | None = None           
   
    _rsi_cache: dict | None = None        
   
    _mfi_cache: dict | None = None        
   
    _bb_cache: dict | None = None            
   
    _atr_cache: dict | None = None         
   
    _break_cache: dict | None = None         

    _cs_pv: np.ndarray | None = None            
   
    _cs_v: np.ndarray | None = None           
    

def precompute_all(
    h: pd.Series,
    l: pd.Series,
    c: pd.Series,
    v: pd.Series
) -> Precomp:
    """
    Compute and cache per-ticker arrays required across indicator scans.

    Precomputations
    ---------------
    - Price and volume arrays: close (c), high (h), low (l), volume (v).
   
    - Simple returns 
    
        r_t = (c_t / c_{t−1}) − 1 as float64.
   
    - ADX series and a boolean `trend_mask`: 
    
        True where ADX >= ADX_ENTRY.
    
    - Volume filter `vol_ok`: 
    
        True where volume >= rolling median over 20 bars.
        
    - Stochastic oscillator:
        
        For K window = STOCH_K and D window = STOCH_D,
        
        low_k  = rolling min(low, K), high_k = rolling max(high, K),
        
        %K_t   = 100 * (close_t − low_k_t) / max(high_k_t − low_k_t, NaN if zero),
        
        %D_t   = SMA_{STOCH_D}(%K).
    
    - On-Balance Volume (OBV):
    
        direction_t = sign(close_t − close_{t−1}) with NaN→0,
    
        OBV_t = cumsum(direction_t * volume_t).
    
    - Caches for EMA, MACD line/signal, RSI, MFI, Bollinger mid/std, ATR, breakout bands,
    price and OBV rolling extrema, and cumulative sums for fast rolling VWAP:
    
        cs_pv = cumsum(typical_price * volume), cs_v = cumsum(volume),
    
        where 
        
        typical_price_t = (high_t + low_t + close_t) / 3.

    Returns
    -------
    Precomp
        A dataclass bundling raw arrays, gating masks, derived indicators, and caches.

    Why
    ---
    Indicator grid scans revisit the same rolling transforms many times. Centralising
    them reduces asymptotic cost from O(K * T) to O(T) for each distinct parameter,
    and allows vectorised or Numba kernels to operate on contiguous arrays.
    """
   
    idx = c.index

    c_np = c.to_numpy(np.float64)
    
    h_np = h.to_numpy(np.float64)
    
    l_np = l.to_numpy(np.float64)
    
    v_np = v.to_numpy(np.float64)

    r = c.pct_change().to_numpy(np.float64)

    adx = adx_series(
        high = h, 
        low = l,
        close = c, 
        window = ADX_WINDOW
    )
    
    trend_mask = (adx >= ADX_ENTRY).fillna(False).to_numpy(np.bool_)
    
    vol_ok = (v >= v.rolling(VOL_CONFIRM_WIN := 20, min_periods = 20).median()).fillna(False).to_numpy(np.bool_)

    low_k = l.rolling(STOCH_K, min_periods = STOCH_K).min()
    
    high_k = h.rolling(STOCH_K, min_periods = STOCH_K).max()
    
    den = (high_k - low_k).replace(0, np.nan)
    
    pct_k = 100 * (c - low_k) / den
    
    stoch_k = pct_k.replace([np.inf, -np.inf], np.nan).to_numpy(np.float64)
    
    stoch_d = pct_k.rolling(STOCH_D, min_periods = STOCH_D).mean().to_numpy(np.float64)

    direction = np.sign(c.diff().to_numpy(np.float64))
   
    direction = np.nan_to_num(direction, nan = 0.0)
   
    obv = (direction * v.to_numpy(np.float64)).cumsum()

    pc = Precomp(
        idx = idx, 
        c = c_np, 
        h = h_np, 
        l = l_np, 
        v = v_np,
        r = r,
        trend_mask = trend_mask, 
        vol_ok = vol_ok,
        rsi = np.array([]), 
        bb_mid = np.array([]), 
        bb_std = np.array([]), 
        atr = np.array([]),
        hi_prev = np.array([]), 
        lo_prev = np.array([]), 
        mfi = np.array([]),
        stoch_k = stoch_k, 
        stoch_d = stoch_d,
        ema_fast = np.array([]), 
        ema_slow = np.array([]),
        obv = obv,
        price_low_prev = np.array([]),
        price_high_prev = np.array([]),
        obv_low_prev = np.array([]),
        obv_high_prev = np.array([]),
    )

    pc._ema_cache = {}
   
    pc._macd_sig_cache = {}
   
    pc._macd_line_cache = {}
   
    pc._rsi_cache = {}
   
    pc._mfi_cache = {}
   
    pc._bb_cache = {}
   
    pc._atr_cache = {}
   
    pc._break_cache = {}
   
    pc._price_extrema_cache = {}
   
    pc._obv_extrema_cache = {}

    pc._ema_cache[EMA_FAST] = pd.Series(pc.c, index = pc.idx).ewm(span = EMA_FAST, adjust = False).mean().to_numpy(np.float64)
    
    pc._ema_cache[EMA_SLOW] = pd.Series(pc.c, index = pc.idx).ewm(span = EMA_SLOW, adjust = False).mean().to_numpy(np.float64)

    pc._rsi_cache[RSI_WINDOW] = ta.momentum.RSIIndicator(pd.Series(pc.c, index = pc.idx), window = RSI_WINDOW).rsi().to_numpy(np.float64)


    def _calc_mfi(
        window: int
    ) -> np.ndarray:
    
        s_h = pd.Series(pc.h, index = pc.idx)
    
        s_l = pd.Series(pc.l, index = pc.idx)
    
        s_c = pd.Series(pc.c, index = pc.idx)
    
        s_v = pd.Series(pc.v, index = pc.idx)
    
        tp = (s_h + s_l + s_c) / 3.0
    
        mf = tp * s_v
    
        up = mf.where(tp.diff() > 0, 0.0)
    
        dn = mf.where(tp.diff() < 0, 0.0)
    
        up_sum = up.rolling(window, min_periods = window).sum()
    
        dn_sum = dn.rolling(window, min_periods = window).sum().replace(0, np.nan)
    
        ratio = (up_sum / dn_sum).replace([np.inf, -np.inf], np.nan)
    
        return (100 - (100 / (1 + ratio))).to_numpy(np.float64)
    
    
    pc._mfi_cache[MFI_WINDOW] = _calc_mfi(
        window = MFI_WINDOW
    )

    s = pd.Series(pc.c, index = pc.idx)
    
    pc._bb_cache[BB_WINDOW] = (
        s.rolling(BB_WINDOW, min_periods = BB_WINDOW).mean().to_numpy(np.float64),
        s.rolling(BB_WINDOW, min_periods = BB_WINDOW).std(ddof = 0).to_numpy(np.float64)
    )

    s_h = pd.Series(pc.h, index = pc.idx)
   
    s_l = pd.Series(pc.l, index = pc.idx)
   
    s_c = pd.Series(pc.c, index = pc.idx)
   
    pc._atr_cache[ATR_WINDOW] = atr_series(
        high = s_h, 
        low = s_l, 
        close = s_c, 
        window = ATR_WINDOW
    ).to_numpy(np.float64)
   
    pc._break_cache[ATR_BREAK_WINDOW] = (
        s_h.rolling(ATR_BREAK_WINDOW, min_periods = ATR_BREAK_WINDOW).max().shift(1).to_numpy(np.float64),
        s_l.rolling(ATR_BREAK_WINDOW, min_periods = ATR_BREAK_WINDOW).min().shift(1).to_numpy(np.float64),
    )

    tp_np = ((h_np + l_np + c_np) / 3.0)
   
    pv = np.nan_to_num(tp_np * v_np, nan = 0.0)
   
    vv = np.nan_to_num(v_np, nan = 0.0)
   
    pc._cs_pv = np.cumsum(pv)
   
    pc._cs_v  = np.cumsum(vv)

    return pc


def _ema_cached(
    pc: Precomp, 
    span: int
) -> np.ndarray:
    """
    Return the exponential moving average (EMA) of close for a given span, with caching.

    Definition
    ----------
    
    EMA_t = α * close_t + (1 − α) * EMA_{t−1}, 
    
    with α = 2 / (span + 1), initialised to pandas' convention (via `ewm` with `adjust=False`).

    Caching avoids recomputation across scans that share the same span.
    """
    
    if pc._ema_cache is None:
    
        pc._ema_cache = {}
    
    span = int(span)
    
    if span not in pc._ema_cache:
    
        pc._ema_cache[span] = pd.Series(pc.c, index=pc.idx).ewm(span = span, adjust = False).mean().to_numpy(np.float64)
    
    return pc._ema_cache[span]


def _macd_line_cached(
    pc: Precomp, 
    fast: int, 
    slow: int
) -> np.ndarray:
    """
    Return the MACD line for (fast, slow) spans, with caching.

    Definition
    ----------
    MACD_line_t = EMA_fast_t − EMA_slow_t, 
    
    where EMA spans satisfy fast < slow.

    The function reuses cached EMAs via `_ema_cached` and memoises the line array by key (fast, slow).
    """

    if pc._macd_line_cache is None:

        pc._macd_line_cache = {}

    key = (int(fast), int(slow))

    if key not in pc._macd_line_cache:

        ef = _ema_cached(
            pc = pc, 
            span = key[0]
        )

        es = _ema_cached(
            pc = pc, 
            span = key[1]
        )

        pc._macd_line_cache[key] = (ef - es)

    return pc._macd_line_cache[key]


def _macd_signal_cached(
    pc: Precomp,
    fast: int, 
    slow: int, 
    signal: int
) -> tuple[np.ndarray, np.ndarray]:
    """
    Return (MACD line, MACD signal) for a (fast, slow, signal) triple.

    Definitions
    -----------
    - MACD line_t = EMA_fast_t − EMA_slow_t.
    
    - MACD signal_t = EMA_{signal}(MACD line)_t.

    The returned MACD line is consistent with `_macd_line_cached`. Results are cached
    to avoid repeated smoothing across grid points.
    """

    if pc._macd_sig_cache is None:
   
        pc._macd_sig_cache = {}
   
    key = (int(fast), int(slow), int(signal))
   
    if key not in pc._macd_sig_cache:
   
        line = _macd_line_cached(
            pc = pc, 
            fast = key[0],
            slow = key[1]
        )
   
        sig = pd.Series(line, index = pc.idx).ewm(span = key[2], adjust = False).mean().to_numpy(np.float64)
   
        pc._macd_sig_cache[key] = sig
   
    return _macd_line_cached(
        pc = pc,
        fast = fast, 
        slow = slow
    ), pc._macd_sig_cache[key]


def _rolling_sum_from_cumsum(
    cs: np.ndarray, 
    w: int
) -> np.ndarray:
    """
    Compute a rolling sum of width w from a cumulative sum array.

    Formula
    -------
    For an array of cumulative sums cs with cs[i] = Σ_{k=0..i} x[k],
    the rolling sum over the last w values at index i >= w−1 is:
  
        roll[i] = cs[i] − cs[i−w]
  
    The first w−1 outputs are NaN due to insufficient history.

    This is used to form a rolling VWAP efficiently.
    """

    w = int(w)
    
    n = cs.shape[0]
    
    out = np.full(n, np.nan, dtype = np.float64)
    
    if w <= 0 or n == 0 or n < w:
    
        return out
    
    out[w-1:] = cs[w-1:] - np.r_[0.0, cs[:-w]]
    
    return out


def _rsi_cached(
    pc: Precomp,
    window: int
) -> np.ndarray:
    """
    Return the Relative Strength Index (RSI) for a specified window, with caching.

    Definition
    ----------
    RSI_t = 100 − 100 / (1 + RS_t), where RS_t is the ratio of smoothed average
    gains to smoothed average losses over `window`. The implementation uses
    `ta.momentum.RSIIndicator`, which internally follows a Wilder-style smoothing.

    Interpretation
    --------------
    RSI is bounded in [0, 100]. Low values (e.g., below 30) often indicate short-term
    weakness; high values (e.g., above 70) indicate short-term strength.
    """

    if pc._rsi_cache is None: 
        
        pc._rsi_cache = {}

    w = int(window)

    if w not in pc._rsi_cache:

        pc._rsi_cache[w] = ta.momentum.RSIIndicator(pd.Series(pc.c, index = pc.idx), window = w).rsi().to_numpy(np.float64)

    return pc._rsi_cache[w]


def _mfi_cached(
    pc: Precomp, 
    window: int
) -> np.ndarray:
    """
    Return the Money Flow Index (MFI) over a specified window, with caching.

    Construction
    ------------
    1) Typical price:
    
        TP_t = (high_t + low_t + close_t) / 3.
    
    2) Money flow: 
    
        MF_t = TP_t * volume_t.
    
    3) Positive flow: 
    
        sum MF_t where TP_t > TP_{t−1}
        
    negative flow: 
        
        sum MF_t where TP_t < TP_{t−1}.
    
    4) Over a rolling window w, compute:
    
        pos_w = Σ positive MF,  neg_w = Σ negative MF (use NaN when neg_w == 0).
   
    5) Money ratio: MR_t = pos_w / neg_w.
   
    6) MFI_t = 100 − 100 / (1 + MR_t).

    Bounded in [0, 100]; used analogously to RSI but volume-weighted.
    """

    if pc._mfi_cache is None:
        
        pc._mfi_cache = {}

    w = int(window)

    if w not in pc._mfi_cache:
       
        s_h = pd.Series(pc.h, index = pc.idx)
       
        s_l = pd.Series(pc.l, index = pc.idx)
       
        s_c = pd.Series(pc.c, index = pc.idx)
       
        s_v = pd.Series(pc.v, index = pc.idx)
       
        tp = (s_h + s_l + s_c) / 3.0
       
        mf = tp * s_v
        
        up = mf.where(tp.diff() > 0, 0.0)
        
        dn = mf.where(tp.diff() < 0, 0.0)
        
        up_sum = up.rolling(w, min_periods = w).sum()
        
        dn_sum = up_sum * 0 + dn.rolling(w, min_periods = w).sum().replace(0, np.nan)
        
        ratio = (up_sum / dn_sum).replace([np.inf, -np.inf], np.nan)
        
        pc._mfi_cache[w] = (100 - (100 / (1 + ratio))).to_numpy(np.float64)
    
    return pc._mfi_cache[w]


def _bb_cached(
    pc: Precomp,
    window: int
) -> tuple[np.ndarray, np.ndarray]:
    """
    Return Bollinger band components for a window: (middle, standard deviation).

    Definitions
    -----------
    For window w:
       
        middle_t = SMA_w(close)_t
       
        std_t    = standard deviation over the last w closes (population ddof=0)

    Complete bands for a multiplier n are middle ± n*std, built on demand by callers.
    """

    if pc._bb_cache is None: 
        
        pc._bb_cache = {}
  
    w = int(window)
  
    if w not in pc._bb_cache:
  
        s = pd.Series(pc.c, index = pc.idx)
  
        mid = s.rolling(w, min_periods = w).mean().to_numpy(np.float64)
  
        std = s.rolling(w, min_periods = w).std(ddof = 0).to_numpy(np.float64)
  
        pc._bb_cache[w] = (mid, std)
  
    return pc._bb_cache[w]


def _atr_cached(
    pc: Precomp, 
    window: int
) -> np.ndarray:
    """
    Return an ATR array for a given window, caching the result.

    ATR is computed as the rolling SMA of True Range over `window`. See `atr_series`.
    """

    if pc._atr_cache is None: 
        
        pc._atr_cache = {}

    w = int(window)

    if w not in pc._atr_cache:

        s_h = pd.Series(pc.h, index = pc.idx)

        s_l = pd.Series(pc.l, index = pc.idx)

        s_c = pd.Series(pc.c, index = pc.idx)

        pc._atr_cache[w] = atr_series(
            high = s_h, 
            low = s_l, 
            close = s_c,
            window = w
        ).to_numpy(np.float64)

    return pc._atr_cache[w]


def _break_cached(
    pc: Precomp,
    window: int
) -> tuple[np.ndarray, np.ndarray]:
    """
    Return rolling breakout levels used by ATR breakout logic.

    For a rolling window w:

    - hi_prev_t = max(high over last w bars) shifted by one bar (yesterday’s rolling high).

    - lo_prev_t = min(low  over last w bars) shifted by one bar (yesterday’s rolling low).

    These are combined with ATR to form dynamic breakout thresholds.
    """

    if pc._break_cache is None: 
        
        pc._break_cache = {}

    w = int(window)

    if w not in pc._break_cache:

        s_h = pd.Series(pc.h, index=pc.idx)

        s_l = pd.Series(pc.l, index=pc.idx)

        hi_prev = s_h.rolling(w, min_periods = w).max().shift(1).to_numpy(np.float64)

        lo_prev = s_l.rolling(w, min_periods = w).min().shift(1).to_numpy(np.float64)

        pc._break_cache[w] = (hi_prev, lo_prev)

    return pc._break_cache[w]


def _price_extrema_cached(
    pc: Precomp,
    window: int
) -> tuple[np.ndarray, np.ndarray]:
    """
    Return rolling price extrema (min and max of close) shifted by one bar.

    For a window w:
   
    - low_prev_t  = min(close over last w bars) shifted by one.
   
    - high_prev_t = max(close over last w bars) shifted by one.

    Shape is (T, 1) to ease broadcasting against (T, P) matrices in scans.
    """

    if not hasattr(pc, "_price_extrema_cache") or pc._price_extrema_cache is None:

        pc._price_extrema_cache = {}

    w = int(window)

    if w not in pc._price_extrema_cache:

        s_c = pd.Series(pc.c, index = pc.idx)

        low_prev  = s_c.rolling(w, min_periods = w).min().shift(1).to_numpy(np.float64)[:, None]

        high_prev = s_c.rolling(w, min_periods = w).max().shift(1).to_numpy(np.float64)[:, None]

        pc._price_extrema_cache[w] = (low_prev, high_prev)

    return pc._price_extrema_cache[w]


def _obv_extrema_cached(
    pc: Precomp,
    window: int
) -> tuple[np.ndarray, np.ndarray]:
    """
    Return rolling OBV extrema (min and max of OBV) shifted by one bar.

    Used to detect divergences where price makes a new extreme but OBV does not.
    """

    if not hasattr(pc, "_obv_extrema_cache") or pc._obv_extrema_cache is None:

        pc._obv_extrema_cache = {}

    w = int(window)

    if w not in pc._obv_extrema_cache:

        s_obv = pd.Series(pc.obv, index = pc.idx)

        low_prev  = s_obv.rolling(w, min_periods = w).min().shift(1).to_numpy(np.float64)[:, None]

        high_prev = s_obv.rolling(w, min_periods = w).max().shift(1).to_numpy(np.float64)[:, None]

        pc._obv_extrema_cache[w] = (low_prev, high_prev)

    return pc._obv_extrema_cache[w]


def eval_bb_grid(
    pc: Precomp,
    num_std_grid: np.ndarray,
    window_grid: np.ndarray,
    is_pro_trend: bool, is_anti_trend: bool, vol_required: bool,
    signal_lag: int = SIGNAL_LAG, cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate a Cartesian grid of Bollinger band multipliers and windows.

    Signals
    -------
    With middle_t and std_t computed over window w, and a multiplier n:

    - Lower band:  lower_t = middle_t − n * std_t
   
    - Upper band:  upper_t = middle_t + n * std_t

    Buy pulse at t:
        
        close_t crosses up lower_t (i.e., close_t > lower_t and close_{t−1} <= lower_{t−1}).
    
    Sell pulse at t:
        
        close_t crosses down upper_t.

    Both pulses are optionally gated by trend and liquidity via `_build_gate`.

    Backtest and Statistics
    -----------------------
    - Pulses → positions via `pulses_to_positions` using configured lag, cooldown,
    and conflict policy.
  
    - Performance statistics via `_eval_stats_from_pos`, including annualised Sharpe,
    one-sided t-statistic, and p-value.
  
    - The best scenario is chosen by `_best_index`, preferring significant outcomes.
  
    - `n_trades` counts position flips (number of sign changes of the position).

    Returns
    -------
    dict
        {
        "window": int,
        "num_std": float,
        "sharpe": float, "t_stat": float, "p_value": float,
        "n": int, "n_trades": int
        }
    """

    NS = num_std_grid.astype(np.float64)
   
    W  = np.array(window_grid, dtype = int)

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )

    buy_blocks = []
    
    sell_blocks = []
    
    offsets = []
   
    for w in W:
   
        mid, sd = _bb_cached(
            pc = pc,
            window = w
        )
   
        upper = mid[:, None] + sd[:, None] * NS[None, :]
   
        lower = mid[:, None] - sd[:, None] * NS[None, :]
      
        c2 = pc.c[:, None]
      
        c2_prev = np.roll(c2, 1, axis=0); c2_prev[0, :] = np.nan
      
        buy_mat  = crosses_over_prerolled(
            x = c2,
            x_prev = c2_prev, 
            y = lower,
            y_prev = np.roll(lower, 1, axis = 0)
        )
      
        sell_mat = crosses_under_prerolled(
            x = c2,
            x_prev = c2_prev, 
            y = upper, 
            y_prev = np.roll(upper, 1, axis = 0)
        )
        lower_prev = np.roll(lower, 1, axis = 0)
        
        lower_prev[0, :] = np.nan
        
        upper_prev = np.roll(upper, 1, axis = 0)
        
        upper_prev[0, :] = np.nan
        
        buy_mat = (c2 > lower) & (c2_prev <= lower_prev)
        
        sell_mat = (c2 < upper) & (c2_prev >= upper_prev)
        
        if not gate.all():
        
            buy_mat  &= gate[:, None]
        
            sell_mat &= gate[:, None]
        
        buy_blocks.append(buy_mat)
        
        sell_blocks.append(sell_mat)
        
        offsets.append((int(w), len(NS)))

    buy_all = np.ascontiguousarray(np.concatenate(buy_blocks,  axis = 1))
   
    sell_all = np.ascontiguousarray(np.concatenate(sell_blocks, axis = 1))

    pos = pulses_to_positions(
        buys = buy_all,
        sells = sell_all,
        signal_lag = signal_lag,
        cooldown_bars = cooldown,
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos, 
        r = pc.r,
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe, 
        t_stat = t_stat, 
        p_val = p_val, 
        min_sharpe = MIN_SHARPE_FOR_SIG,
        alpha = SIGNIFICANCE_ALPHA
    )

    col0 = 0
    
    chosen_w = None
    
    i_ns = None
    
    for (w, n_ns) in offsets:
        
        if j < col0 + n_ns:
        
            chosen_w = w
            
            i_ns = j - col0
            
            break
      
        col0 += n_ns

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
    
    return {
        "window": int(chosen_w),
        "num_std": float(NS[i_ns]),
        "sharpe": float(sharpe[j]),
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]),
        "n": int(n),
        "n_trades": flips
    }


def eval_atr_grid(
    pc: Precomp,
    mult_grid: np.ndarray,
    atr_window: int,
    break_window: int,
    is_pro_trend: bool, is_anti_trend: bool, vol_required: bool,
    signal_lag: int = SIGNAL_LAG, cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate ATR breakout signals over a multiplier grid and fixed windows.

    Construction
    ------------
    Given:
    
    - ATR_{t−1} computed over `atr_window`,
    
    - hi_prev_t, lo_prev_t computed over `break_window` and lagged by one bar,

    Define dynamic thresholds:
     
        up_th_t = hi_prev_t + mult * ATR_{t − 1}
     
        dn_th_t = lo_prev_t − mult * ATR_{t − 1}

    Signals
    -------
    Buy pulse: close_t > up_th_t (upside breakout).
   
    Sell pulse: close_t < dn_th_t (downside breakout).

    Gating, backtest, and statistics as in `eval_bb_grid`. The chosen configuration
    maximises Sharpe under the significance preference and reports the number of flips.
    """

    M = mult_grid.astype(np.float64)
   
    atr = _atr_cached(
        pc = pc,
        window = atr_window
    )
    
    atr_prev = np.roll(atr, 1)
    
    atr_prev[0] = np.nan
    
    hi_prev, lo_prev = _break_cached(
        pc = pc, 
        window = break_window
    )

    up_th = hi_prev[:, None] + atr_prev[:, None] * M[None, :]
   
    dn_th = lo_prev[:, None] - atr_prev[:, None] * M[None, :]
   
    c2 = pc.c[:, None]
   
    buy_mat = (c2 > up_th)
   
    sell_mat = (c2 < dn_th)

    gate = _build_gate(
        pc = pc,
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
   
    if not gate.all():
      
        buy_mat  &= gate[:, None]
      
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
    
    sell_mat = np.ascontiguousarray(sell_mat)

    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat,
        signal_lag = signal_lag, 
        cooldown_bars = cooldown,
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos,
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe, 
        t_stat = t_stat, 
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG, 
        alpha = SIGNIFICANCE_ALPHA
    )
    
    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))

    return {
        "atr_window": int(atr_window),
        "break_window": int(break_window),
        "mult": float(M[j]),
        "sharpe": float(sharpe[j]),
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]),
        "n": int(n),
        "n_trades": flips
    }


@njit(parallel = True, cache = True, fastmath = True)
def eval_mfi_cartesian_kernel(
    mfi, 
    r, 
    buys,
    sells, 
    signal_lag, 
    cooldown,
    mode_id, 
    annualisation
):
    """
    Numba kernel to scan a full Cartesian grid of MFI buy/sell thresholds.

    Inputs
    ------
    mfi : float64[T]
        Money Flow Index series (may contain NaNs).
   
    r : float64[T]
        Simple returns aligned with prices (r_t applies to the move from t−1 to t).
   
    buys, sells : float64
        One-dimensional arrays of threshold levels. A cross is defined against the
        threshold level itself (not a moving series).
   
    signal_lag, cooldown, mode_id : int
        Passed through to the embedded pulse-to-position logic.
   
    annualisation : float
        Used to annualise the Sharpe ratio.

    Signal Logic
    ------------
    One-bar cross logic against constants:
   
    - Buy pulse if MFI crosses up through buy threshold.
   
    Formally: 
    
        (MFI_t > b) and (MFI_{t−1} <= b).
   
    - Sell pulse if MFI crosses down through sell threshold.
    
    Formally: 
    
        (MFI_t < s) and (MFI_{t−1} >= s).

    Backtest
    --------
    Positions are built inline (to avoid allocations) with lag and cooldown. Returns
    are aggregated using Welford’s one-pass algorithm for numerical stability:
   
    - mean and variance are updated online for 
    
        X_t = pos_{t−1} * r_t.

    Outputs
    -------
    sharpe, tstat, pval : float64[Nb, Ns]
        Annualised Sharpe, one-sided t-statistic, and corresponding p-value for each grid point.
    flips : int32[Nb, Ns]
        Number of position flips.
    n_valid : int
        Number of valid daily returns used (shared across grid points).

    Statistical Formulas
    --------------------
    - Sharpe = sqrt(A) * mean(X) / std(X).
    
    - t = mean(X) / (std(X) / sqrt(n)).
    
    - p = 0.5 * erfc( max(t, 0) / sqrt(2) )  (numerically stable one-sided p-value).
    """

    T = mfi.shape[0]
  
    Nb = buys.shape[0]
    
    Ns = sells.shape[0]
  
    sharpe = np.full((Nb, Ns), -np.inf)
  
    tstat = np.zeros((Nb, Ns))
  
    pval = np.ones((Nb, Ns))
  
    flips = np.zeros((Nb, Ns), dtype = np.int32)

    r_eff = r[1:]
   
    valid_eff = ~np.isnan(r_eff)
   
    n_valid = 0
   
    for t in range(r_eff.shape[0]):
   
        if valid_eff[t]:
   
            n_valid += 1

    for ib in prange(Nb):
    
        bt = buys[ib]
    
        for is_ in range(Ns):
    
            st = sells[is_]

            pos_t = 0
        
            cool = 0
        
            flips_count = 0
        
            prev_pos = 0
        
            prev_val = np.nan
        
            pos_series = np.zeros(T, np.int8)

            for t in range(T):
               
                tt = t - signal_lag
               
                b = False
                
                s = False
                
                if tt >= 0 and cool == 0:
                
                    cur = mfi[tt]
                
                    if not np.isnan(cur) and not np.isnan(prev_val):
                
                        b = (cur > bt) and (prev_val <= bt)
                
                        s = (cur < st) and (prev_val >= st)
                
                if cool > 0:
                
                    cool -= 1

                if b and s:
                
                    if mode_id == 0: 
                
                        b = False
                
                    elif mode_id == 1: 
                
                        s = False
                
                    else:             
                
                        b = False
                        
                        s = False

                if b:
                
                    pos_t = 1
                    
                    cool = cooldown
                
                elif s:
                
                    pos_t = -1
                    
                    cool = cooldown

                pos_series[t] = pos_t

                if t > 0 and pos_series[t] != prev_pos:
             
                    flips_count += 1
             
                prev_pos = pos_series[t]

                prev_val = mfi[t]  

            mean = 0.0
            
            m2 = 0.0
            
            cnt = 0
            
            for t in range(1, T):
            
                if not np.isnan(r[t]):
            
                    x = float(pos_series[t - 1]) * float(r[t])
            
                    cnt += 1
            
                    delta = x - mean
            
                    mean += delta / cnt
            
                    m2 += delta * (x - mean)

            if cnt <= 1:
             
                sharpe[ib, is_] = -np.inf
             
                tstat[ib, is_] = 0.0
             
                pval[ib, is_] = 1.0
             
                flips[ib, is_] = flips_count
             
                continue

            var = m2 / (cnt - 1)
            
            sd = np.sqrt(var) if var > 0 else np.nan
            
            if not np.isfinite(sd) or sd <= 0:
            
                sharpe[ib, is_] = -np.inf
            
                tstat[ib, is_] = 0.0
            
                pval[ib, is_] = 1.0
            
                flips[ib, is_] = flips_count
            
                continue

            sr = np.sqrt(annualisation) * (mean / sd)
            
            t = mean / (sd / np.sqrt(cnt))
            
            sharpe[ib, is_] = sr
            
            tstat[ib, is_] = t
            
            z = max(t, 0.0)
            
            p = 0.5 * erfc(z / np.sqrt(2.0))
            
            pval[ib, is_] = p
            
            flips[ib, is_] = flips_count

    return sharpe, tstat, pval, flips, n_valid


def eval_mfi_grid(
    pc: Precomp,
    buy_grid: np.ndarray,
    sell_grid: np.ndarray,
    window_grid: np.ndarray,
    is_pro_trend: bool, is_anti_trend: bool, vol_required: bool,
    signal_lag: int = SIGNAL_LAG, cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE,
    *,
    top_k: int = 128,
):
    """
    Evaluate MFI threshold grids across multiple smoothing windows.

    Process
    -------
    For each window w in `window_grid`:
    
    1) Compute or retrieve MFI(w).
    
    2) Optionally apply the trend/volume gate by setting barred entries to NaN.
    
    3) Call the Numba kernel `eval_mfi_cartesian_kernel` over the buy/sell threshold grid.
    
    4) Collect statistics and identify the best configuration by `_best_index`,
    optionally restricting search to the top-K Sharpe candidates for efficiency.

    Returns
    -------
    dict
        {
        "window": int,
        "buy_thresh": float, "sell_thresh": float,
        "sharpe": float, "t_stat": float, "p_value": float,
        "n": int, "n_trades": int
        }

    Why Use Top-K
    -------------
    A partial selection (`np.argpartition`) focuses any further selection on the
    most promising candidates, reducing overhead for very large grids while leaving
    the final choice to `_best_index` (which applies significance preferences).
    """
   
    B = buy_grid.astype(np.float64)
   
    S = sell_grid.astype(np.float64)
   
    W = np.array(window_grid, dtype = int)

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend,
        vol_required = vol_required
    )

    per_win_stats = []
    
    per_win_shapes = []

    for w in W:
    
        mfi = _mfi_cached(
            pc = pc, 
            window = int(w)
        ).astype(np.float64)
       
        if not gate.all():
    
            mfi = mfi.copy()
    
            mfi[~gate] = np.nan

        sr, tt, pv, flips, n = eval_mfi_cartesian_kernel(
            mfi = mfi,
            r = pc.r.astype(np.float64),
            buys = B,
            sells = S,
            signal_lag = signal_lag,
            cooldown = cooldown,
            mode_id = _mode_id(
                mode = conflict_mode
            ),
            annualisation = float(ANNUALISATION)
        )
       
        Nb, Ns = sr.shape
       
        per_win_stats.append({
            "w": int(w),
            "sr": sr.ravel(),
            "tt": tt.ravel(),
            "pv": pv.ravel(),
            "flips": flips.ravel().astype(np.int64),
            "n": int(n),
        })
       
        per_win_shapes.append((Nb, Ns))

    if not per_win_stats:
      
        return {
            "window": int(W[0] if len(W) else MFI_WINDOW),
            "buy_thresh": float(B[0] if len(B) else MFI_BUY_THRESH),
            "sell_thresh": float(S[0] if len(S) else MFI_SELL_THRESH),
            "sharpe": float("-inf"),
            "t_stat": 0.0,
            "p_value": 1.0,
            "n": 0,
            "n_trades": 0,
        }

    sr_all = np.concatenate([d["sr"] for d in per_win_stats])
  
    offsets = np.cumsum([0] + [d["sr"].shape[0] for d in per_win_stats])
  
    finite_mask = np.isfinite(sr_all)

    if finite_mask.any():
  
        cand_pool = np.flatnonzero(finite_mask)
  
    else:
  
        try:
  
            j_global = int(np.nanargmax(sr_all))
  
            cand_pool = np.array([j_global], dtype=int)
  
        except Exception:
  
            return {
                "window": int(W[0] if len(W) else MFI_WINDOW),
                "buy_thresh": float(B[0] if len(B) else MFI_BUY_THRESH),
                "sell_thresh": float(S[0] if len(S) else MFI_SELL_THRESH),
                "sharpe": float("-inf"),
                "t_stat": 0.0,
                "p_value": 1.0,
                "n": 0,
                "n_trades": 0,
            }

    k = min(top_k, cand_pool.shape[0])
   
    part_idx = np.argpartition(-sr_all[cand_pool], kth = k-1)[:k]
   
    top_global_idx = cand_pool[part_idx]

    sr_top = sr_all[top_global_idx]
    
    tt_top = np.concatenate([d["tt"] for d in per_win_stats])[top_global_idx]
    
    pv_top = np.concatenate([d["pv"] for d in per_win_stats])[top_global_idx]
    
    flips_all = np.concatenate([d["flips"] for d in per_win_stats])
    
    flips_top = flips_all[top_global_idx]

    j_top = _best_index(
        sharpe = sr_top, 
        t_stat = tt_top, 
        p_val = pv_top,
        min_sharpe = MIN_SHARPE_FOR_SIG, 
        alpha = SIGNIFICANCE_ALPHA
    )

    jg = int(top_global_idx[j_top])
    
    win_idx = int(np.searchsorted(offsets, jg, side = "right") - 1)
    
    j_within = int(jg - offsets[win_idx])
    
    Nb, Ns = per_win_shapes[win_idx]
    
    bi = j_within // Ns
    
    si = j_within %  Ns

    w = per_win_stats[win_idx]["w"]
   
    bt = float(B[bi])
   
    st = float(S[si])

    sharpe = float(sr_top[j_top])
    
    t_stat = float(tt_top[j_top])
    
    p_val = float(pv_top[j_top])
    
    n = int(per_win_stats[win_idx]["n"])
    
    flips = int(flips_top[j_top])

    return {
        "window": int(w),
        "buy_thresh": bt,
        "sell_thresh": st,
        "sharpe": sharpe,
        "t_stat": t_stat,
        "p_value": p_val,
        "n": n,
        "n_trades": flips,
    }


def eval_stoch_grid(
    pc: Precomp, 
    lower_grid: np.ndarray, 
    upper_grid: np.ndarray,
    is_pro_trend: bool, 
    is_anti_trend: bool, 
    vol_required: bool,
    signal_lag: int = SIGNAL_LAG, 
    cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate stochastic oscillator signals conditioned by lower/upper bounds.

    Oscillator
    ----------
    
        %K_t = 100 * (close_t − rolling_min(low, K)) / (rolling_max(high, K) − rolling_min(low, K))
    
        %D_t = SMA_D(%K_t)

    Signals
    -------
    - CrossUp: %K crosses above %D, i.e., (%K_t > %D_t) and (%K_{t−1} <= %D_{t−1}).
    
    - CrossDown: %K crosses below %D.

    Buy pulse: 
    
        ( %K_t < lower ) and CrossUp.
    
    Sell pulse: 
    
        ( %K_t > upper ) and CrossDown.

    Vectorisation
    -------------
    For lower grid L and upper grid U, the Cartesian product is formed by repeating
    and tiling the two pulse matrices, yielding an efficient broadcast across all (L, U).

    Backtest and statistics follow the standard pattern (lag/cooldown, gating, Sharpe, t-test).
    """

    k = pc.stoch_k.astype(np.float64)[:, None]
    
    d = pc.stoch_d.astype(np.float64)[:, None]
    
    cross_up = crosses_over(
        x = k,
        y = d
    )    
    
    cross_down = crosses_under(
        x = k, 
        y = d
    )   
    
    L = lower_grid.astype(np.float64)    
   
    U = upper_grid.astype(np.float64)    
   
    buy_l = (k < L[None, :]) & cross_up    
   
    sell_u = (k > U[None, :]) & cross_down   
   
    Nl = buy_l.shape[1]
    
    Nu = sell_u.shape[1]
   
    buy_mat = np.repeat(buy_l, Nu, axis = 1)
    
    sell_mat = np.tile(sell_u, (1, Nl))
    
    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend,
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
    
    if not gate.all():
       
        buy_mat &= gate[:, None]
       
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
    
    sell_mat = np.ascontiguousarray(sell_mat)    
            
    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat, 
        signal_lag = signal_lag, 
        cooldown_bars = cooldown, 
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos, 
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe, 
        t_stat = t_stat,
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG, 
        alpha = SIGNIFICANCE_ALPHA,    
    )

    li = j // Nu
   
    ui = j % Nu

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
   
    return {
        "lower": float(L[li]), 
        "upper": float(U[ui]),
        "sharpe": float(sharpe[j]), 
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]),
        "n": int(n), 
        "n_trades": flips
    }


@njit(parallel = True, cache = True, fastmath = True)
def eval_rsi_cartesian_kernel(
    rsi, 
    r, 
    buys, 
    sells, 
    signal_lag, 
    cooldown,
    mode_id, 
    annualisation
):
    """
    Numba kernel to scan RSI threshold grids with inline backtesting.

    Signals
    -------
    - Buy pulse if RSI crosses up through buy threshold: 
    
        (RSI_t > b) and (RSI_{t−1} <= b).
   
    - Sell pulse if RSI crosses down through sell threshold: 
    
        (RSI_t < s) and (RSI_{t−1} >= s).

    The kernel maintains a per-grid state machine identical to `pulses_to_positions`
    to generate positions with lag and cooldown, then computes:
  
    - Annualised Sharpe = sqrt(A) * mean(X) / std(X),
  
    - t-statistic = mean(X) / (std(X) / sqrt(n)),
  
    - one-sided p-value via erfc, and
  
    - number of flips.

    NaNs in RSI or returns suppress signal generation and contribution to statistics.
    """
   
    T = rsi.shape[0]
   
    Nb = buys.shape[0]
    
    Ns = sells.shape[0]
    
    sharpe = np.full((Nb, Ns), -np.inf)
    
    tstat = np.zeros((Nb, Ns))
    
    pval = np.ones((Nb, Ns))
    
    flips = np.zeros((Nb, Ns), dtype = np.int32)

    r_eff = r[1:]
  
    valid_eff = ~np.isnan(r_eff)
  
    n_valid = 0
  
    for t in range(r_eff.shape[0]):
  
        if valid_eff[t]:
  
            n_valid += 1

    for ib in prange(Nb):
  
        bt = buys[ib]
  
        for is_ in range(Ns):
  
            st = sells[is_]

            pos_t = 0
  
            cool = 0
  
            flips_count = 0
  
            prev_pos = 0

            prev_rsi = np.nan
  
            pos_series = np.zeros(T, np.int8)

            for t in range(T):
  
                tt = t - signal_lag
  
                b = False
                
                s = False
                
                if tt >= 0 and cool == 0:
                
                    cur = rsi[tt]
                
                    if not np.isnan(cur) and not np.isnan(prev_rsi):
                
                        b = (cur > bt) and (prev_rsi <= bt)
                
                        s = (cur < st) and (prev_rsi >= st)
                
                if cool > 0:
                
                    cool -= 1

                if b and s:
                
                    if mode_id == 0:  
                
                        b = False
                
                    elif mode_id == 1:
                
                        s = False
                
                    else:          
                
                        b = False
                        
                        s = False

                if b:
              
                    pos_t = 1
                    
                    cool = cooldown
              
                elif s:
              
                    pos_t = -1
                    
                    cool = cooldown

                pos_series[t] = pos_t

                if t > 0 and pos_series[t] != prev_pos:
             
                    flips_count += 1
             
                prev_pos = pos_series[t]

                prev_rsi = rsi[t]  

            mean = 0.0
            
            m2 = 0.0
            
            cnt = 0
            
            for t in range(1, T):
               
                if not np.isnan(r[t]):
               
                    x = float(pos_series[t - 1]) * float(r[t])
                 
                    cnt += 1
                 
                    delta = x - mean
                 
                    mean += delta / cnt
                 
                    m2 += delta * (x - mean)

            if cnt <= 1:
               
                sharpe[ib, is_] = -np.inf
               
                tstat[ib, is_] = 0.0
               
                pval[ib, is_] = 1.0
               
                flips[ib, is_] = flips_count
               
                continue

            var = m2 / (cnt - 1)

            sd = np.sqrt(var) if var > 0 else np.nan

            if not np.isfinite(sd) or sd <= 0:

                sharpe[ib, is_] = -np.inf

                tstat[ib, is_] = 0.0

                pval[ib, is_] = 1.0

                flips[ib, is_] = flips_count

                continue

            sr = np.sqrt(annualisation) * (mean / sd)

            t = mean / (sd / np.sqrt(cnt))

            sharpe[ib, is_] = sr

            tstat[ib, is_] = t

            z = max(t, 0.0)

            p = 0.5 * erfc(z / np.sqrt(2.0))

            pval[ib, is_] = p

            flips[ib, is_] = flips_count

    return sharpe, tstat, pval, flips, n_valid


def eval_rsi_grid(
    pc: Precomp,
    buy_grid: np.ndarray,
    sell_grid: np.ndarray,
    window_grid: np.ndarray,
    is_pro_trend: bool, is_anti_trend: bool, vol_required: bool,
    signal_lag: int = SIGNAL_LAG, cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE,
    *,
    top_k: int = 128,
):
    """
    Evaluate RSI threshold grids across multiple windows with significance-aware selection.

    Procedure
    ---------
    1) For each window, compute RSI (from cache if available) and apply optional gates.

    2) Invoke `eval_rsi_cartesian_kernel` for that window.

    3) Concatenate window-level results and select the best configuration via
    `_best_index`, favouring p < alpha and Sharpe >= minimum.

    4) Return the best window and thresholds together with performance statistics.

    Returns
    -------
    dict with keys: "window", "buy_thresh", "sell_thresh",
    "sharpe", "t_stat", "p_value", "n", "n_trades".
    """

    B = buy_grid.astype(np.float64)
  
    S = sell_grid.astype(np.float64)
  
    W = np.array(window_grid, dtype = int)

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend,
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )

    per_win_stats = []  
   
    per_win_shapes = [] 

    for w in W:
       
        rsi = _rsi_cached(
            pc = pc,
            windows = int(w)
        ).astype(np.float64)
       
        if not gate.all():
       
            rsi = rsi.copy()
       
            rsi[~gate] = np.nan

        sr, tt, pv, flips, n = eval_rsi_cartesian_kernel(
            rsi = rsi,
            r = pc.r.astype(np.float64),
            buys = B,
            sells = S,
            signal_lag = signal_lag,
            cooldown = cooldown,
            mode_id = _mode_id(
                mode = conflict_mode
            ),
            annualisation = float(ANNUALISATION)
        )
      
        Nb, Ns = sr.shape
      
        per_win_stats.append({
            "w": int(w),
            "sr": sr.ravel(),
            "tt": tt.ravel(),
            "pv": pv.ravel(),
            "flips": flips.ravel().astype(np.int64),
            "n": int(n),
        })
       
        per_win_shapes.append((Nb, Ns))

    if not per_win_stats:
      
        return {
            "window": int(W[0] if len(W) else RSI_WINDOW),
            "buy_thresh": float(B[0] if len(B) else RSI_BUY_THRESH),
            "sell_thresh": float(S[0] if len(S) else RSI_SELL_THRESH),
            "sharpe": float("-inf"),
            "t_stat": 0.0,
            "p_value": 1.0,
            "n": 0,
            "n_trades": 0,
        }

    sr_all = np.concatenate([d["sr"] for d in per_win_stats])  
   
    offsets = np.cumsum([0] + [d["sr"].shape[0] for d in per_win_stats])

    finite_mask = np.isfinite(sr_all)
   
    if finite_mask.any():
   
        cand_pool = np.flatnonzero(finite_mask)
   
    else:
   
        try:
   
            j_global = int(np.nanargmax(sr_all))
   
            cand_pool = np.array([j_global], dtype=int)
   
        except Exception:
   
            return {
                "window": int(W[0] if len(W) else RSI_WINDOW),
                "buy_thresh": float(B[0] if len(B) else RSI_BUY_THRESH),
                "sell_thresh": float(S[0] if len(S) else RSI_SELL_THRESH),
                "sharpe": float("-inf"),
                "t_stat": 0.0,
                "p_value": 1.0,
                "n": 0,
                "n_trades": 0,
            }

    k = min(top_k, cand_pool.shape[0])
   
    part_idx = np.argpartition(-sr_all[cand_pool], kth = k-1)[:k]
  
    top_global_idx = cand_pool[part_idx]

    sr_top = sr_all[top_global_idx]
  
    tt_top = np.concatenate([d["tt"] for d in per_win_stats])[top_global_idx]
  
    pv_top = np.concatenate([d["pv"] for d in per_win_stats])[top_global_idx]
  
    flips_all = np.concatenate([d["flips"] for d in per_win_stats])
  
    flips_top = flips_all[top_global_idx]

    j_top = _best_index(
        sharpe = sr_top,
        t_stat = tt_top,
        p_val = pv_top,
        min_sharpe = MIN_SHARPE_FOR_SIG,
        alpha = SIGNIFICANCE_ALPHA,
    )

    jg = int(top_global_idx[j_top])
  
    win_idx = int(np.searchsorted(offsets, jg, side = "right") - 1)
   
    j_within = int(jg - offsets[win_idx])
   
    Nb, Ns = per_win_shapes[win_idx]
   
    bi = j_within // Ns
   
    si = j_within %  Ns

    w = per_win_stats[win_idx]["w"]
    
    bt = float(B[bi])
    
    st = float(S[si])

    sharpe = float(sr_top[j_top])
    
    t_stat = float(tt_top[j_top])
    
    p_val = float(pv_top[j_top])
    
    n = int(per_win_stats[win_idx]["n"])
    
    flips = int(flips_top[j_top])

    return {
        "window": int(w),
        "buy_thresh": bt,
        "sell_thresh": st,
        "sharpe": sharpe,
        "t_stat": t_stat,
        "p_value": p_val,
        "n": n,
        "n_trades": flips,
    }


def eval_ema_grid(
    pc: Precomp, 
    fast_grid: np.ndarray,
    slow_grid: np.ndarray,
    is_pro_trend: bool, 
    is_anti_trend: bool, 
    vol_required: bool,
    signal_lag: int = SIGNAL_LAG, 
    cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate EMA crossover strategies over grids of (fast, slow) spans.

    Signals
    -------
    - Buy pulse: EMA_fast crosses up EMA_slow.
   
    - Sell pulse: EMA_fast crosses down EMA_slow.
   
    Formally at time t:
   
        buy:  (EMA_f_t > EMA_s_t) and (EMA_f_{t−1} <= EMA_s_{t−1})
   
        sell: (EMA_f_t < EMA_s_t) and (EMA_f_{t−1} >= EMA_s_{t−1})

    Constraints
    -----------
    Only pairs with fast < slow are considered (to maintain the usual MACD-style setup).

    Backtest
    --------
    - Pulses are gated, lagged, cooled down, then converted to positions.
   
    - Statistics are computed via `_eval_stats_from_pos`.
   
    - The best pair is chosen by `_best_index`.

    Returns the chosen (fast, slow) and associated performance metrics.
    """

    F = np.array(fast_grid, dtype = int)

    S = np.array(slow_grid, dtype = int)

    pairs = [(f, s) for f in F for s in S if f < s]

    if not pairs:
     
        return {
            "fast": EMA_FAST, 
            "slow": EMA_SLOW,
            "sharpe": -np.inf,
            "t_stat": 0.0, 
            "p_value": 1.0,
            "n": 0, 
            "n_trades": 0
        }

    uniq = sorted(set(list(F) + list(S)))
    
    ema_map = {
        span: _ema_cached(
            pc = pc, 
            span = int(span)
        ) for span in uniq
    }

    T = pc.c.shape[0]
   
    P = len(pairs)
   
    buy_mat = np.zeros((T, P), dtype = bool)
   
    sell_mat = np.zeros((T, P), dtype = bool)

    for j, (f, s) in enumerate(pairs):
     
        ef = ema_map[f][:, None]
     
        es = ema_map[s][:, None]
     
        ef_prev = np.roll(ef, 1, axis = 0)
        
        ef_prev[0, :] = np.nan
     
        es_prev = np.roll(es, 1, axis=0)
        
        es_prev[0, :] = np.nan
     
        buy_mat[:, j:j+1]  = (ef > es) & (ef_prev <= es_prev)
     
        sell_mat[:, j:j+1] = (ef < es) & (ef_prev >= es_prev)

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
    
    if not gate.all():
    
        buy_mat &= gate[:, None]
        
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
   
    sell_mat = np.ascontiguousarray(sell_mat)

    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat,
        signal_lag = signal_lag, 
        cooldown_bars = cooldown,
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos, 
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe,
        t_stat = t_stat, 
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG,
        alpha = SIGNIFICANCE_ALPHA
    )

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
    
    best_f, best_s = pairs[j]
    
    return {
        "fast": int(best_f), 
        "slow": int(best_s),
        "sharpe": float(sharpe[j]), 
        "t_stat": float(t_stat[j]),
         "p_value": float(p_val[j]),
        "n": int(n), "n_trades": flips
    }


def eval_macd_grid(
    pc: Precomp, 
    fast_grid: np.ndarray, 
    slow_grid: np.ndarray,
    signal_grid: np.ndarray,
    is_pro_trend: bool, 
    is_anti_trend: bool, 
    vol_required: bool,
    signal_lag: int = SIGNAL_LAG, 
    cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate MACD line/signal crossovers across grids of (fast, slow, signal) spans.

    Definitions
    -----------
    - MACD line = EMA_fast(close) − EMA_slow(close).
   
    - Signal = EMA_signal(MACD line).

    Signals
    -------
    - Buy pulse: 
    
        MACD line crosses up the Signal.
   
    - Sell pulse:
    
        MACD line crosses down the Signal.

    Implementation
    --------------
    - The MACD line for each (fast, slow) pair is cached and reused across the signal
    grid to avoid recomputation.
   
    - Pulses are gated and converted into positions with lag and cooldown.
   
    - Performance is measured via `_eval_stats_from_pos` and selected with `_best_index`.

    Returns the best triple and its statistics, including the flip count.
    """

    F = np.array(fast_grid, dtype = int)
    
    S = np.array(slow_grid, dtype = int)
    
    G = np.array(signal_grid, dtype = int)
    
    pairs = [(f, s) for f in F for s in S if f < s]
    
    if not pairs:
    
        return {
            "fast": EMA_FAST, 
            "slow": EMA_SLOW, 
            "signal": 9, 
            "sharpe": -np.inf,
            "t_stat": 0.0,
            "p_value": 1.0, 
            "n": 0,
            "n_trades": 0
        }

    line_map = {}

    for (f, s) in pairs:

        line_map[(f, s)] = _macd_line_cached(
            pc = pc, 
            fast = f,
            slow = s
        )

    T = pc.c.shape[0]
  
    P = len(pairs) * len(G)
  
    buy_mat = np.zeros((T, P), dtype = bool)
  
    sell_mat = np.zeros((T, P), dtype = bool)

    col = 0
   
    for (f, s) in pairs:
   
        line = line_map[(f, s)]
   
        for g in G:
   
            _, sig = _macd_signal_cached(
                pc = pc,
                fast = f, 
                slow = s, 
                signal = g
            )
   
            ln = line[:, None]
            
            sg = sig[:, None]
   
            ln_prev = np.roll(ln, 1, axis = 0)
            
            ln_prev[0, :] = np.nan
   
            sg_prev = np.roll(sg, 1, axis = 0)
            
            sg_prev[0, :] = np.nan

            buy_mat[:, col:col+1]  = (ln > sg) & (ln_prev <= sg_prev)
           
            sell_mat[:, col:col+1] = (ln < sg) & (ln_prev >= sg_prev)
           
            col += 1

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
    
    if not gate.all():
    
        buy_mat &= gate[:, None]
        
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
  
    sell_mat = np.ascontiguousarray(sell_mat)

    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat,
        signal_lag = signal_lag,
        cooldown_bars = cooldown,
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos,
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe,
        t_stat = t_stat, 
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG, 
        alpha = SIGNIFICANCE_ALPHA
    )

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
    
    pair_idx = j // len(G)
    
    sig_idx  = j % len(G)
    
    best_f, best_s = pairs[pair_idx]
    
    best_sig = int(G[sig_idx])
    
    return {
        "fast": int(best_f),
        "slow": int(best_s), 
        "signal": best_sig,
        "sharpe": float(sharpe[j]),
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]),
        "n": int(n), 
        "n_trades": flips
    }


def eval_vwap_grid(
    pc: Precomp, 
    window_grid: np.ndarray,
    is_pro_trend: bool,
    is_anti_trend: bool,
    vol_required: bool,
    signal_lag: int = SIGNAL_LAG,
    cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate rolling VWAP crossovers across window sizes.

    Rolling VWAP
    ------------
    For window w:
    
        num_t = Σ_{k=t−w+1..t} (typical_price_k * volume_k),
    
        den_t = Σ_{k=t−w+1..t} volume_k,
    
        RVWAP_t = num_t / den_t  (undefined if den_t == 0).

    `typical_price_k` is (high_k + low_k + close_k) / 3.

    Signals
    -------
    - Buy pulse: close crosses up RVWAP.
    
    - Sell pulse: close crosses down RVWAP.

    Gating, backtest, and statistics use the standard mechanisms. The best window is
    selected using `_best_index`.
    """

    W = np.array(window_grid, dtype = int)

    T = pc.c.shape[0]

    P = len(W)

    buy_mat = np.zeros((T, P), dtype = bool)
    
    sell_mat = np.zeros((T, P), dtype = bool)

    for j, w in enumerate(W):
      
        num = _rolling_sum_from_cumsum(
            cs = pc._cs_pv,
            w = w
        )
      
        den = _rolling_sum_from_cumsum(
            cs = pc._cs_v, 
            w = w
        )
      
        rvwap = (num / np.where(den == 0.0, np.nan, den))[:, None]
      
        c2 = pc.c[:, None]
      
        buy_mat[:, j:j+1] = crosses_over(
            x = c2, 
            y = rvwap
        )
      
        sell_mat[:, j:j + 1] = crosses_under(
            x = c2,
            y = rvwap
        )

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
    
    if not gate.all():
    
        buy_mat &= gate[:, None]
    
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
    
    sell_mat = np.ascontiguousarray(sell_mat)

    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat,
        signal_lag = signal_lag, 
        cooldown_bars = cooldown,
        mode_id = _mode_id(conflict_mode)
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos, 
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe,
        t_stat = t_stat,
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG,
        alpha = SIGNIFICANCE_ALPHA
    )

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
    
    return {
        "window": int(W[j]),
        "sharpe": float(sharpe[j]),
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]), 
        "n": int(n), 
        "n_trades": flips
    }


def eval_obv_grid(
    pc: Precomp,
    lookback_grid: np.ndarray,
    is_pro_trend: bool,
    is_anti_trend: bool,
    vol_required: bool,
    signal_lag: int = SIGNAL_LAG,
    cooldown: int = COOLDOWN_BARS,
    conflict_mode: str = CONFLICT_MODE
):
    """
    Evaluate OBV divergence signals over a grid of lookback windows.

    Divergence Logic
    ----------------
    For window w, define rolling, one-bar-lagged extrema:
   
    - price_low_prev  = min(close over last w bars) at t−1,
   
    - price_high_prev = max(close over last w bars) at t−1,
   
    - obv_low_prev    = min(OBV over last w bars) at t−1,
   
    - obv_high_prev   = max(OBV over last w bars) at t−1.

    Pulses:
   
    - Buy divergence:  
    
        price makes a lower low but OBV does not make a lower low:
   
        (close_t < price_low_prev_t) and (OBV_t >= obv_low_prev_t).
   
    - Sell divergence: 
    
        price makes a higher high but OBV does not:
       
        (close_t > price_high_prev_t) and (OBV_t <= obv_high_prev_t).

    These pattern-based pulses are gated, turned into positions, and scored.
    The best lookback is selected via `_best_index`.
    """

    L = np.array(lookback_grid, dtype = int)
   
    T = pc.c.shape[0]
   
    P = len(L)

    obv = pc.obv  

    buy_mat = np.zeros((T, P), dtype=bool)
   
    sell_mat = np.zeros((T, P), dtype=bool)

    for j, w in enumerate(L):
   
        price_low_prev, price_high_prev = _price_extrema_cached(
            pc = pc,
            window = w
        )
   
        obv_low_prev, obv_high_prev = _obv_extrema_cached(
            pc = pc,
            window = w
        )
   
        c2  = pc.c[:, None]
   
        obv2 = obv[:, None]
   
        buy_mat[:, j:j+1]  = (c2 < price_low_prev) & (obv2 >= obv_low_prev)
   
        sell_mat[:, j:j+1] = (c2 > price_high_prev) & (obv2 <= obv_high_prev)

    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro_trend, 
        is_anti_trend = is_anti_trend, 
        vol_required = vol_required
    )
    
    if not gate.all():
    
        buy_mat &= gate[:, None]
    
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
    
    sell_mat = np.ascontiguousarray(sell_mat)

    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat,
        signal_lag = signal_lag,
        cooldown_bars = cooldown,
        mode_id = _mode_id(
            mode = conflict_mode
        )
    )
    
    sharpe, t_stat, p_val, n = _eval_stats_from_pos(
        pos = pos, 
        r = pc.r,
        annualisation = ANNUALISATION
    )

    j = _best_index(
        sharpe = sharpe,
        t_stat = t_stat,
        p_val = p_val,
        min_sharpe = MIN_SHARPE_FOR_SIG, 
        alpha = SIGNIFICANCE_ALPHA
    )

    flips = int(np.sum(np.abs(np.diff(pos[:, j].astype(np.int16))) > 0))
    
    return {
        "lookback": int(L[j]),
        "sharpe": float(sharpe[j]), 
        "t_stat": float(t_stat[j]),
        "p_value": float(p_val[j]),
        "n": int(n), 
        "n_trades": flips
    }


def _coarse_then_full_mfi(
    pc: Precomp, 
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage MFI parameter tuning with an optional local refinement.

    Stage 1 (Coarse Screen)
    -----------------------
    Evaluate a small grid (`COARSE_GRIDS["mfi"]`) across all MFI window sizes.
    If the best coarse result does not satisfy significance (p < SCREEN_ALPHA and
    Sharpe >= SCREEN_MIN_SHARPE), return it and mark `passed=False`.

    Stage 2 (Refinement)
    --------------------
    If `REFINE_STRATEGY == "local"`, restrict buy/sell grids to the index-neighbourhood
    around the coarse best using `_neighbor_window` with width `LOCAL_WIDTH`.
    Otherwise, scan the full grids. Return the refined best and `passed=True`.

    Gating flags (trend/liquidity) are honoured in both stages.
    """

    g = COARSE_GRIDS["mfi"]

    coarse = eval_mfi_grid(
        pc = pc,
        buy_grid = np.array(g["buy"], dtype = float),
        sell_grid = np.array(g["sell"], dtype = float),
        window_grid = np.array(MFI_WINDOW_GRID),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti, 
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
    
    if not _is_sig(
        sharpe = coarse["sharpe"], 
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE, 
        alpha = SCREEN_ALPHA
    ):
        
        return coarse, False

    
    if REFINE_STRATEGY == "local":
    
        buy_grid = _neighbor_window(
            full_grid = MFI_BUY_GRID,  
            center = coarse["buy_thresh"],  
            width = LOCAL_WIDTH
        )
        
        sell_grid = _neighbor_window(
            full_grid = MFI_SELL_GRID, 
            center = coarse["sell_thresh"], 
            width = LOCAL_WIDTH)
   
    else:
   
        buy_grid = MFI_BUY_GRID
        
        sell_grid = MFI_SELL_GRID
   
    full = eval_mfi_grid(
        pc = pc,
        buy_grid = np.array(buy_grid, dtype = float),
        sell_grid = np.array(sell_grid, dtype = float),
        window_grid = np.array(MFI_WINDOW_GRID),
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_bb(
    pc: Precomp, 
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning: coarse screen then refine (full or local neighborhood).

    Parameters
    ----------
    pc : Precomp
    is_pro, is_anti, vol_req : bool
        Gating flags.
    Returns
    -------
    (best_dict, passed) : tuple
        best_dict : dict of the same shape returned by the corresponding `eval_*`.
        passed    : bool indicating whether the coarse screen met significance
                    (`SCREEN_MIN_SHARPE`, `SCREEN_ALPHA`) and thus refinement ran.

    Strategy
    --------
    1) Evaluate a small, cheap COARSE_GRIDS set.
    
    2) If best coarse result is not significant, return it with passed=False.
    
    3) Else, refine:
    
    - if REFINE_STRATEGY == "full": run the full grid;
    
    - if "local": restrict to ±LOCAL_WIDTH neighbors around the coarse best.
    """

    g = COARSE_GRIDS["bb"]

    coarse = eval_bb_grid(
        pc = pc,
        num_std_grid = np.array(g["num_std"], dtype = float),
        window_grid = np.array(BB_WINDOW_GRID),
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti, 
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
    
    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE, 
        alpha = SCREEN_ALPHA
    ):
        
        return coarse, False

    if REFINE_STRATEGY == "local":
   
        grid = _neighbor_window(
            full_grid = BB_STD_GRID,
            center = coarse["num_std"], 
            width = LOCAL_WIDTH
        )
   
    else:
   
        grid = BB_STD_GRID
   
    full = eval_bb_grid(
        pc = pc, 
        num_std_grid = np.array(grid, dtype = float),
        window_grid = np.array(BB_WINDOW_GRID),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti, 
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_atr(
    pc: Precomp,
    atr_win_grid: list[int] | np.ndarray,
    break_win_grid: list[int] | np.ndarray,
    is_pro: bool, is_anti: bool, vol_req: bool
):
    """
    Two-stage ATR breakout tuning across separate ATR and breakout windows.

    Stage 1
    -------
    Scan coarse multipliers for each combination of `atr_win_grid` and `break_win_grid`;
    select the best coarse configuration by Sharpe.

    Stage 2
    -------
    Refine the multiplier over the full `ATR_MULT_GRID` while fixing the windows to
    those that produced the best coarse result.

    Returns
    -------
    dict
        Best refined configuration with keys: "atr_window", "break_window", "mult",
        and associated statistics.
    """

    coarse_M = np.array(COARSE_GRIDS["atr"]["mult"], dtype = float)
  
    best = None
  
    for aw in atr_win_grid:
  
        for bw in break_win_grid:
  
            res = eval_atr_grid(
                pc = pc, 
                mult_grid = coarse_M,
                atr_window = aw, 
                break_window = bw,
                is_pro_trend = is_pro, 
                is_anti_trend = is_anti, 
                vol_required = vol_req
            )
            
            if (best is None) or (res["sharpe"] > best["sharpe"]):
            
                best = res

    refined = eval_atr_grid(
        pc = pc,
        mult_grid = np.array(ATR_MULT_GRID, dtype = float),
        atr_window = best["atr_window"],
        break_window = best["break_window"],
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti, 
        vol_required = vol_req
    )
    
    return refined


def _coarse_then_full_rsi(
    pc: Precomp,
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning: coarse screen then refine (full or local neighborhood).

    Parameters
    ----------
    pc : Precomp
    is_pro, is_anti, vol_req : bool
        Gating flags.
    Returns
    -------
    (best_dict, passed) : tuple
        best_dict : dict of the same shape returned by the corresponding `eval_*`.
        passed    : bool indicating whether the coarse screen met significance
                    (`SCREEN_MIN_SHARPE`, `SCREEN_ALPHA`) and thus refinement ran.

    Strategy
    --------
    1) Evaluate a small, cheap COARSE_GRIDS set.
   
    2) If best coarse result is not significant, return it with passed=False.
   
    3) Else, refine:
   
    - if REFINE_STRATEGY == "full": run the full grid;
   
    - if "local": restrict to ±LOCAL_WIDTH neighbors around the coarse best.
    
    - If significant, refine either locally (±LOCAL_WIDTH in the grid index) or
    across full buy/sell grids. Return (best_dict, passed).
    """

    g = COARSE_GRIDS["rsi"]

    coarse = eval_rsi_grid(
        pc = pc,
        buy_grid = np.array(g["buy"], dtype = float),
        sell_grid = np.array(g["sell"], dtype = float),
        window_grid = np.array(RSI_WINDOW_GRID, dtype = float),
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti, 
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
   
    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE, 
        alpha = SCREEN_ALPHA
    ):
    
        return coarse, False

    if REFINE_STRATEGY == "local":

        buy_grid = _neighbor_window(
            full_grid = RSI_BUY_GRID,  
            center = coarse["buy_thresh"], 
            width = LOCAL_WIDTH
        )

        sell_grid = _neighbor_window(
            full_grid = RSI_SELL_GRID,
            center = coarse["sell_thresh"],
            width = LOCAL_WIDTH
        )

    else:

        buy_grid = RSI_BUY_GRID
        
        sell_grid = RSI_SELL_GRID

    full = eval_rsi_grid(
        pc = pc,
        buy_grid = np.array(buy_grid, dtype = float),
        sell_grid = np.array(sell_grid, dtype = float),
        window_grid = np.array(RSI_WINDOW_GRID, dtype = float),
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti, 
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_stoch(
    pc: Precomp,
    is_pro: bool, 
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning: coarse screen then refine (full or local neighborhood).

    Parameters
    ----------
    pc : Precomp
    is_pro, is_anti, vol_req : bool
        Gating flags.
    Returns
    -------
    (best_dict, passed) : tuple
        best_dict : dict of the same shape returned by the corresponding `eval_*`.
        passed    : bool indicating whether the coarse screen met significance
                    (`SCREEN_MIN_SHARPE`, `SCREEN_ALPHA`) and thus refinement ran.

    Strategy
    --------
    1) Evaluate a small, cheap COARSE_GRIDS set.
    
    2) If best coarse result is not significant, return it with passed=False.
    
    3) Else, refine:
    
    - if REFINE_STRATEGY == "full": run the full grid;
    
    - if "local": restrict to ±LOCAL_WIDTH neighbors around the coarse best.
    
    - If significant, refine locally or fully, holding the oscillator windows fixed
    (STOCH_K and STOCH_D are part of the precomputation).
    """

    g = COARSE_GRIDS["stoch"]

    coarse = eval_stoch_grid(
        pc = pc,
        lower_grid = np.array(g["lower"], dtype = float),
        upper_grid = np.array(g["upper"], dtype = float),
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti,
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE,
        alpha = SCREEN_ALPHA
    ):
        
        return coarse, False

    
    if REFINE_STRATEGY == "local":
    
        lower_grid = _neighbor_window(
            full_grid = STOCH_LOW_GRID, 
            center = coarse["lower"], 
            width = LOCAL_WIDTH
        )
    
        upper_grid = _neighbor_window(
            full_grid = STOCH_HIGH_GRID, 
            center = coarse["upper"],
            width = LOCAL_WIDTH
        )
    
    else:
    
        lower_grid = STOCH_LOW_GRID
        
        upper_grid = STOCH_HIGH_GRID
    
    full = eval_stoch_grid(
        pc = pc,
        lower_grid = np.array(lower_grid, dtype = float),
        upper_grid = np.array(upper_grid, dtype = float),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req, 
        signal_lag = SIGNAL_LAG, 
        cooldown = COOLDOWN_BARS, 
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_ema(
    pc: Precomp,
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning for EMA spans.

    - Coarse screen over small sets of fast/slow spans.
    
    - If significant, refine across either the local neighbourhood or the full
    `EMA_FAST_GRID` × `EMA_SLOW_GRID` (with fast < slow).
    """

    g = COARSE_GRIDS["ema"]

    coarse = eval_ema_grid(
        pc = pc,
        fast_grid = np.array(g["fast"], dtype = int),
        slow_grid = np.array(g["slow"], dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )

    if not _is_sig(
        sharpe = coarse["sharpe"], 
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE,
        alpha = SCREEN_ALPHA
    ):
    
        return coarse, False

    if REFINE_STRATEGY == "local":
    
        fast_grid = _neighbor_window(
            full_grid = EMA_FAST_GRID,
            center = coarse["fast"], 
            width = LOCAL_WIDTH
        )
    
        slow_grid = _neighbor_window(
            full_grid = EMA_SLOW_GRID, 
            center = coarse["slow"], 
            width = LOCAL_WIDTH
        )
    
    else:
    
        fast_grid = EMA_FAST_GRID
    
        slow_grid = EMA_SLOW_GRID

    full = eval_ema_grid(
        pc = pc,
        fast_grid = np.array(fast_grid, dtype = int),
        slow_grid = np.array(slow_grid, dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_macd(
    pc: Precomp,
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning for MACD (fast, slow, signal) spans.

    - Coarse screen on small grids for all three spans.
   
    - If significant, refine locally or across the full grids as configured.
    """
    
    g = COARSE_GRIDS["macd"]

    coarse = eval_macd_grid(
        pc = pc,
        fast_grid = np.array(g["fast"], dtype = int),
        slow_grid = np.array(g["slow"], dtype = int),
        signal_grid = np.array(g["signal"], dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )

    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE, 
        alpha = SCREEN_ALPHA
    ):
    
        return coarse, False

    if REFINE_STRATEGY == "local":
    
        fast_grid = _neighbor_window(
            full_grid = EMA_FAST_GRID,     
            center = coarse["fast"],   
            width = LOCAL_WIDTH
        )
    
        slow_grid = _neighbor_window(
            full_grid = EMA_SLOW_GRID,     
            center = coarse["slow"],   
            width = LOCAL_WIDTH
        )
    
        signal_grid = _neighbor_window(
            full_grid = MACD_SIGNAL_GRID, 
            center = coarse["signal"], 
            width = LOCAL_WIDTH
        )
    
    else:
    
        fast_grid = EMA_FAST_GRID
    
        slow_grid = EMA_SLOW_GRID
    
        signal_grid = MACD_SIGNAL_GRID

    full = eval_macd_grid(
        pc = pc,
        fast_grid = np.array(fast_grid, dtype = int),
        slow_grid = np.array(slow_grid, dtype = int),
        signal_grid = np.array(signal_grid, dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_vwap(
    pc: Precomp,
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning for rolling VWAP window.

    - Coarse screen over a small set of windows.
   
    - If significant, refine locally or across the full `VWAP_WIN_GRID`.
    """
    
    g = COARSE_GRIDS["vwap"]

    coarse = eval_vwap_grid(
        pc = pc,
        window_grid = np.array(g["window"], dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )

    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE,
        alpha = SCREEN_ALPHA
    ):
        
        return coarse, False

    if REFINE_STRATEGY == "local":
    
        win_grid = _neighbor_window(
            full_grid = VWAP_WIN_GRID,
            center = coarse["window"],
            width = LOCAL_WIDTH
        )
    
    else:
    
        win_grid = VWAP_WIN_GRID

    full = eval_vwap_grid(
        pc = pc,
        window_grid = np.array(win_grid, dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def _coarse_then_full_obv(
    pc: Precomp,
    is_pro: bool,
    is_anti: bool,
    vol_req: bool
):
    """
    Two-stage tuning for OBV divergence lookback.

    - Coarse screen over a small set of lookbacks.
    - If significant, refine locally or over the full `OBV_LOOKBACK_SCAN`.
    """

    g = COARSE_GRIDS["obv"]

    coarse = eval_obv_grid(
        pc = pc,
        lookback_grid = np.array(g["lookback"], dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )

    if not _is_sig(
        sharpe = coarse["sharpe"],
        p_value = coarse["p_value"],
        min_sharpe = SCREEN_MIN_SHARPE, 
        alpha = SCREEN_ALPHA
    ):
    
        return coarse, False

    if REFINE_STRATEGY == "local":
      
        lb_grid = _neighbor_window(
            full_grid = OBV_LOOKBACK_SCAN, 
            center = coarse["lookback"], 
            width = LOCAL_WIDTH
        )
   
    else:
   
        lb_grid = OBV_LOOKBACK_SCAN

    full = eval_obv_grid(
        pc = pc,
        lookback_grid = np.array(lb_grid, dtype = int),
        is_pro_trend = is_pro,
        is_anti_trend = is_anti,
        vol_required = vol_req,
        signal_lag = SIGNAL_LAG,
        cooldown = COOLDOWN_BARS,
        conflict_mode = CONFLICT_MODE
    )
    
    return full, True


def weight_from_stats(
    sharpe, 
    t_stat,
    p_value,
    *,
    sr_bh_ann: float | None = None,
    alpha: float = SIGNIFICANCE_ALPHA,
    min_sr: float = MIN_SHARPE_FOR_SIG
) -> int:
    """
    Map performance versus buy-and-hold to a discrete weight in {0,1,2,3,4}.

    Inputs
    ------
    sharpe : float
   
        Annualised Sharpe of the indicator strategy.
   
    t_stat : float
   
        One-sided t-statistic for daily mean > 0 (used only for gate consistency).
   
    p_value : float
   
        One-sided p-value; if p_value >= alpha or sharpe < min_sr, weight is 0.
   
    sr_bh_ann : float | None
   
        Annualised buy-and-hold Sharpe. If None or invalid, treated as 0.
   
    alpha : float
   
        Significance level for the gate (default SIGNIFICANCE_ALPHA).
   
    min_sr : float
   
        Minimum Sharpe to pass the economic gate (default MIN_SHARPE_FOR_SIG).

    Rule
    ----
    If the gates pass, compute improvement = sharpe − sr_bh_ann and assign:
      
        improvement >= 1.00 → 4
      
        improvement >= 0.50 → 3
      
        improvement >= 0.25 → 2
      
        improvement >  0.00 → 1
      
        otherwise           → 0

    Rationale
    ---------
    This favours indicators that add risk-adjusted value over passive exposure, while
    discarding statistically weak or economically trivial configurations.
    """
   
    if not np.isfinite(sharpe):
   
        return 0

    if (p_value >= alpha) or (sharpe < min_sr):
   
        return 0

    sr_bh_ann = 0.0 if (sr_bh_ann is None or not np.isfinite(sr_bh_ann)) else sr_bh_ann
   
    improve = sharpe - sr_bh_ann
   
    if improve <= 0:
   
        return 0

    if improve >= 1.00: 
        
        return 4
   
    if improve >= 0.50: 
        
        return 3
   
    if improve >= 0.25: 
        
        return 2
   
    return 1


def default_params_for(
    name: str
) -> dict:
    """
    Return safe default parameters for a named indicator.

    Purpose
    -------
    Provides a deterministic fallback when calibration is unavailable or has failed.
    Values mirror canonical textbook defaults (e.g., RSI 14, 30/70; BB 20, 2σ; ATR 14).

    Returns
    -------
    dict
        Parameter dictionary appropriate for the given `name`. Unknown names return {}.
    """

    if name == "rsi":
       
        return {
            "buy_thresh": RSI_BUY_THRESH, 
            "sell_thresh": RSI_SELL_THRESH,
            "window": RSI_WINDOW
        }

    if name == "mfi":
       
        return {
            "buy_thresh": MFI_BUY_THRESH, 
            "sell_thresh": MFI_SELL_THRESH,
            "window": MFI_WINDOW
        }

    if name == "bb":
       
        return {
            "num_std": BB_STD, 
            "window": BB_WINDOW
        }

    if name == "atr":
       
        return {
            "mult": ATR_MULTIPLIER, 
            "atr_window": ATR_WINDOW,
            "break_window": ATR_BREAK_WINDOW
        }

    if name == "stoch":
       
        return {
            "lower": 20.0, 
            "upper": 80.0
        }

    if name == "ema":
       
        return {
            "fast": EMA_FAST,
            "slow": EMA_SLOW
        }

    if name == "macd":
        
        return {
            "fast": EMA_FAST,
            "slow": EMA_SLOW,
            "signal": 9
        }

    if name == "vwap":
       
        return {
            "window": VWAP_WINDOW
        }

    if name == "obv":

        return {
            "lookback": OBV_LOOKBACK,
            "mode": "divergence"
        }

    return {}


def tune_thresholds_if_significant(
    name: str,
    base: IndicatorResult,
    pc: Precomp
) -> IndicatorResult:
    """
    Calibrate a single indicator for one ticker, respecting optional coarse screening.

    Inputs
    ------
    name : str
    
        Indicator name among {"rsi","mfi","bb","atr","stoch","ema","macd","vwap","obv"}.
    
    base : IndicatorResult
    
        Baseline result used if calibration is disabled or coarse screen fails.
    
    pc : Precomp
    
        Precomputed arrays and caches for the ticker.

    Behaviour
    ---------
    - If `USE_COARSE_SCREEN` is True and the indicator supports it, run the
    corresponding two-stage tuner. If the coarse stage fails the significance gate,
    return defaults with the coarse statistics.
    
    - Otherwise, or if coarse screen passes, execute the full grid evaluation for
    the indicator and return the best parameter set with its performance.

    Returns
    -------
    IndicatorResult
    
        Populated with tuned parameters, annualised Sharpe, t-statistic, p-value,
        and an approximate `n_trades` (flip count).
    """

    is_pro = (name in PRO_TREND)
   
    is_anti = (name in ANTI_TREND)
   
    vol_req = (name in VOL_REQUIRED)

    if name == "rsi" and USE_COARSE_SCREEN:
       
        best, passed = _coarse_then_full_rsi(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"],
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
        
        return IndicatorResult(
            name = name,
            params = {
                "buy_thresh": best["buy_thresh"],
                "sell_thresh": best["sell_thresh"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "mfi" and USE_COARSE_SCREEN:
      
        best, passed = _coarse_then_full_mfi(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
      
        if not passed:
      
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"],
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
      
        return IndicatorResult(
            name = name,
            params = {
                "buy_thresh": best["buy_thresh"],
                "sell_thresh": best["sell_thresh"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "bb" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_bb(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti,
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name,
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"], 
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
        
        return IndicatorResult(
            name = name,
            params = {
                "num_std": best["num_std"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "atr" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_atr(
            pc = pc, 
            atr_win_grid = ATR_WINDOW_GRID,
            break_win_grid = ATR_BREAK_GRID, 
            is_pro = is_pro,
            is_anti = is_anti,
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"], 
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
        
        return IndicatorResult(
            name = name,
            params = {
                "mult": best["mult"],
                "atr_window": best["atr_window"],
                "break_window": best["break_window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "stoch" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_stoch(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"], 
                t_stat = best["t_stat"],
                p_value = best["p_value"],
                n_trades = best["n_trades"]
            )
        
        return IndicatorResult(
            name = name,
            params = {
                "lower": best["lower"],
                "upper": best["upper"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "ema" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_ema(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"],
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
            
        return IndicatorResult(
            name = name,
            params = {
                "fast": best["fast"], 
                "slow": best["slow"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "macd" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_macd(
            pc = pc, 
            is_pro = is_pro,
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
            
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"], 
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
            
        return IndicatorResult(
            name = name,
            params = {
                "fast": best["fast"], 
                "slow": best["slow"], 
                "signal": best["signal"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "vwap" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_vwap(
            pc = pc,
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(name),
                sharpe = best["sharpe"],
                t_stat = best["t_stat"],
                p_value = best["p_value"],
                n_trades = best["n_trades"]
            )

        return IndicatorResult(
            name = name,
            params = {
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "obv" and USE_COARSE_SCREEN:
        
        best, passed = _coarse_then_full_obv(
            pc = pc, 
            is_pro = is_pro, 
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        if not passed:
        
            return IndicatorResult(
                name = name, 
                params = default_params_for(
                    name = name
                ),
                sharpe = best["sharpe"], 
                t_stat = best["t_stat"],
                p_value = best["p_value"], 
                n_trades = best["n_trades"]
            )
      
        return IndicatorResult(
            name = name,
            params = {
                "lookback": best["lookback"], 
                "mode": "divergence"
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "rsi":
        
        best = eval_rsi_grid(
            pc = pc,
            buy_grid = np.array(RSI_BUY_GRID, dtype = float),
            sell_grid = np.array(RSI_SELL_GRID, dtype = float),
            window_grid = np.array(RSI_WINDOW_GRID, dtype = float),
            is_pro_trend = is_pro,
            is_anti_trend = is_anti,
            vol_required = vol_req,
            signal_lag = SIGNAL_LAG,
            cooldown = COOLDOWN_BARS,
            conflict_mode = CONFLICT_MODE
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "buy_thresh": best["buy_thresh"],
                "sell_thresh": best["sell_thresh"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "mfi":
        
        best = eval_mfi_grid(
            pc = pc,  
            buy_grid = np.array(MFI_BUY_GRID, dtype = float),
            sell_grid = np.array(MFI_SELL_GRID, dtype = float),
            window_grid = np.array(MFI_WINDOW_GRID, dtype = int),
            is_pro_trend = is_pro,
            is_anti_trend = is_anti,
            vol_required = vol_req,
            signal_lag = SIGNAL_LAG, 
            cooldown = COOLDOWN_BARS, 
            conflict_mode = CONFLICT_MODE
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "buy_thresh": best["buy_thresh"],
                "sell_thresh": best["sell_thresh"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "bb":
        
        best = eval_bb_grid(
            pc = pc,
            num_std_grid = np.array(BB_STD_GRID, dtype = float),
            window_grid = np.array(BB_WINDOW_GRID, dtype = int),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti, 
            vol_required = vol_req,
            signal_lag = SIGNAL_LAG, 
            cooldown = COOLDOWN_BARS, 
            conflict_mode = CONFLICT_MODE
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "num_std": best["num_std"],
                "window": best["window"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "atr":
        
        best = _coarse_then_full_atr(
            pc = pc, 
            atr_win_grid = ATR_WINDOW_GRID, 
            break_win_grid = ATR_BREAK_GRID,
            is_pro = is_pro,
            is_anti = is_anti, 
            vol_req = vol_req
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "mult": best["mult"],
                "atr_window": best["atr_window"],
                "break_window": best["break_window"]
            },
            sharpe = best["sharpe"],
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "stoch":
        
        best = eval_stoch_grid(
            pc = pc,
            lower_grid = np.array(STOCH_LOW_GRID, dtype = float),
            upper_grid = np.array(STOCH_HIGH_GRID, dtype = float),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti,
            vol_required = vol_req,
            signal_lag = SIGNAL_LAG, 
            cooldown = COOLDOWN_BARS,
            conflict_mode = CONFLICT_MODE
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "lower": best["lower"],
                "upper": best["upper"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "ema":
        
        best = eval_ema_grid(
            pc = pc,
            fast_grid = np.array(EMA_FAST_GRID, dtype = int),
            slow_grid = np.array(EMA_SLOW_GRID, dtype = int),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti, 
            vol_required = vol_req
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "fast": best["fast"], 
                "slow": best["slow"]
            },
            sharpe = best["sharpe"],
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "macd":
        
        best = eval_macd_grid(
            pc = pc,
            fast_grid = np.array(MACD_FAST_GRID, dtype = int),
            slow_grid = np.array(MACD_SLOW_GRID, dtype = int),
            signal_grid = np.array(MACD_SIGNAL_GRID, dtype = int),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti, 
            vol_required = vol_req
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "fast": best["fast"],
                "slow": best["slow"], 
                "signal": best["signal"]
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    if name == "vwap":
        
        best = eval_vwap_grid(
            pc = pc, 
            window_grid = np.array(VWAP_WIN_GRID, dtype = int),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti,
            vol_required = vol_req
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "window": best["window"]
            },
            sharpe = best["sharpe"],
            t_stat = best["t_stat"],
            p_value = best["p_value"], 
            n_trades = best["n_trades"]
        )

    if name == "obv":
        
        best = eval_obv_grid(
            pc = pc, 
            lookback_grid = np.array(OBV_LOOKBACK_SCAN, dtype = int),
            is_pro_trend = is_pro, 
            is_anti_trend = is_anti, 
            vol_required = vol_req
        )
        
        return IndicatorResult(
            name = name,
            params = {
                "lookback": best["lookback"],
                "mode": "divergence"
            },
            sharpe = best["sharpe"], 
            t_stat = best["t_stat"],
            p_value = best["p_value"],
            n_trades = best["n_trades"]
        )

    return base


def calibrate_ticker(
    pc: Precomp,
    c: pd.Series, 
    tk: str
) -> tuple[dict[str, dict], list[dict]]:
    """
    Run calibration for all indicators on a single ticker and compose persistence rows.

    Process
    -------
    1) Compute the buy-and-hold Sharpe via `_bh_sharpe_from_returns`.
   
    2) For each indicator in `INDICATOR_ORDER`:
   
    a) Tune parameters via `tune_thresholds_if_significant`.
   
    b) Convert the result to a relative weight using `weight_from_stats`,
        where improvement is measured versus buy-and-hold Sharpe.
   
    c) Collect a result entry and a calibration row ready for Excel persistence.
   
    3) Timestamps (Start, End, LastUpdated) are made tz-naïve for Excel.

    Outputs
    -------
    (results, rows)
   
    - results : dict[str, dict]
        {
        indicator: {
            "params": dict, "weight": int,
            "sharpe": float, "t_stat": float, "p_value": float,
            "n_trades": int
        }, ...
        }
   
    - rows : list[dict]
   
        Each row contains the canonical calibration columns for persistence.

    Notes on Backtest Methodology
    -----------------------------
    - Positions are generated from buy/sell pulses with explicit signal lag and
    cooldown to reduce look-ahead bias and over-trading.
   
    - Returns are daily simple returns; strategy returns use position_{t−1} * r_t.
   
    - Annualisation assumes 252 trading days per annum.
    """

    start = pd.Timestamp(c.index.min())
    
    end = pd.Timestamp(c.index.max())
    
    if start.tz is not None:
    
        start = start.tz_convert("UTC").tz_localize(None)
    
    if end.tz is not None:
    
        end = end.tz_convert("UTC").tz_localize(None)
    
    now_naive = pd.Timestamp.now(tz = "UTC").tz_localize(None)
    
    bh_sharpe = _bh_sharpe_from_returns(
        r = pc.r, 
        annualisation = ANNUALISATION
    )

    results: dict[str, dict] = {}
    
    rows: list[dict] = []

    for name in INDICATOR_ORDER:
        
        logger.info(f"Tuning {name} for {tk}")
       
        base = IndicatorResult(
            name = name, 
            params = default_params_for(name = name), 
            sharpe = 0.0, 
            t_stat = 0.0, 
            p_value = 1.0,
            n_trades = 0
        )
        
        best = tune_thresholds_if_significant(
            name = name,
            base = base, 
            pc = pc
        )
        
        w_rel = weight_from_stats(
            sharpe = best.sharpe,
            t_stat = best.t_stat,
            p_value = best.p_value,
            sr_bh_ann = bh_sharpe
        )

        results[name] = {
            "params": best.params,
            "weight": w_rel,
            "sharpe": best.sharpe,
            "t_stat": best.t_stat,
            "p_value": best.p_value,
            "n_trades": best.n_trades
        }

        rows.append({
            "Ticker": tk,
            "Indicator": name,
            "ParamJSON": json.dumps(best.params),
            "Weight": w_rel,      
            "Sharpe": best.sharpe,
            "SharpeBH": bh_sharpe,    
            "TStat": best.t_stat,
            "PValue": best.p_value,
            "NTrades": best.n_trades,
            "Start": start, 
            "End": end, 
            "LastUpdated": now_naive
        })

    return results, rows


def _pos_for_indicator(
    name: str, 
    params: dict,
    pc: Precomp
) -> np.ndarray:
    """
    Construct a single-scenario position series for one indicator and a given
    parameterisation, including gating and pulse-to-position mapping.

    Parameters
    ----------
    name : {"rsi","mfi","bb","atr","stoch","ema","macd","vwap","obv"}
        Indicator identifier.
    params : dict
        Indicator-specific parameters for the current ticker. Missing entries
        fall back to module defaults (e.g., `RSI_WINDOW`, `BB_STD`).
    pc : Precomp
        Precomputed arrays for this ticker (prices, returns, cached indicators,
        trend and volume gates, and auxiliary extrema).

    Returns
    -------
    np.ndarray of int8, shape (T,)
        Position series `pos_t ∈ {-1, 0, +1}` produced by:
    
        1) building buy/sell pulse vectors at bar resolution for the specified
           indicator and parameters;
    
        2) applying gating masks:
    
           - pro-trend indicators are masked by `trend_mask`;
    
           - anti-trend indicators are masked by the complement of
             `trend_mask`;
    
           - volume-dependent indicators are masked by `vol_ok`;
    
        3) converting pulses to a persistent position via the Numba kernel
           `pulses_to_positions`, with global `SIGNAL_LAG`, `COOLDOWN_BARS`,
           and `CONFLICT_MODE`.

    Pulse Definitions (by indicator)
    --------------------------------
    All crosses are evaluated with prior values to avoid spurious triggers at
    initialisation, and thresholds derived from rolling constructs are shifted
    appropriately where needed.

    RSI (anti-trend):
    
        Let `RSI_t` be the Wilder RSI over window `w`.
    
        Buy pulse at `t` if `RSI_t > b` and `RSI_{t−1} <= b` (upward cross of
        `buy_thresh = b`). 
        
        Sell pulse if `RSI_t < s` and `RSI_{t−1} >= s`
        (downward cross of `sell_thresh = s`).

    MFI (neutral):
       
        Let `MFI_t` be Money Flow Index over window `w`.
       
        Buy at `MFI_t > b` and `MFI_{t−1} <= b`. Sell at `MFI_t < s` and
        `MFI_{t−1} >= s`.

    Bollinger Bands (anti-trend):
      
        Over window `w`, `mid_t = SMA_w(C)_t`, `sd_t = std_w(C)_t` (ddof=0).
      
        For multiplier `nσ`:
      
          `upper_t = mid_t + nσ * sd_t`, `lower_t = mid_t − nσ * sd_t`.
      
        Buy pulse if `C_t` crosses above `lower_t`:
      
          `(C_t > lower_t) and (C_{t−1} <= lower_{t−1})`.
      
        Sell pulse if `C_t` crosses below `upper_t`:
      
          `(C_t < upper_t) and (C_{t−1} >= upper_{t−1})`.

    ATR breakout (pro-trend):
      
        ATR over `w_atr`, prior extrema over `w_break`:
      
          `hi_prev_t = max(H_{t−w_break:t−1})`,
      
          `lo_prev_t = min(L_{t−w_break:t−1})`,
      
          `ATR_prev_t = ATR_{t−1}`.
      
        For multiplier `m`:
      
          `up_th_t = hi_prev_t + m * ATR_prev_t`,
      
          `dn_th_t = lo_prev_t − m * ATR_prev_t`.
      
        Buy pulse if `C_t > up_th_t`. Sell pulse if `C_t < dn_th_t`.

    Stochastic %K/%D (anti-trend):
      
        `%K_t = 100 * (C_t − min L) / (max H − min L)` over `k` bars; `%D` is
        SMA over `d` bars.
      
        Crosses are computed as `(K_t > D_t and K_{t−1} <= D_{t−1})` (up) and
        `(K_t < D_t and K_{t−1} >= D_{t−1})` (down).
      
        Buy pulse if `%K_t < lower` and cross up; sell pulse if `%K_t > upper`
        and cross down.

    EMA crossover (pro-trend):
      
        `EMA_f` and `EMA_s` with `f < s`. Buy pulse on `EMA_f` crossing above
        `EMA_s`; sell pulse on `EMA_f` crossing below `EMA_s`.

    MACD (pro-trend):
      
        MACD line `M_t = EMA_f(C)_t − EMA_s(C)_t`, signal `S_t = EMA_g(M)_t`.
      
        Buy pulse on `M` crossing above `S`; sell pulse on `M` crossing below `S`.

    Rolling VWAP (pro-trend):
      
        Over window `w`, `VWAP_t = (Σ TP_i * V_i) / (Σ V_i)` with
        `TP_i = (H_i + L_i + C_i) / 3`. 
        
        Buy pulse on cross above VWAP
        
        Sell on cross below VWAP.

    OBV divergence (pro-trend):
        
        OBV cumulative series, with price/OBV prior extrema over `w`:
        
            Buy pulse if `C_t < min(C_{t−w:t−1})` and `OBV_t >= min(OBV_{t − w:t − 1})`
            (bullish divergence).
        
            Sell pulse if `C_t > max(C_{t − w:t − 1})` and
            `OBV_t <= max(OBV_{t − w:t − 1})` (bearish divergence).

    Pulse → Position
    ----------------
    Given buy and sell boolean pulses, the kernel applies:
   
    - effective time `t_eff = t − SIGNAL_LAG` (if negative, ignored),
   
    - a per-scenario cooldown `COOLDOWN_BARS` after any flip,
   
    - conflict handling at `t_eff` according to:
   
      "sell_wins" (sell overrides buy), "buy_wins" (buy overrides), or
   
      "mutual_exclude" (ignore both).
   
    The resulting vector is returned as `pos[:, 0]` (shape `(T,)`).

    Notes
    -----
    - All crosses are evaluated with prerolled prior values to avoid double-
      counting and to handle NaN warm-ups gracefully.
   
    - Gating reduces false positives by restricting signals to market regimes
      for which the indicator is designed (pro-trend or anti-trend) and by
      requiring adequate trading activity where applicable.
    """
    
    is_pro = name in PRO_TREND
    
    is_anti = name in ANTI_TREND
    
    vol_req = name in VOL_REQUIRED

    T = pc.c.shape[0]
    
    gate = _build_gate(
        pc = pc, 
        is_pro_trend = is_pro, 
        is_anti_trend = is_anti, 
        vol_required = vol_req
    )

    if name == "bb":
       
        w = int(params.get("window", BB_WINDOW))
       
        ns = float(params.get("num_std", BB_STD))
       
        mid, std = _bb_cached(
            pc = pc,
            window = w
        )
       
        upper = mid + std * ns
       
        lower = mid - std * ns

        c2 = pc.c[:, None]
       
        c_prev = np.roll(c2, 1, axis = 0)
        
        c_prev[0, :] = np.nan
       
        buy_mat  = (c2 > lower[:, None]) & (c_prev <= lower[:, None])
       
        sell_mat = (c2 < upper[:, None]) & (c_prev >= upper[:, None])

    elif name == "atr":
       
        m = float(params.get("mult", ATR_MULTIPLIER))
       
        aw = int(params.get("atr_window", ATR_WINDOW))
       
        bw = int(params.get("break_window", ATR_BREAK_WINDOW))

        atr = _atr_cached(
            pc = pc, 
            window = aw
        )
       
        atr_prev = np.roll(atr, 1)
        
        atr_prev[0] = np.nan
       
        hi_prev, lo_prev = _break_cached(
            pc = pc, 
            window = bw
        )

        up_th = hi_prev + atr_prev * m
        
        dn_th = lo_prev - atr_prev * m
        
        c2 = pc.c[:, None]
        
        buy_mat = (c2 > up_th[:, None])
        
        sell_mat = (c2 < dn_th[:, None])

    elif name == "mfi":
        
        bt = float(params.get("buy_thresh", MFI_BUY_THRESH))
        
        st = float(params.get("sell_thresh", MFI_SELL_THRESH))
        
        w  = int(params.get("window", MFI_WINDOW))
        
        mfi = _mfi_cached(
            pc = pc,
            window = w
        )
        
        mfi_prev = np.roll(mfi, 1)
        
        mfi_prev[0] = np.nan

        buy_mat = ((mfi[:, None] > bt) & (mfi_prev[:, None] <= bt))
        
        sell_mat = ((mfi[:, None] < st) & (mfi_prev[:, None] >= st))

    elif name == "stoch":
       
        lo = float(params.get("lower", 20.0))
       
        hi = float(params.get("upper", 80.0))
       
        k = pc.stoch_k[:, None]
       
        d = pc.stoch_d[:, None]

        k_prev = np.roll(k, 1, axis = 0)
        
        k_prev[0, :] = np.nan
       
        d_prev = np.roll(d, 1, axis = 0)
        
        d_prev[0, :] = np.nan
       
        cross_up = (k > d) & (k_prev <= d_prev)
       
        cross_down = (k < d) & (k_prev >= d_prev)

        buy_mat  = (k < lo) & cross_up
        
        sell_mat = (k > hi) & cross_down

    elif name == "rsi":
       
        bt = float(params.get("buy_thresh", RSI_BUY_THRESH))
       
        st = float(params.get("sell_thresh", RSI_SELL_THRESH))
       
        w = int(params.get("window", RSI_WINDOW))
       
        rsi = _rsi_cached(
            pc = pc,
            window = w
        )
       
        rsi_prev = np.roll(rsi, 1)
        
        rsi_prev[0] = np.nan

        buy_mat = ((rsi[:, None] > bt) & (rsi_prev[:, None] <= bt))
      
        sell_mat = ((rsi[:, None] < st) & (rsi_prev[:, None] >= st))

    elif name == "ema":
        
        f = int(params.get("fast", EMA_FAST))
        
        s = int(params.get("slow", EMA_SLOW))
        
        ef = _ema_cached(
            pc = pc,
            span = f
        )[:, None]
        
        es = _ema_cached(
            pc = pc, 
            span = s
        )[:, None]

        ef_prev = np.roll(ef, 1, axis = 0); ef_prev[0, :] = np.nan
        
        es_prev = np.roll(es, 1, axis = 0); es_prev[0, :] = np.nan
        
        buy_mat  = (ef > es) & (ef_prev <= es_prev)
        
        sell_mat = (ef < es) & (ef_prev >= es_prev)

    elif name == "macd":
        
        f = int(params.get("fast", EMA_FAST))
        
        s = int(params.get("slow", EMA_SLOW))
        
        g = int(params.get("signal", 9))
        
        line, sig = _macd_signal_cached(
            pc = pc,
            fast = f, 
            slow = s, 
            signal = g
        )
        
        ln = line[:, None]
        
        sg = sig[:, None]
        
        ln_prev = np.roll(ln, 1, axis = 0); ln_prev[0, :] = np.nan
        
        sg_prev = np.roll(sg, 1, axis = 0); sg_prev[0, :] = np.nan
        
        buy_mat = (ln > sg) & (ln_prev <= sg_prev)
        
        sell_mat = (ln < sg) & (ln_prev >= sg_prev)

    elif name == "vwap":
        
        w = int(params.get("window", VWAP_WINDOW))
        
        num = _rolling_sum_from_cumsum(
            cs = pc._cs_pv, 
            w = w
        )
        
        den = _rolling_sum_from_cumsum(
            cs = pc._cs_v,  
            w = w
        )
        
        rvwap = (num / np.where(den == 0.0, np.nan, den))[:, None]
        
        c2 = pc.c[:, None]
        
        c_prev = np.roll(c2, 1, axis = 0)
        
        c_prev[0, :] = np.nan
        
        buy_mat  = (c2 > rvwap) & (c_prev <= rvwap)
        
        sell_mat = (c2 < rvwap) & (c_prev >= rvwap)

    elif name == "obv":
        
        lb = int(params.get("lookback", OBV_LOOKBACK))
        
        price_low_prev, price_high_prev = _price_extrema_cached(
            pc = pc,
            window = lb
        )
        
        obv_low_prev, obv_high_prev = _obv_extrema_cached(
            pc = pc,
            window = lb
        )
        
        c2  = pc.c[:, None]
        
        obv2 = pc.obv[:, None]
        
        buy_mat = (c2 < price_low_prev) & (obv2 >= obv_low_prev)
        
        sell_mat = (c2 > price_high_prev) & (obv2 <= obv_high_prev)

    else:

        return np.zeros(T, dtype = np.int8)

    if not gate.all():
       
        buy_mat &= gate[:, None]
       
        sell_mat &= gate[:, None]

    buy_mat = np.ascontiguousarray(buy_mat)
    
    sell_mat = np.ascontiguousarray(sell_mat)    
        
    pos = pulses_to_positions(
        buys = buy_mat, 
        sells = sell_mat, 
        signal_lag = SIGNAL_LAG, 
        cooldown_bars = COOLDOWN_BARS, 
        mode_id = _mode_id(
            mode = CONFLICT_MODE
        )
    )
   
    return pos[:, 0] 


def score_for_ticker(
    pc: Precomp,
    tk: str, 
    calib_raw: dict[str, dict]
) -> pd.Series:
    """
    Compute the composite integer signal score for a single ticker by
    aggregating weighted indicator positions.

    Parameters
    ----------
    pc : Precomp
        Precomputed arrays for the ticker (from `precompute_all`).
    tk : str
        Ticker symbol; used as the name of the returned series.
    calib_raw : dict[str, dict]
        Calibration map for this ticker:
        `calib_raw[name] = {"params": dict, "weight": int, "sharpe": float,
        "t_stat": float, "p_value": float, "n_trades": int}`.
        Missing entries are filled from `default_params_for(name)` and
        `DEFAULT_WEIGHTS`.

    Returns
    -------
    pd.Series (dtype int16)
        Time-indexed series `S_t` with the same index as the price, where
        
            `S_t = clip( Σ_i w_i * pos_i(t), −SCORE_CLAMP, +SCORE_CLAMP )`,
            
        `w_i ∈ {0,1,2,3,4}` are indicator weights obtained during calibration,
        and `pos_i(t) ∈ {−1,0,+1}` are per-indicator positions constructed by
        `_pos_for_indicator` for the calibrated parameters.

    Construction
    ------------
    For each indicator in `INDICATOR_ORDER`:
     
      1) obtain `(w_i, params_i)` from `calib_raw` (fall back to defaults);
     
      2) if `w_i > 0`, compute `pos_i(t)` via `_pos_for_indicator`, which
         performs gating by regime and volume, and applies the global lag,
         cooldown, and conflict policy during pulse-to-position mapping;
     
      3) accumulate `w_i * pos_i(t)` into an `int16` running total.

    Interpretation
    --------------
    The score is a dimensionless, direction-only aggregation of indicator
    agreement:
     
      - `S_t > 0` indicates a majority of *weighted* pro-long signals;
     
      - `S_t < 0` indicates dominance of *weighted* pro-short signals;
     
      - magnitude reflects consensus strength up to `SCORE_CLAMP`.

    Statistical Context
    -------------------
    Indicator weights arise from a significance gate:
   
      - annualised Sharpe threshold `Sharpe >= MIN_SHARPE_FOR_SIG`,
   
      - one-sided p-value `p_value < SIGNIFICANCE_ALPHA` for `H1: mu > 0`,
   
      - positive improvement over buy-and-hold Sharpe.
   
    The weights represent relative outperformance in-sample, not certainty of
    future returns. Positions are aligned to avoid look-ahead:
   
        `X_t = pos_{t−1} * r_t`.

    Notes
    -----
    - The composite score is not a forecast of expected return and is not
      directly comparable across tickers without calibration context.
  
    - The return value is clipped to `±SCORE_CLAMP` to reduce the impact of
      extreme indicator alignment.
    """

    calib: dict[str, dict] = {}

    for name in INDICATOR_ORDER:

        if name in calib_raw:

            calib[name] = calib_raw[name]

        else:

            calib[name] = {
                "params": default_params_for(name),
                "weight": DEFAULT_WEIGHTS.get(name, 1),
                "sharpe": 0.0,
                "t_stat": 0.0,
                "p_value": 1.0, 
                "n_trades": 0
            }

    score = np.zeros(pc.c.shape[0], dtype=np.int16)
   
    for name in INDICATOR_ORDER:
   
        w = int(calib[name].get("weight", 0))
   
        if w <= 0:
   
            continue
   
        params = calib[name].get("params", {})
   
        pos = _pos_for_indicator(
            name = name, 
            params = params, 
            pc = pc
        ).astype(np.int16)
   
        score = score + (w * pos)

    score = np.clip(score, -SCORE_CLAMP, SCORE_CLAMP)
  
    return pd.Series(score, index = pc.idx, dtype = "int16", name = tk)


def main() -> None:
    """
    Orchestrate end-to-end loading, calibration (optional), scoring, and
    persistence of results to the Excel workbook.

    Pipeline
    --------
    1) Load OHLCV data
     
       `data = load_data(config.DATA_FILE)` expects sheets "High", "Low",
     
       "Close", and "Volume" with a DatetimeIndex.

    2) Determine calibration mode
     
       If `UPDATE_PARAM` is True, run per-ticker calibration across all
       indicators; otherwise, load the last saved calibration via
       `load_calibration_map()` and use those parameters and weights.

    3) Per-ticker processing
     
       For each `tk` in `config.tickers`:
     
         a) extract column series for OHLCV (`_col` enforces presence);
     
         b) build `Precomp` via `precompute_all(h, l, c, v)`;
     
         c) if calibrating:
     
              - call `calibrate_ticker` to tune indicators via grid search,
                compute Sharpe, t-statistic, and p-value using the backtest
                alignment `X_t = pos_{t−1} * r_t`, count flips, and derive
                integer weights based on improvement over buy-and-hold Sharpe;
     
              - accumulate rows for upsert into the "Calibration" sheet;
     
            else:
     
              - use cached parameters/weights; if absent, fall back to
                on-the-fly calibration (safe default);
     
         d) compute the composite integer score series with
            `score_for_ticker(pc, tk, calib_raw)`. This series sums
            `weight * position` across indicators and is clamped to
            `±SCORE_CLAMP`.

    4) Persist calibration (if updated)
     
       If `UPDATE_PARAM` is True, upsert accumulated calibration rows into
       the "Calibration" sheet using `write_calibration_rows`.

    5) Persist composite scores
     
       Concatenate all ticker series into a DataFrame, restrict the index to
       dates later than `config.YEAR_AGO`, and write to the "Signal Scores"
       sheet via `save_scores`. The writer applies Excel conditional
       formatting: values > 0 shaded green; values < 0 shaded red.

    Statistical and Modelling Details
    ---------------------------------
    
    - Returns are simple arithmetic returns `r_t = (C_t / C_{t−1}) − 1`.
    
    - Realised strategy returns are aligned as `X_t = pos_{t−1} * r_t`.
    
    - Annualised Sharpe is `sqrt(252) * (mean(X) / sd(X))`. The t-statistic is
      `mean(X) / (sd(X) / sqrt(n))`, with one-sided p-values for `H1: mean > 0`.
    
    - The calibration significance gate requires
      `p_value < SIGNIFICANCE_ALPHA` and `Sharpe >= MIN_SHARPE_FOR_SIG`.
    
    - Weights in {0,1,2,3,4} reflect Sharpe improvement over buy-and-hold.

    Side Effects
    ------------
    - Logs progress and any per-ticker exceptions. Tickers failing validation
      or computation are skipped and reported.
    
    - Overwrites or creates the "Signal Scores" sheet and upserts the
      "Calibration" sheet in the workbook at `config.DATA_FILE`.

    Notes
    -----
    - The procedure is in-sample and does not model execution costs or
      constraints. The composite score is intended as a regime/consensus
      indicator rather than a standalone trading signal.
    """
    
    data = load_data(
        excel_file = config.DATA_FILE
    )
    
    tickers = list(config.tickers)

    logger.info("Scoring %d tickers", len(tickers))

    series_list: List[pd.Series] = []
   
    skipped: List[str] = []
   
    all_rows: List[dict] = []

    cache_map: dict[str, dict[str, dict]] | None = None
   
    if not UPDATE_PARAM:
   
        cache_map = load_calibration_map()

    for tk in tickers:
   
        try:
            c = _col(
                df = data["close"], 
                tk = tk, 
                name = "Close"
            )
            
            h = _col(
                df = data["high"], 
                tk = tk,
                name = "High"
            )
            
            l  =  _col(
                df = data["low"], 
                tk = tk, 
                name = "Low"
            )
            
            v = _col(
                df = data["volume"],
                tk = tk,
                name = "Volume"
            )
            
            pc = precompute_all(
                h = h,
                l = l,
                c = c,
                v = v
            )

            if UPDATE_PARAM:
             
                calib_raw, rows = calibrate_ticker(
                    pc = pc, 
                    c = c,
                    tk = tk
                )
             
                all_rows.extend(rows)
         
            else:
         
                cached = cache_map.get(tk) if cache_map is not None else None
         
                if cached is None:
         
                    calib_raw, rows = calibrate_ticker(
                        pc = pc,
                        c = c, 
                        tk = tk
                    )
         
                else:
         
                    calib_raw = cached

            s = score_for_ticker(
                pc = pc, 
                tk = tk,
                calib_raw = calib_raw
            )
            
            if s.empty:
            
                raise ValueError(f"Empty score series for {tk}")
            
            series_list.append(s)

        except Exception as e:
            
            logger.exception("Skipping %s due to error: %s", tk, e)
            
            skipped.append(tk)

    if UPDATE_PARAM and all_rows:
        
        logger.info("Writing %d calibration rows ...", len(all_rows))
        
        write_calibration_rows(
            rows = all_rows
        )

    if not series_list:
      
        raise RuntimeError("No valid score series produced; cannot build score DataFrame.")

    score_df = pd.concat(series_list, axis = 1).sort_index()
   
    score_df = score_df.loc[score_df.index > pd.to_datetime(config.YEAR_AGO)]

    logger.info("Writing results back to workbook (skipped %d tickers)", len(skipped))
   
    save_scores(
        excel_file = config.DATA_FILE,
        scores = score_df
    )
   
    logger.info("Done – sheet 'Signal Scores' updated.")


if __name__ == "__main__":
   
    main()
