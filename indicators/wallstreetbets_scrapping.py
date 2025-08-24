"""
Scrapes posts and comments from r/wallstreetbets, extracts ticker mentions, analyses sentiment using NLTK’s VADER, 
and saves aggregated results with conditional formatting in Excel.
"""

import re
import time
import logging
import requests
import pandas as pd
from collections import Counter
from typing import List, Dict, Any

from nltk.sentiment import SentimentIntensityAnalyzer
import nltk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import config


nltk.download('vader_lexicon', quiet = True)

logger = logging.getLogger(__name__)

logger.setLevel(logging.INFO)

console_handler = logging.StreamHandler()

console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

console_handler.setFormatter(console_formatter)

logger.addHandler(console_handler)


STOPWORDS = {
    "the",
    "and", 
    "to", 
    "a", 
    "in", 
    "of", 
    "for", 
    "is", 
    "on",
    "that",
    "with", 
    "as",
    "at", 
    "this", 
    "it", 
    "by", 
    "from", 
    "are", 
    "an",
    "i",
    "you",
    "but", 
    "if",
    "or",
    "be", 
    "was",
    "so", 
    "we", 
    "they", 
    "not",
    "have", 
    "has", 
    "im", 
    "can",
    "there",
    "what", 
    "will", 
    "all",
    "just",
    "https",
    "s", 
    "trump",
    "ukraine",
    "minerals",
    "there",
    "why", 
    "wondering",
    "apparently", 
    "everyone", 
    "there", 
    "says", 
    "their",
    "about",
    "reddit", 
    "r",
    "wallstreetbets", 
    "amp", 
    "http", 
    "www", 
    "com",
    "figured", 
    "last", 
    "coffee",
    "people",
    "college", 
    "university", 
    "student", 
    "obvious",
    "me", 
    "bpunt", 
    "slurping", 
    "see", 
    "guys", 
    "should", 
    "do", 
    "people", 
    "per", 
    "usually"
}


NOTTICKERS = {
    "CEO",
    "CFO", 
    "CTO", 
    "COO", 
    "WSB", 
    "DD", 
    "YOLO",
    "TOS", 
    "AOC", 
    "GDP",
    "OTM", 
    "GAIN", 
    "IS", 
    "UK", 
    "THE",
    "US",
    "NEVER",
    "OK",
    "FDA", 
    "AM", 
    "PM",
    "RH", 
    "EV", 
    "IPO",
    "ATH", 
    "TOS", 
    "TD", 
    "EDIT",
    "TLDR", 
    "ROPE",
    "STAY",
    "SAFE",
    "AUTO", 
    "BOT",
    "AI", 
    "IT", 
    "ELON",
    "MUSK",
    "SEC",
    "TICK",
    "TOCK",
    "USD", 
    "CPU",
    "IS",
    "DOE",
    "OP", 
    "NY", 
    "DOJ",
    "IRA",
    "NOT",
    "ZERO", 
    "II", 
    "III",
    "IV",
    "NFA", 
    "IN",
    "BUY",
    "BUT",
    "JOBS",
    "THEIR", 
    "WAS", 
    "GOOD",
    "EU",
    "DVD",
    "EPS",
    "IM", 
    "RIF"
}


def get_wsb_posts(
    limit: int = 100
) -> List[Dict[str, Any]]:
    """
    Fetch top posts from r/wallstreetbets via Reddit's public JSON endpoint.

    Parameters
    ----------
    limit : int, default 100
        Maximum number of posts to request (passed as the `?limit=` query parameter).

    Returns
    -------
    list[dict[str, Any]]
        A list of post "children" objects from the Reddit listing JSON
        (each item typically exposes `.get("data", {})` with post fields).
        Returns an empty list if an HTTP or parsing error occurs.

    Network / Robustness
    --------------------
    - Uses a desktop-style `User-Agent`.
    - Wraps `requests.get(...).raise_for_status()` in try/except and logs on failure.
    - Endpoint: ``https://www.reddit.com/r/wallstreetbets/.json?limit={limit}``

    Notes
    -----
    This unauthenticated endpoint is rate-limited and can throttle or block
    excessive requests. If your volume grows, consider authenticated API access
    via PRAW or Reddit's official OAuth2.
    """


    url = f"https://www.reddit.com/r/wallstreetbets/.json?limit={limit}"

    headers = {
        'User-Agent': 'Mozilla/5.0 (compatible; TickerScraper/1.0)'
    }

    try:

        response = requests.get(url, headers=headers, timeout = 10)

        response.raise_for_status()

    except requests.RequestException as exc:

        logger.error("Error fetching posts from Reddit: %s", exc)

        return []

    data = response.json()

    return data.get("data", {}).get("children", [])


def get_comments_for_post(
    post_id: str
) -> List[str]:
    """
    Fetch all (top-level + nested) comments for a given r/wallstreetbets post ID
    and return the plain-text bodies.

    Parameters
    ----------
    post_id : str
        Reddit post ID (e.g., 'abc123').

    Returns
    -------
    list[str]
        A flat list of comment bodies. Returns [] if the request fails,
        if the response does not include a comments listing, or if there
        are simply no comments.

    Network / Backoff
    -----------------
    - Endpoint: ``https://www.reddit.com/r/wallstreetbets/comments/{post_id}/.json``
    - Retries up to `max_retries = 2` times with simple exponential backoff
    (sleep 5s, then 10s) when a 429 (rate limit) is detected.
    - Other HTTP exceptions are logged and cause an early return [].

    Implementation details
    ----------------------
    The Reddit comments JSON is a list of two listings:
    1) the post
    2) the comments tree

    We extract the second element, traverse `data.children` recursively, and collect
    comment bodies where `kind == "t1"`.

    Pseudocode of recursion:
        - If node.kind != 't1' → skip
        - Append `node.data.body` if present
        - If `node.data.replies` is a dict, recurse into `replies.data.children`

    Caveats
    -------
    - Deleted/removed comments may have empty or missing `body`.
    - Deeply nested threads are handled, but the total comment count is subject
    to Reddit's collapsed/continued threads behavior in the JSON view.
    """


    url = f"https://www.reddit.com/r/wallstreetbets/comments/{post_id}/.json"

    headers = {'User-Agent': 'Mozilla/5.0 (compatible; TickerScraper/1.0)'}

    max_retries = 2

    delay = 5

    for attempt in range(max_retries):

        try:

            response = requests.get(url, headers=headers, timeout=10)

            response.raise_for_status()

            break
       
        except requests.RequestException as exc:

            if "429" in str(exc):

                logger.warning(
                    "Rate limited fetching comments for post %s (attempt %d/%d). Sleeping %d seconds.",
                    post_id, attempt + 1, max_retries, delay
                )

                time.sleep(delay)

                delay *= 2

            else:

                logger.error("Error fetching comments for post %s: %s", post_id, exc)

                return []
    else:

        logger.error("Failed to fetch comments for post %s after %d retries.", post_id, max_retries)

        return []

    data = response.json()

    if len(data) < 2:

        return []

    comment_data = data[1].get("data", {}).get("children", [])

    comments = []

    def extract_comments(
        comments_list: List[Dict[str, Any]]
    ) -> None:
        """
        Recursively traverse comment threads and collect comment bodies.
        """

        for item in comments_list:

            if item.get("kind") != "t1":

                continue

            comment_body = item.get("data", {}).get("body", "")

            if comment_body:

                comments.append(comment_body)

            replies = item.get("data", {}).get("replies")

            if replies and isinstance(replies, dict):

                next_children = replies.get("data", {}).get("children", [])

                extract_comments(
                    comments_list = next_children
                )

    extract_comments(
        comments_list = comment_data
    )

    return comments


def extract_tickers(
    text: str
) -> List[str]:
    """
        Extract likely stock ticker tokens from free text using a simple regex heuristic.

        Parameters
        ----------
        text : str
            Arbitrary text (title, selftext, or comment body).

        Returns
        -------
        list[str]
            Uppercase tokens that look like tickers (2–5 capital letters), filtered
            to exclude common all-caps words/acronyms listed in `NOTTICKERS`.

        Method
        ------
        - Regex pattern: ``r'\\b[A-Z]{2,5}\\b'``
        * ``\\b`` are word boundaries to avoid partial matches.
        * ``[A-Z]{2,5}`` restricts to 2–5 uppercase letters (e.g., 'AAPL', 'TSLA').
        - Post-filter to drop false positives using the curated `NOTTICKERS` set.

        Limitations
        -----------
        - Symbols with dots/hyphens (e.g., 'BRK.B', 'RDS-A', 'SHOP.TO') are **not** captured.
        - Multi-word tickers or casings with lowercase letters are excluded.
        - This is intentionally conservative; expand the regex and false-positive filters
        if you target international symbols or more exotic tickers.
        """


    pattern = r'\b[A-Z]{2,5}\b'

    tickers = re.findall(pattern, text)

    return [ticker for ticker in tickers if ticker not in NOTTICKERS]


def extract_words(
    text: str
) -> List[str]:
    """
    Tokenise text into lowercased words and remove common stopwords.

    Parameters
    ----------
    text : str
        Input text.

    Returns
    -------
    list[str]
        A list of lowercase tokens where each token matches ``\\b\\w+\\b`` and is
        not contained in the `STOPWORDS` set.

    Notes
    -----
    - This is a lightweight tokeniser for bag-of-words features (e.g., top-3
    words per ticker). It intentionally avoids stemming/lemmatisation to keep
    the vocabulary interpretable.
    - Emojis, punctuation-only tokens, and URLs are effectively discarded by the regex.
    """

    words = re.findall(r'\b\w+\b', text.lower())

    return [word for word in words if word not in STOPWORDS]


def format_sheet_as_table(
    excel_file: str, 
    sheet_name: str
) -> None:
    """
    Format an existing Excel worksheet as an OpenXML table for readability and filtering.

    Parameters
    ----------
    excel_file : str
        Path to the .xlsx file.
    sheet_name : str
        Worksheet name to format (must already exist).

    Effects
    -------
    - Wraps the used range A1:({last_col}{last_row}) in an Excel table.
    - Applies `TableStyleMedium9` with row striping.
    - Saves the workbook in-place.

    Returns
    -------
    None

    Robustness
    ----------
    - If the sheet does not exist, logs a warning and returns.
    - Any exception during workbook operations is caught and logged.

    Notes
    -----
    OpenPyXL tables add native Excel features: filtering, banded rows, and
    styling. This function is typically called after writing a DataFrame to a sheet.
    """

    try:

        wb = load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:

            logger.warning("Sheet '%s' not found in %s", sheet_name, excel_file)

            return

        ws = wb[sheet_name]

        max_row = ws.max_row

        max_col = ws.max_column

        last_col_letter = get_column_letter(max_col)

        table_range = f"A1:{last_col_letter}{max_row}"

        table_name = sheet_name.replace(" ", "") + "Table"

        table = Table(displayName=table_name, ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        ws.add_table(table)

        wb.save(excel_file)

        logger.info("Sheet '%s' formatted as a table in %s.", sheet_name, excel_file)

    except Exception as exc:

        logger.error("Error formatting sheet '%s' as table: %s", sheet_name, exc)


def main() -> None:
    """
    End-to-end Reddit WSB sentiment and ticker-mention pipeline.

    Steps
    -----
    1) **Fetch posts**:
    - Pull up to 100 latest items from r/wallstreetbets (`get_wsb_posts`).

    2) **Initialise sentiment**:
    - Create an NLTK VADER `SentimentIntensityAnalyzer`.
    - Extend its lexicon with domain-specific weights (`custom_words`).
        VADER computes:
        - 'compound' score ∈ [-1, 1], a normalised sum of valence signals:
            Let s = Σ (token valences with heuristics).
            Then, per VADER, the normalisation is:
                compound = s / √(s² + α),  with α = 15,
            producing an S-shaped squashing.

    3) **Process each post**:
    - Extract tickers from post title (`extract_tickers`).
    - Compute post sentiment on the concatenated "title + selftext".
    - Tokenise “words” (lowercased, minus STOPWORDS) via `extract_words`.
    - Accumulate per-ticker:
        * Mentions count
        * Bag-of-words counts (Counter)
        * A list of 'compound' sentiment scores

    4) **Fetch and process comments**:
    - For each post, get all nested comments via `get_comments_for_post`.
    - For each comment containing tickers:
        * Compute VADER compound sentiment on the comment text
        * Update per-ticker mentions, word Counter, and sentiment list

    5) **Aggregate statistics**:
    - For each ticker with sentiment list {s_i}_{i=1..n}:
        * Average sentiment (population mean):
                μ = (1/n) Σ s_i
        * Dispersion (population standard deviation):
                σ = √( (1/n) Σ (s_i − μ)² )
            (When n = 1, set σ = 0.)
    - Top words: take the three most common tokens from the ticker Counter.

    6) **Sort & export**:
    - Sort tickers by total mentions (descending).
    - Build `findings_df` with columns:
        ['ticker', 'mentions', 'avg_sentiment', 'sentiment_std', 'top_words'].
    - Write to the `Sentiment Findings` sheet in `config.FORECAST_FILE`
        (append/replace mode).
    - Apply conditional formatting on the 'avg_sentiment' column:
        * red fill if < 0
        * green fill if > 0
    - Finally, call `format_sheet_as_table(...)` to convert the sheet to an Excel table.

    Performance / Rate Limits
    -------------------------
    - Inserts `time.sleep(1)` per post to be friendlier to Reddit's rate limits.
    - `get_comments_for_post` adds simple exponential backoff for 429s.

    Caveats
    -------
    - Regex-based ticker extraction is heuristic and conservative; adapt it for
    international or dotted tickers.
    - VADER is lexicon- and rule-based; sarcasm, memes, and non-standard slang
    can yield noisy scores. The custom lexicon here reduces (but does not
    eliminate) such effects.

    Outputs
    -------
    - Excel file at `config.FORECAST_FILE` containing a sheet named
    'Sentiment Findings' with mentions and sentiment summaries by ticker.
    """


    logger.info("Scraping WallStreetBets posts...")

    posts = get_wsb_posts(
        limit = 100
    )

    if not posts:

        logger.warning("No posts fetched. Exiting.")

        return

    sia = SentimentIntensityAnalyzer()

    custom_words = {
        "buy": 4.0, 
        "sell": -4.0, 
        "bull": 4, 
        "bullish": 4, 
        "bear": -4,
        "bearish": -4, 
        "moon": 4, 
        "rocket": 4, 
        "crash": -4, 
        "yolo": 2,
        "tendies": 2, 
        "stonks": 2, 
        "omg": 1, 
        "fomo": 1,
        "fml": -1,
        "up": 4, 
        "down": -4, 
        "long": 4, 
        "short": -4, 
        "overvalued": -3,
        "undervalued": 3, 
        "pump": 4, 
        "dump": -4,
        "bagholder": -2,
        "moonshot": 2, 
        "highs": 2, 
        "lows": -2, 
        "underrated": 3,
        "overrated": -3, 
        "pumping": 2,
        "dumping": -3,
        "recession": -2,
        "depression": -2, 
        "crisis": -2, 
        "soaring": 3,
        "soars": 2,
        "soared": 2, 
        "plunges": -3,
        "plunged": -2,
        "plunge": -3,
        "surge": 4, 
        "surges": 4, 
        "surged": 2, 
        "collapses": -4,
        "collapsed": -2, 
        "collapse": -4, 
        "rally": 2,
        "rallies": 2,
        "rallied": 2, 
        "crashes": -4,
        "crashed": -2,
        "buy the dip": 4,
        "put": -4, 
        "call": 4, 
        "buying": 4, 
        "selling": -4,
        "cheap": 4,
        "expensive": -4, 
        "late": -4
    }

    sia.lexicon.update(custom_words)

    ticker_word_counter: Dict[str, Counter] = {}

    ticker_mentions: Dict[str, int] = {}

    ticker_sentiments: Dict[str, List[float]] = {}

    logger.info("Processing %d posts (including comments)...", len(posts))

    for i, post in enumerate(posts):


        logger.info("Processing Post %d/%d", i + 1, len(posts))

        post_data = post.get("data", {})

        post_id = post_data.get("id", "")

        title = post_data.get("title", "")

        selftext = post_data.get("selftext", "")

        combined_post_text = f"{title} {selftext}"

        post_tickers = extract_tickers(
            text = title
        )

        if post_tickers:

            sentiment_scores = sia.polarity_scores(combined_post_text)

            words_in_post = extract_words(
                text = combined_post_text
            )

            for tkr in post_tickers:

                ticker_mentions[tkr] = ticker_mentions.get(tkr, 0) + 1

                ticker_word_counter.setdefault(tkr, Counter()).update(words_in_post)

                ticker_sentiments.setdefault(tkr, []).append(sentiment_scores['compound'])

        comments = get_comments_for_post(
            post_id = post_id
        )

        for comment_text in comments:

            comment_tickers = extract_tickers(
                text = comment_text
            )

            if not comment_tickers:

                continue

            comment_sentiment = sia.polarity_scores(comment_text)['compound']

            comment_words = extract_words(
                text = comment_text
            )

            for tkr in comment_tickers:

                ticker_mentions[tkr] = ticker_mentions.get(tkr, 0) + 1

                ticker_word_counter.setdefault(tkr, Counter()).update(comment_words)

                ticker_sentiments.setdefault(tkr, []).append(comment_sentiment)

        time.sleep(1)

    ticker_avg_sentiment = {}

    ticker_sentiment_std = {}

    for tkr, sentiment_values in ticker_sentiments.items():

        avg_score = sum(sentiment_values) / len(sentiment_values)

        ticker_avg_sentiment[tkr] = avg_score

        if len(sentiment_values) > 1:

            variance = sum((val - avg_score) ** 2 for val in sentiment_values) / len(sentiment_values)

            ticker_sentiment_std[tkr] = variance ** 0.5

        else:

            ticker_sentiment_std[tkr] = 0.0

    ticker_top_words = {
        tkr: ", ".join([word for word, _ in ticker_word_counter[tkr].most_common(3)])
        for tkr in ticker_word_counter
    }

    sorted_tickers = sorted(ticker_mentions, key=ticker_mentions.get, reverse=True)

    findings = []

    for tkr in sorted_tickers:

        findings.append({
            "ticker": tkr,
            "mentions": ticker_mentions[tkr],
            "avg_sentiment": ticker_avg_sentiment.get(tkr, 0),
            "sentiment_std": ticker_sentiment_std.get(tkr, 0),
            "top_words": ticker_top_words.get(tkr, "")
        })

    findings_df = pd.DataFrame(findings)

    try:

        with pd.ExcelWriter(config.FORECAST_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:

            findings_df.to_excel(writer, sheet_name='Sentiment Findings', index=False)

            ws = writer.sheets['Sentiment Findings']

            try:

                col_idx = findings_df.columns.get_loc("avg_sentiment") + 1  

                col_letter = get_column_letter(col_idx)

            except Exception as exc:

                logger.error("Error finding the 'avg_sentiment' column index: %s", exc)

                col_letter = "C"  

            max_row = ws.max_row

            red_fill = PatternFill(start_color = 'FFC7CE', end_color = 'FFC7CE', fill_type = 'solid')

            green_fill = PatternFill(start_color = 'C6EFCE', end_color = 'C6EFCE', fill_type = 'solid')

            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{max_row}",
                CellIsRule(operator = 'lessThan', formula = ['0'], fill = red_fill)
            )
         
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{max_row}",
                CellIsRule(operator = 'greaterThan', formula = ['0'], fill = green_fill)
            )

        logger.info("Sentiment findings successfully saved to Excel in '%s'.", config.FORECAST_FILE)

    except Exception as exc:
      
        logger.error("Failed to write sentiment findings to Excel: %s", exc)

    format_sheet_as_table(
        excel_file = config.FORECAST_FILE, 
        sheet_name = 'Sentiment Findings'
    )


if __name__ == "__main__":
    
    main()

