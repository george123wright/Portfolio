import numpy as np
import yfinance as yf
import pandas as pd
import datetime as dt
import scipy.stats as st
import logging
from typing import Tuple, Any, Dict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from maps.industry_mapping import IndustryMap
from maps.sector_map import SectorMap
pd.set_option('future.no_silent_downcasting', True)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
ch = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
ch.setFormatter(formatter)
logger.addHandler(ch)

today: dt.date = dt.date.today()

excel_file: str = f"Portfolio_Optimisation_Data_{today}.xlsx"

output_excel_file: str = f"Portfolio_Optimisation_Forecast_{today}.xlsx"

currency = pd.read_excel(
    f'/Users/georgewright/Portfolio_Optimisation_Data_{today}.xlsx',
    sheet_name='Currency',
    index_col=0,
    parse_dates=True,
    engine='openpyxl'
)

usdcny = currency.loc['USDCNY']['Last']

usdcad = currency.loc['USDCAD']['Last']

usdcny = pd.to_numeric(usdcny, errors='coerce')

usdcad = pd.to_numeric(usdcad, errors='coerce')

gbpusd = pd.to_numeric(currency.loc['GBPUSD']['Last'], errors='coerce')

eurusd = pd.to_numeric(currency.loc['EURUSD']['Last'], errors='coerce')

usdchf = pd.to_numeric(currency.loc['USDCHF']['Last'], errors='coerce')

usdhkd = pd.to_numeric(currency.loc['USDHKD']['Last'], errors='coerce')

usdcny = pd.to_numeric(currency.loc['USDCNY']['Last'], errors='coerce')

usddkk = pd.to_numeric(currency.loc['USDDKK']['Last'], errors='coerce')

logger.info("Downloading Analyst Data from Yahoo Finance ...")

try:

    close = (
        pd.read_excel(excel_file, sheet_name="Close", index_col=0, parse_dates=True, engine="openpyxl")
        .sort_index(ascending=True)
    )

    close.columns = close.columns.astype(str)

    rets = (
        pd.read_excel(excel_file, sheet_name="Historic Returns", index_col=0, parse_dates=True, engine="openpyxl")
        .sort_index(ascending=True)
    )

    rets.columns = rets.columns.astype(str)

except Exception as e:

    logger.error(
        "Failed to read data from Excel. Ensure the file '%s' is not open in another application. Error: %s",
        excel_file, e
    )

    raise


def drawdown(return_series: pd.Series) -> pd.DataFrame:
    """
    Calculate wealth index, previous peaks, and drawdowns.
    """

    wealth_index = 1000 * (1 + return_series).cumprod()

    previous_peaks = wealth_index.cummax()

    drawdowns = (wealth_index - previous_peaks) / previous_peaks

    return pd.DataFrame({"Wealth": wealth_index, "Previous Peak": previous_peaks, "Drawdown": drawdowns})


def safe_first_value(df: pd.DataFrame, row_label: str, default: float = np.nan) -> float:
    """
    Safely retrieve the first value for a given row label from a DataFrame.
    Logs a warning if the row or index does not exist.
    """
  
    try:
  
        return df.loc[row_label].iloc[0]
  
    except Exception as e:
  
        logger.warning("Missing or invalid data for '%s': %s", row_label, e)
  
        return default


tickers: list = close.columns.tolist()

ticker_objs: Dict[str, yf.Ticker] = {}

targetsYF: Dict[str, Any] = {}

for ticker in tickers:

    try:

        ticker_obj = yf.Ticker(ticker)

        ticker_objs[ticker] = ticker_obj

        targetsYF[ticker] = ticker_obj.info

    except Exception as e:

        logger.warning("Failed to retrieve info for %s: %s", ticker, e)

desired_keys = [
    'targetMeanPrice', 'targetMedianPrice', 'targetLowPrice', 'targetHighPrice',
    'numberOfAnalystOpinions', 'dividendYield', 'beta', 'overallRisk',
    'recommendationKey', 'sharesShort', 'sharesShortPriorMonth', 'sharesOutstanding',
    'earningsGrowth', 'revenueGrowth', 'debtToEquity', 'returnOnAssets', 'returnOnEquity',
    'priceToBook', 'trailingEps', 'forwardEps', 'industryKey', 'sectorKey',
    'country', 'priceToSalesTrailing12Months', 'enterpriseValue', 'enterpriseToRevenue',
    'priceEpsCurrentYear', 'forwardPE', 'trailingPE', 'profitMargins', 'fullExchangeName', 
    'marketCap', 'bookValue', 'lastDividendValue'
]

def get_historical_data(tickers: list) -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Retrieve historical financial, cashflow, and balance sheet data for each ticker.
    If data shapes differ, each DataFrame is truncated to the minimum number of columns.
    """

    financial_data = {}
    
    for ticker in tickers:

        try:
            
            logger.info("Obtaining Data for ticker %s", ticker)

            stock = ticker_objs.get(ticker, yf.Ticker(ticker))

            financials = stock.financials

            cashflow = stock.cashflow

            balance_sheet = stock.balance_sheet

            rev_estimate = stock.revenue_estimate

            eps_estimate = stock.earnings_estimate

            min_shape = min(financials.shape[1], cashflow.shape[1], balance_sheet.shape[1])

            financials = financials.iloc[:, :min_shape]

            cashflow = cashflow.iloc[:, :min_shape]

            balance_sheet = balance_sheet.iloc[:, :min_shape]

            rev_estimate = rev_estimate.loc['+1y'][['low', 'avg', 'high', 'numberOfAnalysts']]

            eps_estimate = eps_estimate.loc['+1y'][['low', 'avg', 'high', 'numberOfAnalysts']]

            financial_data[ticker] = {
                'financials': financials,
                'cashflow': cashflow,
                'balance_sheet': balance_sheet,
                'rev_estimate': rev_estimate,
                'eps_estimate': eps_estimate
            }

        except Exception as e:

            logger.warning("Error retrieving historical data for %s: %s", ticker, e)

    return financial_data


financial_data = get_historical_data(tickers)

net_income_dict = {}

operating_cash_flow_dict = {}

total_assets_dict = {}

average_assets_dict = {}

previous_return_on_assets_dict = {}

return_on_assets_dict = {}

long_term_debt_dict = {}

previous_long_term_debt_dict = {}

current_assets_dict = {}

current_liabilities_dict = {}

previous_current_liabilities_dict = {}

current_ratio_dict = {}

previous_current_ratio_dict = {}

new_shares_issued_dict = {}

gross_margin_dict = {}

previous_gross_margin_dict = {}

insider_purchases_dict = {}

asset_turnover_dict = {}

prev_asset_turnover_dict = {}

low_rev_estimate_dict = {}

avg_rev_estimate_dict = {}

high_rev_estimate_dict = {}

num_analyst_rev_dict = {}

low_eps_estimate_dict = {}

avg_eps_estimate_dict = {}

high_eps_estimate_dict = {}

num_analyst_eps_dict = {}

asset_growth_rate_dict = {}

asset_growth_rate_vol_dict = {}

book_value_dict = {}

total_liabilities_dict = {}

total_liabilities_growth_dict = {}

total_liabilities_growth_vol_dict = {}

tax_rate_dict = {}

for ticker in tickers:

    try:

        data = financial_data.get(ticker, {})

        financials = data.get('financials')

        cashflow = data.get('cashflow')

        balance_sheet = data.get('balance_sheet')

        rev_estimate = data.get('rev_estimate')

        eps_estimate = data.get('eps_estimate')

        if financials is None or cashflow is None or balance_sheet is None:

            continue

        net_income_val = safe_first_value(financials, 'Net Income')

        operating_cash_flow_val = safe_first_value(cashflow, 'Operating Cash Flow')

        total_assets_val = safe_first_value(balance_sheet, 'Total Assets')

        if 'Total Assets' in balance_sheet.index:

            total_assets_series = balance_sheet.loc['Total Assets'].dropna()[::-1] 

            if len(total_assets_series) >= 3:

                prev_total_assets = total_assets_series.iloc[-2]

                prev_prev_total_assets = total_assets_series.iloc[-3]

                asset_growth_rates = total_assets_series.pct_change().dropna()

                asset_growth_rate = asset_growth_rates.mean()

                asset_growth_rate_vol = total_assets_series.std()

            else:

                prev_total_assets = np.nan

                prev_prev_total_assets = np.nan

                asset_growth_rate = np.nan

                asset_growth_rate_vol = np.nan
            
            if 'Research And Development' in financials.index:
            
                rnd = safe_first_value(financials, 'Research And Development', default=0)
                
                rnd_assets = rnd / total_assets_val if total_assets_val != 0 else np.nan
        
            else:
                
                rnd = np.nan
                
                rnd_assets = np.nan
            
            if 'Capital Expenditure' in cashflow.index:
            
                capital_expenditure = safe_first_value(cashflow, 'Capital Expenditure', default=0)
                
                capex_intensity = capital_expenditure / total_assets_val if total_assets_val != 0 else np.nan
            
            else:
                
                capital_expenditure = np.nan
                
                capex_intensity = np.nan
                
            if 'EBIT' in financials.index:
                
                ebit = safe_first_value(financials, 'EBIT', default=0)
                
                ebit_assets = ebit / total_assets_val if total_assets_val != 0 else np.nan
            
            else:
                
                ebit = np.nan
                
                ebit_assets = np.nan
                
        else:

            prev_total_assets = np.nan

            prev_prev_total_assets = np.nan

            asset_growth_rate = np.nan

            asset_growth_rate_vol = np.nan
            
            rnd = np.nan
            
            rnd_assets = np.nan
            
            capital_expenditure = np.nan
            
            capex_intensity = np.nan
            
            ebit = np.nan
            
            ebit_assets = np.nan

        if 'Net Income' in financials.index and financials.loc['Net Income'].size > 1:

            prev_net_income = financials.loc['Net Income'].iloc[1]

        else:

            prev_net_income = 0
        
        if 'Common Stock Equity' in balance_sheet.index and not balance_sheet.loc['Common Stock Equity'].empty:

            book_value = balance_sheet.loc['Common Stock Equity', balance_sheet.columns[0]]

        elif 'Stockholders Equity' in balance_sheet.index and not balance_sheet.loc['Stockholders Equity'].empty:

            book_value = balance_sheet.loc['Stockholders Equity', balance_sheet.columns[0]]

        else:

            book_value = np.nan
            
        if 'Working Capital' in balance_sheet.index:
            
            working_capital = safe_first_value(balance_sheet, 'Working Capital', default=0)
        
        else:
            
            working_capital = np.nan
        

        if 'Total Liabilities Net Minority Interest' in balance_sheet.index:

            tot_liab_series = balance_sheet.loc['Total Liabilities Net Minority Interest'].dropna()[::-1]

            if len(tot_liab_series) >= 1:

                tot_liab_c = tot_liab_series.iloc[-1]

            else:

                tot_liab_c = np.nan

            if len(tot_liab_series) > 1:

                tot_liabs_growth = tot_liab_series.pct_change().dropna()

                tot_liab_growth = tot_liabs_growth.mean()

                tot_liab_growth_vol = tot_liab_series.std()

            else:

                tot_liab_growth = np.nan

                tot_liab_growth_vol = np.nan

        else:

            tot_liab_c = np.nan

            tot_liab_growth = np.nan

            tot_liab_growth_vol = np.nan

        average_assets_val = (
            total_assets_val if pd.isna(prev_total_assets)
            else (total_assets_val + prev_total_assets) / 2
        )

        previous_average_assets_val = (
            prev_total_assets if pd.isna(prev_prev_total_assets)
            else (prev_total_assets + prev_prev_total_assets) / 2
        )

        return_on_assets_val = (
            net_income_val / average_assets_val if average_assets_val != 0 else np.nan
        )

        previous_return_on_assets_val = (
            prev_net_income / previous_average_assets_val
            if previous_average_assets_val != 0 else np.nan
        )

        if 'Long Term Debt' in balance_sheet.index:

            long_term_debt_val = safe_first_value(balance_sheet, 'Long Term Debt', default=0)

            try:

                prev_long_term_debt = balance_sheet.loc['Long Term Debt'].iloc[1]

            except Exception:

                prev_long_term_debt = np.nan

        else:

            long_term_debt_val = 0

            prev_long_term_debt = 0

        if 'Current Assets' in balance_sheet.index and 'Current Liabilities' in balance_sheet.index:

            current_assets_val = safe_first_value(balance_sheet, 'Current Assets', default=0)

            current_liabilities_val = safe_first_value(balance_sheet, 'Current Liabilities', default=0)

            try:

                prev_current_assets = balance_sheet.loc['Current Assets'].iloc[1]

            except Exception:

                prev_current_assets = np.nan

            try:

                prev_current_liabilities = balance_sheet.loc['Current Liabilities'].iloc[1]

            except Exception:

                prev_current_liabilities = np.nan

        else:

            current_assets_val = 0

            current_liabilities_val = 0

            prev_current_assets = np.nan

            prev_current_liabilities = np.nan

        current_ratio_val = (

            current_assets_val / current_liabilities_val

            if current_liabilities_val != 0 else np.nan

        )

        previous_current_ratio_val = (

            prev_current_assets / prev_current_liabilities

            if pd.notna(prev_current_assets)

               and pd.notna(prev_current_liabilities)

               and prev_current_liabilities != 0

            else np.nan
        )

        if 'Issuance Of Capital Stock' in cashflow.index:

            new_shares_issued_val = safe_first_value(cashflow, 'Issuance Of Capital Stock', default=0)

        else:

            new_shares_issued_val = 0

        if 'Gross Profit' in financials.index and 'Total Revenue' in financials.index:

            total_revenue = safe_first_value(financials, 'Total Revenue', default=0)

            gross_profit = safe_first_value(financials, 'Gross Profit', default=0)

            gross_margin_val = gross_profit / total_revenue if total_revenue != 0 else 0

            try:

                prev_total_revenue = financials.loc['Total Revenue'].iloc[1]

                prev_gross_profit = financials.loc['Gross Profit'].iloc[1]

                previous_gross_margin_val = (

                    prev_gross_profit / prev_total_revenue if prev_total_revenue != 0 else np.nan
                )

            except Exception:

                previous_gross_margin_val = np.nan

        else:

            gross_margin_val = 0

            previous_gross_margin_val = 0
            
        if 'Tax Rate For Calcs' in financials.index:

            tax_rate = safe_first_value(financials, 'Tax Rate For Calcs', default=0)

        else:

            tax_rate = 0.21

        if 'Total Revenue' in financials.index and 'Current Assets' in balance_sheet.index:

            total_revenue = safe_first_value(financials, 'Total Revenue', default=0)

            try:

                prev_total_revenue = financials.loc['Total Revenue'].iloc[1]

            except Exception:

                prev_total_revenue = np.nan

            asset_turnover = (

                total_revenue / average_assets_val

                if average_assets_val != 0 else np.nan

            )

            prev_asset_turnover = (

                prev_total_revenue / previous_average_assets_val

                if previous_average_assets_val != 0 else np.nan

            )

        else:

            asset_turnover = np.nan

            prev_asset_turnover = np.nan

        ticker_obj = ticker_objs.get(ticker, yf.Ticker(ticker))

        try:

            insider_trans = ticker_obj.insider_purchases

            if not insider_trans.empty and len(insider_trans) > 2:

                insider_purchase = insider_trans.iloc[2]['Shares']

            else:

                insider_purchase = np.nan

        except Exception as e:

            logger.warning("Failed to retrieve insider purchases for %s: %s", ticker, e)

            insider_purchase = np.nan

        rev_low_estimate = rev_estimate['low']

        rev_avg_estimate = rev_estimate['avg']

        rev_high_estimate = rev_estimate['high']

        rev_num_analysts = rev_estimate['numberOfAnalysts']

        eps_low_estimate = eps_estimate['low']

        eps_avg_estimate = eps_estimate['avg']

        eps_high_estimate = eps_estimate['high']

        eps_num_analysts = eps_estimate['numberOfAnalysts']

        net_income_dict[ticker] = net_income_val

        operating_cash_flow_dict[ticker] = operating_cash_flow_val

        total_assets_dict[ticker] = total_assets_val

        average_assets_dict[ticker] = average_assets_val

        previous_return_on_assets_dict[ticker] = previous_return_on_assets_val

        return_on_assets_dict[ticker] = return_on_assets_val

        long_term_debt_dict[ticker] = long_term_debt_val

        previous_long_term_debt_dict[ticker] = prev_long_term_debt

        current_assets_dict[ticker] = current_assets_val

        current_liabilities_dict[ticker] = current_liabilities_val

        previous_current_liabilities_dict[ticker] = prev_current_liabilities

        current_ratio_dict[ticker] = current_ratio_val

        previous_current_ratio_dict[ticker] = previous_current_ratio_val

        new_shares_issued_dict[ticker] = new_shares_issued_val

        gross_margin_dict[ticker] = gross_margin_val

        previous_gross_margin_dict[ticker] = previous_gross_margin_val

        insider_purchases_dict[ticker] = insider_purchase

        asset_turnover_dict[ticker] = asset_turnover

        prev_asset_turnover_dict[ticker] = prev_asset_turnover

        low_rev_estimate_dict[ticker] = rev_low_estimate

        avg_rev_estimate_dict[ticker] = rev_avg_estimate

        high_rev_estimate_dict[ticker] = rev_high_estimate

        num_analyst_rev_dict[ticker] = rev_num_analysts

        low_eps_estimate_dict[ticker] = eps_low_estimate

        avg_eps_estimate_dict[ticker] = eps_avg_estimate

        high_eps_estimate_dict[ticker] = eps_high_estimate

        num_analyst_eps_dict[ticker] = eps_num_analysts

        book_value_dict[ticker] = book_value

        total_liabilities_dict[ticker] = tot_liab_c

        total_liabilities_growth_dict[ticker] = tot_liab_growth

        asset_growth_rate_dict[ticker] = asset_growth_rate

        asset_growth_rate_vol_dict[ticker] = asset_growth_rate_vol

        total_liabilities_growth_vol_dict[ticker] = tot_liab_growth_vol

        tax_rate_dict[ticker] = tax_rate

    except Exception as e:

        logger.warning("Could not process financial metrics for %s: %s", ticker, e)

        continue

targets_df = (
    pd.DataFrame(targetsYF)
    .T[desired_keys]
    .fillna(0)
    .infer_objects()
)

targets_df.index.name = "Ticker"

targets_df.columns = targets_df.columns.astype(str)

targets_df = targets_df.reindex(sorted(targets_df.index))

def map_yfinance_industry(yf_industry: str) -> str:
    """
    Map a yfinance industry string to a high-level industry name.
    Defaults to "Diversified" if not recognized.
    """

    return IndustryMap.get(yf_industry, "Diversified")

def map_yfinance_sector(yf_sector: str) -> str:
    """
    Map a yfinance sector string to a high-level sector name.
    Defaults to "Other" if not recognized.
    """

    return SectorMap.get(yf_sector, "Other")

targets_df['industryKey'] = targets_df['industryKey'].apply(map_yfinance_industry)

targets_df['sectorKey'] = targets_df['sectorKey'].apply(map_yfinance_sector)

targets_df.rename(columns={'industryKey': 'Industry'}, inplace=True)

targets_df.rename(columns={'sectorKey': 'Sector'}, inplace=True)

targets_df['Net Income'] = targets_df.index.map(net_income_dict)

targets_df['Operating Cash Flow'] = targets_df.index.map(operating_cash_flow_dict)

targets_df['Total Assets'] = targets_df.index.map(total_assets_dict)

targets_df['Average Assets'] = targets_df.index.map(average_assets_dict)

targets_df['Return on Assets'] = targets_df.index.map(return_on_assets_dict)

targets_df['Previous Return on Assets'] = targets_df.index.map(previous_return_on_assets_dict)

targets_df['Long Term Debt'] = targets_df.index.map(long_term_debt_dict)

targets_df['Previous Long Term Debt'] = targets_df.index.map(previous_long_term_debt_dict)

targets_df['Current Assets'] = targets_df.index.map(current_assets_dict)

targets_df['Current Liabilities'] = targets_df.index.map(current_liabilities_dict)

targets_df['Previous Current Liabilities'] = targets_df.index.map(previous_current_liabilities_dict)

targets_df['Current Ratio'] = targets_df.index.map(current_ratio_dict)

targets_df['Previous Current Ratio'] = targets_df.index.map(previous_current_ratio_dict)

targets_df['New Shares Issued'] = targets_df.index.map(new_shares_issued_dict)

targets_df['Gross Margin'] = targets_df.index.map(gross_margin_dict)

targets_df['Previous Gross Margin'] = targets_df.index.map(previous_gross_margin_dict)

targets_df['Insider Purchases'] = targets_df.index.map(insider_purchases_dict)

targets_df['Asset Turnover'] = targets_df.index.map(asset_turnover_dict)

targets_df['Previous Asset Turnover'] = targets_df.index.map(prev_asset_turnover_dict)

targets_df['Low Revenue Estimate'] = targets_df.index.map(low_rev_estimate_dict)

targets_df['Avg Revenue Estimate'] = targets_df.index.map(avg_rev_estimate_dict)

targets_df['High Revenue Estimate'] = targets_df.index.map(high_rev_estimate_dict)

targets_df['Number Analyst Revenue Estimate'] = targets_df.index.map(num_analyst_rev_dict)

targets_df['Low EPS Estimate'] = targets_df.index.map(low_eps_estimate_dict)

targets_df['Avg EPS Estimate'] = targets_df.index.map(avg_eps_estimate_dict)

targets_df['High EPS Estimate'] = targets_df.index.map(high_eps_estimate_dict)

targets_df['Number Analyst EPS Estimate'] = targets_df.index.map(num_analyst_eps_dict)

targets_df['Asset Growth Rate'] = targets_df.index.map(asset_growth_rate_dict)

targets_df['Total Liabilities'] = targets_df.index.map(total_liabilities_dict)

targets_df['Total Liabilities Growth Rate'] = targets_df.index.map(total_liabilities_growth_dict)

targets_df['Book Value'] = targets_df.index.map(book_value_dict)

targets_df['Tax Rate'] = targets_df.index.map(tax_rate_dict)

logger.info("Download Complete")

for ticker in tickers:

    if ticker.endswith('.L'):
        
        targets_df.loc[ticker, 'targetMeanPrice'] /= 100
        targets_df.loc[ticker, 'targetMedianPrice'] /= 100
        targets_df.loc[ticker, 'targetLowPrice'] /= 100
        targets_df.loc[ticker, 'targetHighPrice'] /= 100
        targets_df.loc[ticker, 'forwardEps'] /= 100
        targets_df.loc[ticker, 'priceEpsCurrentYear'] /= 100            
        targets_df.loc[ticker, 'priceToBook'] /= 100
        
        targets_df.loc[ticker, 'marketCap'] /= gbpusd
        
    if ticker.endswith('.TO'):
        
        targets_df.loc[ticker, 'marketCap'] /= usdcad
        targets_df.loc[ticker, 'enterpriseValue'] /= usdcad
        
    if ticker.endswith('.MC') or ticker.endswith('.PA') or ticker.endswith('.AS'):
        
        targets_df.loc[ticker, 'marketCap'] *= eurusd
        targets_df.loc[ticker, 'enterpriseValue'] *= eurusd
    
    if ticker.endswith('.SW'):
        
        targets_df.loc[ticker, 'marketCap'] *= usdchf
        targets_df.loc[ticker, 'enterpriseValue'] *= usdchf
    
    if ticker.endswith('.HK'):
        
        targets_df.loc[ticker, 'marketCap'] /= usdhkd
        targets_df.loc[ticker, 'enterpriseValue'] /= usdhkd
    
    if ticker.endswith('.CO'):
        
        targets_df.loc[ticker, 'marketCap'] /= usddkk
        targets_df.loc[ticker, 'enterpriseValue'] /= usddkk
                
    if ticker in ['ASX']:
        
        usdtwd = 30
        targets_df.loc[ticker, 'Low Revenue Estimate'] /= usdtwd
        targets_df.loc[ticker, 'Avg Revenue Estimate'] /= usdtwd
        targets_df.loc[ticker, 'High Revenue Estimate'] /= usdtwd
        targets_df.loc[ticker, 'enterpriseValue'] /= usdtwd
        targets_df.loc[ticker, 'bookValue'] /= usdtwd
        targets_df.loc[ticker, 'Net Income'] /= usdtwd
        targets_df.loc[ticker, 'Operating Cash Flow'] /= usdtwd
        targets_df.loc[ticker, 'Total Assets'] /= usdtwd
        targets_df.loc[ticker, 'Average Assets'] /= usdtwd
        targets_df.loc[ticker, 'Long Term Debt'] /= usdtwd
        targets_df.loc[ticker, 'Previous Long Term Debt'] /= usdtwd
        targets_df.loc[ticker, 'Current Assets'] /= usdtwd
        targets_df.loc[ticker, 'Current Liabilities'] /= usdtwd
        targets_df.loc[ticker, 'Previous Current Liabilities'] /= usdtwd
        targets_df.loc[ticker, 'Previous Long Term Debt'] /= usdtwd
        targets_df.loc[ticker, 'Total Liabilities'] /= usdtwd
        targets_df.loc[ticker, 'Book Value'] /= usdtwd
        
    if ticker in ['BABA']:
        
        targets_df.loc[ticker, 'Low Revenue Estimate'] /= usdcny
        targets_df.loc[ticker, 'Avg Revenue Estimate'] /= usdcny
        targets_df.loc[ticker, 'High Revenue Estimate'] /= usdcny
        targets_df.loc[ticker, 'enterpriseValue'] /= usdcny
        targets_df.loc[ticker, 'bookValue'] /= usdcny
        targets_df.loc[ticker, 'Net Income'] /= usdcny
        targets_df.loc[ticker, 'Operating Cash Flow'] /= usdcny
        targets_df.loc[ticker, 'Total Assets'] /= usdcny
        targets_df.loc[ticker, 'Average Assets'] /= usdcny
        targets_df.loc[ticker, 'Long Term Debt'] /= usdcny
        targets_df.loc[ticker, 'Previous Long Term Debt'] /= usdcny
        targets_df.loc[ticker, 'Current Assets'] /= usdcny
        targets_df.loc[ticker, 'Current Liabilities'] /= usdcny
        targets_df.loc[ticker, 'Previous Current Liabilities'] /= usdcny
        targets_df.loc[ticker, 'Previous Long Term Debt'] /= usdcny
        targets_df.loc[ticker, 'Total Liabilities'] /= usdcny
        targets_df.loc[ticker, 'Book Value'] /= usdcny
        targets_df.loc[ticker, 'priceEpsCurrentYear'] /= usdcny
        targets_df.loc[ticker, 'Low EPS Estimate'] /= usdcny
        targets_df.loc[ticker, 'Avg EPS Estimate'] /= usdcny
        targets_df.loc[ticker, 'High EPS Estimate'] /= usdcny
        

def compute_z_score(n: int) -> float:
    """
    Compute z-score given a count of opinions (number of analysts).
    """
    if n == 2:
        n += 1
    alpha = 1 / n
    return st.norm.ppf(1 - alpha)


def rets_variable_yahoo(meanY: float, medianY: float, minY: float, maxY: float,
                        nY: float, price: float, beta: float, div: float = 0,
                        threshold: float = 0.10) -> Tuple[float, float, float, float]:
    """
    Compute expected return, volatility, and related statistics based on analyst data.
    """
   
    if nY < 2:
   
        nY = 2
   
    if price == 0:
   
        logger.warning("Price is 0; cannot compute returns.")
   
        return 0, 0, 0, beta

    if meanY > 0 and abs(meanY - medianY) / meanY > threshold:

        medRetsY = (medianY / price) - 1

        minRetsY = (minY / price) - 1

        maxRetsY = (maxY / price) - 1

        varY = (((medRetsY - minRetsY) ** 2) + ((maxRetsY - medRetsY) ** 2)) / 12

        expected_return = (minRetsY + 2 * medRetsY + maxRetsY) / 4

        sigma_return = np.sqrt(varY)

    else:

        expected_return = ((meanY / price)) - 1

        z_score = compute_z_score(int(nY))

        sigma_return = (maxY - minY) / (2 * z_score * price)

    return expected_return, sigma_return, sigma_return, beta

latest_prices_series = pd.Series({ticker: close[ticker].iloc[-1] for ticker in tickers})

latest_prices_series = latest_prices_series.reindex(targets_df.index)

latest_prices = latest_prices_series.to_dict()

targets_df['Current Price'] = targets_df.index.map(latest_prices)

logger.info("Computing Analyst Predictions ...")

analyst_results: Dict[str, Tuple[float, float, float, float]] = {}

for ticker in tickers:

    if ticker in targets_df.index:

        data = targets_df.loc[ticker]

        price = latest_prices[ticker]

        meanY = data['targetMeanPrice']

        medianY = data['targetMedianPrice']

        minY = data['targetLowPrice']

        maxY = data['targetHighPrice']

        nY = data['numberOfAnalystOpinions']

        div = data['dividendYield']

        beta = data['beta']

        if price == 0:

            logger.warning("Price for %s is 0; skipping.", ticker)

            continue

        analyst_results[ticker] = rets_variable_yahoo(
            meanY, medianY, minY, maxY,
            nY, price, beta, div
        )

one_year_from_today = today + dt.timedelta(days=365)

Analyst_Target_Data = []

for ticker, result in analyst_results.items():

    Analyst_Target_Data.append({
        "Ticker": ticker,
        "Current Price": latest_prices[ticker],
        f"Target Price ({one_year_from_today})": latest_prices[ticker] * (1 + result[0]),
        "Predicted Returns": result[0],
        "Predicted Volatility": result[1],
        "Standard Error": result[2],
        "Beta": result[3]
    })

Analyst_Target_df = pd.DataFrame(Analyst_Target_Data)

Analyst_Target_df = Analyst_Target_df.set_index("Ticker")

Analyst_Target_df = Analyst_Target_df.reindex(targets_df.index)

Analyst_Target_df = Analyst_Target_df.reset_index()

Analyst_Target_df = pd.DataFrame(Analyst_Target_Data)

Analyst_Target_df.columns = Analyst_Target_df.columns.astype(str)

logger.info("Uploading Data to Excel ...")

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')


def format_sheet_as_table(excel_file: str, sheet_name: str) -> None:
    """
    Convert the specified sheet into an Excel table with a preset style.
    """

    try:

        wb = load_workbook(excel_file)

    except Exception as exc:

        logger.error(
            "Error opening Excel file '%s': %s. Ensure the file is not open in another application.",
            excel_file, exc
        )

        return

    if sheet_name not in wb.sheetnames:

        logger.warning("Sheet '%s' not found in %s", sheet_name, excel_file)

        wb.close()

        return

    ws = wb[sheet_name]

    ws._tables.clear()

    max_row = ws.max_row

    max_col = ws.max_column

    last_col_letter = get_column_letter(max_col)

    table_range = f"A1:{last_col_letter}{max_row}"

    table_name = sheet_name.replace(" ", "") + "Table"

    table = Table(displayName=table_name, ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
  
    table.tableStyleInfo = style
  
    ws.add_table(table)

    wb.save(excel_file)
 
    wb.close()
  
    logger.info("Formatted sheet '%s' as a table.", sheet_name)

with pd.ExcelWriter(output_excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
  
    targets_df.to_excel(writer, sheet_name='Analyst Data', index=True)
  
    Analyst_Target_df.to_excel(writer, sheet_name='Analyst Target', index=False)

    ws1 = writer.sheets['Analyst Data']
 
    max_row_ws1 = ws1.max_row

    current_price_col = None
  
    for cell in ws1[1]:
   
        if cell.value == "Current Price":
  
            current_price_col = cell.column_letter
  
            break

    if current_price_col is None:
       
        logger.error("Current Price column not found in 'Analyst Data'. Using default 'I'.")
       
        current_price_col = "I"

    for col in ['B', 'C', 'D', 'E']:
       
        formula_red = f"{col}2<{current_price_col}2"
        formula_green = f"{col}2>{current_price_col}2"

        ws1.conditional_formatting.add(
            f"{col}2:{col}{max_row_ws1}",
            FormulaRule(formula=[formula_red], fill=red_fill)
        )

        ws1.conditional_formatting.add(
            f"{col}2:{col}{max_row_ws1}",
            FormulaRule(formula=[formula_green], fill=green_fill)
        )

    ws2 = writer.sheets['Analyst Target']
    
    max_row_ws2 = ws2.max_row

    ws2.conditional_formatting.add(
        f"C2:C{max_row_ws2}",
        FormulaRule(formula=["C2<B2"], fill=red_fill)
    )

    ws2.conditional_formatting.add(
        f"C2:C{max_row_ws2}",
        FormulaRule(formula=["C2>B2"], fill=green_fill)
    )

    ws2.conditional_formatting.add(
        f"D2:D{max_row_ws2}",
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
    )

    ws2.conditional_formatting.add(
        f"D2:D{max_row_ws2}",
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
    )

format_sheet_as_table(output_excel_file, 'Analyst Data')

format_sheet_as_table(output_excel_file, 'Analyst Target')

logger.info("Data has been uploaded to Excel with conditional formatting applied, and all sheets are now tables.")
